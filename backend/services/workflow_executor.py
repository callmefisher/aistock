import pandas as pd
import numpy as np
import os
import glob
import json
from typing import Dict, List, Optional, Any, Union
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path

from services.path_resolver import get_resolver
from utils.stock_code_normalizer import (
    normalize_stock_code,
    extract_numeric_code,
    match_stock_code_flexible,
    is_public_file as check_is_public_file
)

logger = logging.getLogger(__name__)

# 模块级缓存：只读匹配源数据 + 公共文件
# key: (dir_path, max_mtime) → value: cached data
_match_source_cache: Dict[str, tuple] = {}  # dir_path → (max_mtime, stock_dict)
_public_file_cache: Dict[str, tuple] = {}   # file_path → (mtime, DataFrame/dict)
_PUBLIC_FILE_CACHE_MAX = 50  # 最大缓存条目数，超出时淘汰最早的条目


def _get_dir_mtime(dir_path: str) -> float:
    """获取目录下所有 Excel 文件的最大修改时间"""
    max_mtime = 0.0
    for ext in ("*.xlsx", "*.xls"):
        for f in glob.glob(os.path.join(dir_path, ext)):
            mt = os.path.getmtime(f)
            if mt > max_mtime:
                max_mtime = mt
    return max_mtime


def invalidate_public_cache(dir_path: str = None):
    """清除公共文件缓存，上传/删除文件后调用"""
    if dir_path:
        keys_to_remove = [k for k in _public_file_cache if k.startswith(dir_path)]
        for k in keys_to_remove:
            del _public_file_cache[k]
        logger.info(f"已清除公共文件缓存: {dir_path} ({len(keys_to_remove)}个)")
    else:
        _public_file_cache.clear()
        logger.info("已清除所有公共文件缓存")


def invalidate_match_source_cache(dir_path: str = None):
    """清除匹配源缓存，上传/删除文件后调用"""
    if dir_path:
        keys_to_remove = [k for k in _match_source_cache if k.startswith(dir_path)]
        for k in keys_to_remove:
            del _match_source_cache[k]
        logger.info(f"已清除匹配源缓存: {dir_path} ({len(keys_to_remove)}个)")
    else:
        _match_source_cache.clear()
        logger.info("已清除所有匹配源缓存")


def auto_adjust_excel_width(output_path: str, fixed_width: int = 20, all_sheets: bool = False, center_align: bool = False):
    """设置固定列宽（不遍历单元格，极快）。可选居中对齐。"""
    try:
        wb = load_workbook(output_path)
        sheets = wb.worksheets if all_sheets else [wb.active]
        for ws in sheets:
            ws.auto_filter.ref = ws.dimensions
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = fixed_width
            if center_align:
                from openpyxl.styles import Alignment
                _center = Alignment(horizontal="center", vertical="center", wrap_text=False)
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = _center
        wb.save(output_path)
    except Exception as e:
        logger.warning(f"设置列宽失败: {output_path}, {e}")


def apply_pledge_ratio_coloring(wb):
    """质押比例相邻两列红绿对比：右>左→深红，右<左→浅绿，相等/空→不标。
    直接在传入的 openpyxl Workbook 上修改，不保存；由调用方负责 save。
    """
    from openpyxl.styles import PatternFill
    # 浅红：质押比例变大（比深红 FFC00000 更浅，黑字可读）
    red_fill = PatternFill(start_color="FFFFA7A7", end_color="FFFFA7A7", fill_type="solid")
    green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")

    def _to_float(val):
        if val is None:
            return None
        try:
            s = str(val).strip().replace("%", "")
            if s == "" or s.lower() == "nan":
                return None
            return float(s)
        except (ValueError, TypeError):
            return None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        ratio_col_idx = [i + 1 for i, h in enumerate(header) if str(h or "").startswith("质押比例")]
        if len(ratio_col_idx) < 2:
            continue
        for row in range(2, ws.max_row + 1):
            for k in range(1, len(ratio_col_idx)):
                right = ws.cell(row, ratio_col_idx[k])
                left = ws.cell(row, ratio_col_idx[k - 1])
                a = _to_float(right.value)
                b = _to_float(left.value)
                if a is None or b is None:
                    continue
                if a > b:
                    right.fill = red_fill
                elif a < b:
                    right.fill = green_fill


def clean_df_for_json(df: pd.DataFrame) -> List[Dict]:
    df_clean = df.fillna('')
    records = df_clean.head(100).to_dict('records')
    for record in records:
        for k, v in record.items():
            if isinstance(v, (float, np.floating)) and (np.isnan(v) or np.isinf(v)):
                record[k] = ''
            elif isinstance(v, (float, int)) and (v != v or abs(v) == float('inf')):
                record[k] = ''
    return records


class WorkflowExecutor:
    def __init__(self, base_dir: str = None, workflow_type: str = ""):
        if base_dir is None:
            import platform
            system = platform.system()
            if system == "Darwin" or os.path.exists("/Users/xiayanji"):
                base_dir = "/Users/xiayanji/qbox/aistock/data/excel"
            else:
                base_dir = "/app/data/excel"
        self.base_dir = base_dir
        self.workflow_type = workflow_type
        self.resolver = get_resolver(base_dir, workflow_type)
        self.today = datetime.now().strftime("%Y-%m-%d")
        os.makedirs(self.base_dir, exist_ok=True)

    def _load_match_source(self, source_path: str, code_col_candidates: List[str], name_col_candidates: List[str]) -> Dict[str, str]:
        """加载匹配源数据，带 mtime 缓存。支持逐行多列回退取值。"""
        current_mtime = _get_dir_mtime(source_path)
        cached = _match_source_cache.get(source_path)
        if cached and cached[0] == current_mtime:
            logger.info(f"匹配源缓存命中: {source_path} ({len(cached[1])}条)")
            return cached[1]

        stock_dict = {}
        excel_files = self._get_excel_files_in_dir(source_path)
        for excel_file in excel_files:
            try:
                src_df = pd.read_excel(excel_file, dtype=str)
                code_cols = [c for c in code_col_candidates if c in src_df.columns]
                name_col = next((c for c in name_col_candidates if c in src_df.columns), None)
                if not code_cols or not name_col:
                    continue

                # 向量化 coalesce：按优先级从多个代码列中取第一个非空值
                combined_code = pd.Series('', index=src_df.index)
                for cc in reversed(code_cols):
                    vals = src_df[cc].fillna('').astype(str).str.strip()
                    mask = vals != ''
                    combined_code[mask] = vals[mask]

                names = src_df[name_col].fillna('').astype(str)
                for code_val, name_val in zip(combined_code, names):
                    nc = normalize_stock_code(code_val)
                    if nc:
                        stock_dict[nc] = normalize_stock_code(name_val)
            except Exception as e:
                logger.warning(f"读取{excel_file}失败: {e}")

        # 预建反向索引：同时存 原始key 和 纯数字key，让匹配 O(1)
        expanded = {}
        for key, val in stock_dict.items():
            expanded[key] = val
            numeric = extract_numeric_code(key)
            if numeric and numeric != key:
                expanded[numeric] = val

        _match_source_cache[source_path] = (current_mtime, expanded)
        logger.info(f"匹配源加载并缓存: {source_path} ({len(stock_dict)}条, 索引{len(expanded)}条)")
        return expanded

    def _read_public_file_cached(self, filepath: str, skiprows: int = 1) -> pd.DataFrame:
        """读取公共文件，带 mtime 缓存"""
        current_mtime = os.path.getmtime(filepath)
        cached = _public_file_cache.get(filepath)
        if cached and cached[0] == current_mtime:
            logger.info(f"公共文件缓存命中: {os.path.basename(filepath)}")
            return cached[1].copy()

        df = pd.read_excel(filepath, skiprows=skiprows)
        _public_file_cache[filepath] = (current_mtime, df)
        logger.info(f"公共文件加载并缓存: {os.path.basename(filepath)} ({len(df)}行)")
        return df.copy()

    def _get_daily_dir(self, date_str: Optional[str] = None) -> str:
        return self.resolver.get_daily_dir(date_str)

    def _resolve_path(self, file_path: str, date_str: Optional[str] = None) -> str:
        if os.path.isabs(file_path):
            return file_path
        daily_dir = self._get_daily_dir(date_str)
        return os.path.join(daily_dir, file_path)

    def _get_excel_files_in_dir(self, daily_dir: str) -> List[str]:
        pattern = os.path.join(daily_dir, "*.xlsx")
        files = glob.glob(pattern)
        pattern = os.path.join(daily_dir, "*.xls")
        files.extend(glob.glob(pattern))
        return sorted(files)

    def _derive_pledge_source(self, file_name: str, sheet_name: str) -> str:
        """质押类型来源识别：文件名含关键字优先（中大盘 / 小盘），sheet 名含关键字兜底。

        真实命名形如 '质押和大宗交易 中大盘0421.xlsx' / '质押和大宗交易 小盘 0421.xlsx'，
        关键字可能在字符串中部。必须先判"中大盘"（"小盘"是"中大盘"的子串，若先判小盘会误归）。
        """
        fn = str(file_name or "").strip()
        if "中大盘" in fn:
            return "中大盘"
        if "小盘" in fn:
            return "小盘"
        sn = str(sheet_name or "").strip()
        if "中大盘" in sn:
            return "中大盘"
        if "小盘" in sn:
            return "小盘"
        return "小盘"

    def _maybe_flatten_pledge_multiheader(
        self,
        df: pd.DataFrame,
        filepath: str,
        sheet_name: str,
    ) -> pd.DataFrame:
        """质押 sheet 的列名标准化：

        情况 A（单行表头含换行）：列名是"质押比例\n[截止日期]2025-04-01"
          → 提取日期，改为"质押比例2025-04-01"

        情况 B（双行表头）：列名只有"质押比例"（pandas .1/.2 后缀或 Unnamed 兄弟列），
          用 openpyxl 读前两行原始单元格拼接。

        只改"质押比例"前缀列；其他列原样保留。
        """
        import re
        _date_re = re.compile(r"(\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2})")

        try:
            new_cols = list(df.columns)
            changed = False
            # 情况 A：单行表头含 \n 或 含日期
            for i, c in enumerate(new_cols):
                if not isinstance(c, str):
                    continue
                # 以"质押比例"开头（忽略空白 / 换行）
                stripped = c.lstrip()
                if stripped.startswith("质押比例"):
                    m = _date_re.search(c)
                    if m:
                        new_cols[i] = f"质押比例{m.group(1)}"
                        changed = True
                    elif "最新" in c and c != "质押比例":
                        new_cols[i] = "质押比例最新"
                        changed = True

            if changed:
                df = df.copy()
                df.columns = new_cols
                ratio_count = sum(1 for c in new_cols if isinstance(c, str) and c.startswith("质押比例"))
                logger.info(f"质押：列名规整 {filepath}#{sheet_name}, 质押比例列数={ratio_count}")
                return df

            # 情况 B：双行表头（需要 openpyxl 读前两行拼接）
            ratio_like = [
                c for c in df.columns
                if isinstance(c, str) and c.split(".")[0].strip() == "质押比例"
            ]
            if len(ratio_like) < 2:
                return df
            from openpyxl import load_workbook as _lwb
            wb = _lwb(filepath, read_only=True, data_only=True)
            try:
                if sheet_name not in wb.sheetnames:
                    return df
                ws = wb[sheet_name]
                row1 = []
                row2 = []
                for i, row in enumerate(ws.iter_rows(max_row=2, values_only=True)):
                    if i == 0:
                        row1 = list(row)
                    elif i == 1:
                        row2 = list(row)
                    else:
                        break
            finally:
                wb.close()
            if not row1 or not row2:
                return df
            ncols = len(df.columns)
            def _fill_forward(arr):
                out = []
                last = ""
                for v in arr:
                    s = str(v).strip() if v is not None else ""
                    if s:
                        last = s
                    out.append(last)
                return out
            row1_ff = _fill_forward(row1)[:ncols]
            row2_ff = _fill_forward(row2)[:ncols]
            while len(row1_ff) < ncols:
                row1_ff.append("")
            while len(row2_ff) < ncols:
                row2_ff.append("")
            new_cols2 = []
            for i, col in enumerate(df.columns):
                r1 = row1_ff[i]
                r2 = row2_ff[i]
                if r1 == "质押比例":
                    m = _date_re.search(r2 or "")
                    if m:
                        new_cols2.append(f"质押比例{m.group(1)}")
                        continue
                new_cols2.append(col)
            if new_cols2 != list(df.columns):
                df = df.copy()
                df.columns = new_cols2
                logger.info(f"质押：双行表头展平 {filepath}#{sheet_name}")
            return df
        except Exception as e:
            logger.warning(f"质押：列名规整失败，继续使用原列名 {filepath}#{sheet_name}: {e}")
            return df

    def _sync_pledge_final_to_public(self, final_file_path: str, date_str: Optional[str] = None) -> bool:
        """把最终输出文件同步到 质押/public 目录。

        顺序：先复制成功 → 删除 public 中除了新复制文件外的其他文件（原子语义的兜底）。
        仅在 workflow_type=='质押' 时生效；失败不抛，仅 warning（不阻塞主流程）。
        """
        if self.workflow_type != "质押":
            return False
        if not final_file_path or not os.path.exists(final_file_path):
            logger.warning(f"[质押 public 同步] 源文件不存在，跳过: {final_file_path}")
            return False
        try:
            import shutil
            public_dir = self.resolver.get_public_directory(date_str)
            os.makedirs(public_dir, exist_ok=True)
            src_name = os.path.basename(final_file_path)
            dst_path = os.path.join(public_dir, src_name)
            # 先复制到目标（若目标已存在会覆盖）
            shutil.copy2(final_file_path, dst_path)
            logger.info(f"[质押 public 同步] 已复制 {src_name} → {public_dir}")
            # 复制成功后清理 public 中其他文件
            removed = 0
            for entry in os.listdir(public_dir):
                entry_path = os.path.join(public_dir, entry)
                if os.path.abspath(entry_path) == os.path.abspath(dst_path):
                    continue
                if os.path.isfile(entry_path):
                    try:
                        os.remove(entry_path)
                        removed += 1
                    except Exception as e:
                        logger.warning(f"[质押 public 同步] 删除 {entry_path} 失败: {e}")
            logger.info(f"[质押 public 同步] 清理 public 中其他文件 {removed} 个，保留: {src_name}")
            return True
        except Exception as e:
            logger.warning(f"[质押 public 同步] 失败（不影响主流程）: {e}")
            return False

    _PLEDGE_FIXED_PREFIX_COLS = (
        "证券代码", "证券简称", "最新公告日",
        "百日新高", "站上20日线", "国央企", "所属板块",
    )

    # 源表里这 4 类信息列（含同义词变体）应被丢弃，由 match_* 步骤权威填充
    _PLEDGE_INFO_COL_SYNONYMS = {
        "百日新高": ("百日新高", "百日最高价", "百日最高"),
        "站上20日线": ("站上20日线", "20日线", "20日均线", "20日均价"),
        "国央企": ("国央企", "国企", "国有企业"),
        "所属板块": ("所属板块", "一级板块", "板块"),
    }

    def _reorder_pledge_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """按固定前 7 列 + 原始剩余列（源序、去重）返回新 DataFrame。
        丢弃 来源 / 序号 列；缺失的前 7 列补空列；重复列名自动去重。
        源表里"百日新高/站上20日线/国央企/所属板块"的原始值及同义词变体一律丢弃，
        由后续 match_* 步骤权威填充；match 未跑时留空列。
        """
        df = df.copy()
        if df.columns.duplicated().any():
            # 重复列追加 _1, _2 后缀
            seen = {}
            new_cols = []
            for c in df.columns:
                if c in seen:
                    seen[c] += 1
                    new_cols.append(f"{c}_{seen[c]}")
                else:
                    seen[c] = 0
                    new_cols.append(c)
            df.columns = new_cols

        # 丢弃源表的 4 类信息列的**同义词变体**——但保留权威名（match_* 步骤的产出列）
        # 即：drop "20日均线"/"国企" 等变体，但保留 "站上20日线"/"国央企"（由 match_* 填充）
        info_synonyms_set = set()
        for canonical, variants in self._PLEDGE_INFO_COL_SYNONYMS.items():
            for v in variants:
                if v == canonical:
                    continue  # 权威名不丢
                info_synonyms_set.add(v)
        info_cols_present = [c for c in df.columns if c in info_synonyms_set]
        if info_cols_present:
            df = df.drop(columns=info_cols_present)

        for col in self._PLEDGE_FIXED_PREFIX_COLS:
            if col not in df.columns:
                df[col] = ""
        drop_cols = {"来源", "序号"}
        prefix = list(self._PLEDGE_FIXED_PREFIX_COLS)
        rest = [c for c in df.columns if c not in prefix and c not in drop_cols]
        return df[prefix + rest]

    def _load_pledge_baseline(self, public_dir: str) -> Dict[str, Any]:
        """读 public 目录所有 xlsx，返回**按来源分的**最大日期基准（运行时 finalize 使用）。

        返回结构：{"中大盘": Timestamp or None, "小盘": Timestamp or None}
        规则：public xlsx 中 sheet 名含"中大盘" → 贡献"中大盘"基准；含"小盘" → 贡献"小盘"基准。
        新文件某行 `最新公告日 > 对应来源的 baseline` → 红标；baseline 为空 → 不标。
        失败时对应 key 返回 None。

        注意：stock_pools 表**不**在此读取；它只用于下载导出时的独立红标（见 statistics_api）。
        """
        baseline: Dict[str, Any] = {"中大盘": None, "小盘": None}

        def _update(source_key: str, ts) -> None:
            if source_key not in baseline:
                return
            cur = baseline[source_key]
            if cur is None or ts > cur:
                baseline[source_key] = ts

        # 读 public 目录下所有 xlsx
        try:
            if public_dir and os.path.isdir(public_dir):
                from openpyxl import load_workbook as _lwb
                for fname in os.listdir(public_dir):
                    if not fname.lower().endswith(".xlsx"):
                        continue
                    fpath = os.path.join(public_dir, fname)
                    try:
                        wb = _lwb(fpath, read_only=True, data_only=True)
                        try:
                            for sheet_name in wb.sheetnames:
                                # sheet 名含关键字决定来源（"中大盘" 必须先判，避免被 "小盘" 子串误归）
                                sn = str(sheet_name or "").strip()
                                if "中大盘" in sn:
                                    src_key = "中大盘"
                                elif "小盘" in sn:
                                    src_key = "小盘"
                                else:
                                    # 未知 sheet 名 → 两边都不记入，跳过
                                    continue
                                ws = wb[sheet_name]
                                rows = ws.iter_rows(values_only=True)
                                try:
                                    header = next(rows)
                                except StopIteration:
                                    continue
                                header = [str(h) if h is not None else "" for h in header]
                                date_idxs = [
                                    i for i, h in enumerate(header)
                                    if h == "最新公告日" or "股权质押公告日期" in h
                                ]
                                if not date_idxs:
                                    continue
                                for row in rows:
                                    if not row:
                                        continue
                                    for di in date_idxs:
                                        if di >= len(row):
                                            continue
                                        dv = row[di]
                                        if dv is None or dv == "":
                                            continue
                                        try:
                                            ts = pd.to_datetime(dv, errors="coerce")
                                            if pd.isna(ts):
                                                continue
                                        except Exception:
                                            continue
                                        _update(src_key, ts)
                        finally:
                            wb.close()
                    except Exception as e:
                        logger.warning(f"[质押 baseline] 读取 public 文件失败 {fname}: {e}")
                        continue
        except Exception as e:
            logger.warning(f"[质押 baseline] 扫描 public 目录失败: {e}")

        logger.info(f"[质押 baseline] 中大盘={baseline['中大盘']}, 小盘={baseline['小盘']}")
        return baseline

    def finalize_pledge_if_needed(
        self,
        last_output_path: Optional[str],
        date_str: Optional[str] = None,
    ) -> bool:
        """质押工作流 run_workflow 循环末尾调用一次。

        - 非质押类型 → return False
        - 读 last_output_path 为 DataFrame → finalize → 覆盖 last_output_path 并同步 public
        - 失败仅 warning，不抛
        """
        if self.workflow_type != "质押":
            return False
        if not last_output_path or not os.path.exists(last_output_path):
            logger.warning(f"[质押 finalize] last output 不存在，跳过: {last_output_path}")
            return False
        try:
            # 读所有 sheet 合并，兜底历史 match_sector 可能已分 sheet 的情况
            try:
                sheet_map = pd.read_excel(last_output_path, sheet_name=None)
                if isinstance(sheet_map, dict) and sheet_map:
                    df = pd.concat(list(sheet_map.values()), ignore_index=True)
                else:
                    df = pd.read_excel(last_output_path)
            except Exception:
                df = pd.read_excel(last_output_path)
            public_dir = self.resolver.get_public_directory(date_str)
            # 关键：finalize 输出路径必须与 last_output_path 一致，
            # 这样 run_workflow 的 DB 保存和前端下载读到的是 finalize 后的文件
            output_path = last_output_path
            self._finalize_pledge_output(df, date_str, output_path, public_dir)
            self._sync_pledge_final_to_public(output_path, date_str)
            return True
        except Exception as e:
            logger.error(f"[质押 finalize] 失败: {e}")
            return False

    def _finalize_pledge_output(
        self,
        df: pd.DataFrame,
        date_str: str,
        output_path: str,
        public_dir: str,
    ) -> None:
        """质押最终输出：列重排 + 分 sheet + 质押比例红绿对比 + 最新公告日首次出现绿标。

        Sheet1 = 中大盘{date_str}, Sheet2 = 小盘{date_str}；两者固定存在（空则仅表头）。
        Baseline（从 public_dir xlsx）用于判定最新公告日是否首次出现；stock_pools 表
        仅在下载端（statistics_api）参与。所有样式（质押比例红绿 / 最新公告日红）在此
        一次性施加，避免被 openpyxl 中间步骤写入清除。
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        baseline = self._load_pledge_baseline(public_dir)

        if df is None:
            df = pd.DataFrame()

        src_col = df["来源"] if "来源" in df.columns else pd.Series(["小盘"] * len(df), index=df.index)
        df_big = df[src_col == "中大盘"].copy()
        df_small = df[src_col != "中大盘"].copy()

        df_big = self._reorder_pledge_columns(df_big)
        df_small = self._reorder_pledge_columns(df_small)

        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, sub_df in (
            (f"中大盘{date_str}", df_big),
            (f"小盘{date_str}", df_small),
        ):
            ws = wb.create_sheet(sheet_name)
            header = ["序号"] + list(sub_df.columns)
            ws.append(header)
            for i, (_, row) in enumerate(sub_df.iterrows(), start=1):
                ws.append([i] + [row[c] for c in sub_df.columns])

        # 质押比例相邻红绿条件格式（抽成模块级，下载端也复用）
        apply_pledge_ratio_coloring(wb)

        # 最新公告日首次出现红标（按 sheet 分基准：中大盘 sheet 用中大盘 baseline；小盘亦然）
        # 颜色：浅红（与质押比例变大的深红区分），同"新行增"语义
        new_row_red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header = [c.value for c in ws[1]]
            if "最新公告日" not in header:
                continue
            # 按 sheet 名前缀选对应 baseline
            sn = str(sheet_name or "").strip()
            if sn.startswith("中大盘"):
                sheet_baseline = baseline.get("中大盘") if isinstance(baseline, dict) else None
            elif sn.startswith("小盘"):
                sheet_baseline = baseline.get("小盘") if isinstance(baseline, dict) else None
            else:
                sheet_baseline = None
            date_col = header.index("最新公告日") + 1
            for row in range(2, ws.max_row + 1):
                date_val = ws.cell(row, date_col).value
                if not date_val:
                    continue
                try:
                    ts = pd.to_datetime(date_val, errors="coerce")
                    if pd.isna(ts):
                        continue
                except Exception:
                    continue
                if sheet_baseline is not None and ts > sheet_baseline:
                    ws.cell(row, date_col).fill = new_row_red_fill

        dirname = os.path.dirname(output_path)
        if dirname:
            os.makedirs(dirname, exist_ok=True)
        wb.save(output_path)
        # 统一样式：两 sheet 全部加筛选 / 列宽 / 居中
        auto_adjust_excel_width(output_path, fixed_width=22, all_sheets=True, center_align=True)
        logger.info(f"[质押 finalize] 已写出 {output_path}")

    def _detect_header_and_parse(self, df_all: pd.DataFrame, known_col_names: set, filepath: str) -> pd.DataFrame:
        """统一处理单行表头、双行表头、分组头三种情况。

        关键场景：部分列名在 Row1 就正确（如序号、证券代码），但另一部分列（如更新日期）
        在 Row1 是分组头（如"申报进程日期"），Row2 才是实际列名。
        策略：如果某个数据行包含列名中没有的已知列名，说明需要重映射。
        """
        # 收集当前列名中的已知列名
        col_name_set = set()
        for c in df_all.columns:
            s = str(c).strip()
            if s and not s.startswith('Unnamed'):
                col_name_set.add(s)
        cols_known = col_name_set & known_col_names

        # 扫描前10行数据，找包含已知列名的行（可能是真实表头行）
        best_row_idx = -1
        best_new_names = set()  # 该行提供的、列头中没有的已知列名
        best_total = 0
        for idx in range(min(10, len(df_all))):
            row_vals = set()
            for val in df_all.iloc[idx]:
                if pd.notna(val) and isinstance(val, str) and val.strip():
                    row_vals.add(val.strip())
            row_known = row_vals & known_col_names
            new_names = row_known - cols_known  # 该行带来的新列名
            total = len(row_known)
            # 优先选提供新列名最多的行；相同时选总匹配最多的
            if len(new_names) > len(best_new_names) or (len(new_names) == len(best_new_names) and total > best_total):
                best_new_names = new_names
                best_total = total
                best_row_idx = idx

        # 如果数据行提供了列头中没有的已知列名 → 需要重映射（限前3行，防止误匹配数据行）
        if len(best_new_names) >= 2 and best_total >= 2 and best_row_idx <= 3:
            header_row = df_all.iloc[best_row_idx]
            new_columns = []
            for orig_col, new_name in zip(df_all.columns, header_row):
                if pd.notna(new_name) and isinstance(new_name, str) and new_name.strip():
                    new_columns.append(new_name.strip())
                else:
                    new_columns.append(str(orig_col).strip())
            # 去重列名
            seen = {}
            unique_columns = []
            for col_name in new_columns:
                if col_name in seen:
                    seen[col_name] += 1
                    unique_columns.append(f"{col_name}_{seen[col_name]}")
                else:
                    seen[col_name] = 0
                    unique_columns.append(col_name)
            df = df_all.iloc[best_row_idx + 1:].copy()
            df.columns = unique_columns
            logger.info(f"表头重映射: 第{best_row_idx+2}行为实际表头(新增列名{best_new_names}), 最终列: {unique_columns}: {filepath}")
            return df

        # 列名已经正确，找数据起始行（序号=1）
        seq_col = None
        for col in df_all.columns:
            if '序号' in str(col):
                seq_col = col
                break
        if seq_col:
            for idx in range(min(20, len(df_all))):
                val = df_all.iloc[idx][seq_col]
                try:
                    if pd.notna(val) and int(float(str(val).strip())) == 1:
                        if idx > 0:
                            df = df_all.iloc[idx:].copy()
                            logger.info(f"跳过前{idx}行元数据: {filepath}")
                            return df
                        break
                except (ValueError, TypeError):
                    continue

        logger.info(f"普通表头，直接读取: {filepath}")
        return df_all.copy()

    async def execute_step(
        self,
        step_type: str,
        step_config: Dict,
        input_data: Optional[pd.DataFrame] = None,
        date_str: Optional[str] = None
    ) -> Dict[str, Any]:
        try:
            if step_type == "import_excel":
                return await self._import_excel(step_config, date_str)
            elif step_type == "merge_excel":
                return await self._merge_excel(step_config, date_str)
            elif step_type == "dedup":
                return await self._dedup(step_config, input_data, date_str)
            elif step_type == "smart_dedup":
                return await self._smart_dedup(step_config, input_data, date_str)
            elif step_type == "extract_columns":
                return await self._extract_columns(step_config, input_data, date_str)
            elif step_type == "export_excel":
                return await self._export_excel(step_config, input_data, date_str)
            elif step_type == "match_high_price":
                return await self._match_high_price(step_config, input_data, date_str)
            elif step_type == "match_ma20":
                return await self._match_ma20(step_config, input_data, date_str)
            elif step_type == "match_soe":
                return await self._match_soe(step_config, input_data, date_str)
            elif step_type == "match_sector":
                return await self._match_sector(step_config, input_data, date_str)
            elif step_type == "condition_intersection":
                return await self._condition_intersection(step_config, date_str)
            elif step_type == "export_ma20_trend":
                return await self._export_ma20_trend(step_config, date_str)
            elif step_type == "export_high_price_trend":
                return await self._export_high_price_trend(step_config, date_str)
            elif step_type == "ranking_sort":
                return await self._ranking_sort(step_config, input_data, date_str)
            elif step_type == "pledge_trend_analysis":
                return await self._pledge_trend_analysis(step_config, input_data, date_str)
            elif step_type == "pending":
                return {
                    "success": True,
                    "message": "步骤待定，跳过执行",
                    "data": input_data
                }
            else:
                return {
                    "success": False,
                    "message": f"未知步骤类型: {step_type}"
                }
        except Exception as e:
            logger.error(f"步骤执行失败: {step_type}, {str(e)}")
            return {
                "success": False,
                "message": f"步骤执行失败: {str(e)}"
            }

    async def _import_excel(self, config: Dict, date_str: Optional[str] = None) -> Dict[str, Any]:
        file_path = config.get("file_path")

        if not file_path:
            daily_dir = self._get_daily_dir(date_str)
            return {
                "success": False,
                "message": f"未指定文件路径，请将文件放入: {daily_dir}"
            }

        resolved_path = self._resolve_path(file_path, date_str)

        if not os.path.exists(resolved_path):
            return {
                "success": False,
                "message": f"文件不存在: {resolved_path}"
            }

        try:
            df = pd.read_excel(resolved_path)
            return {
                "success": True,
                "message": f"成功导入Excel，共{len(df)}行",
                "data": clean_df_for_json(df),
                "rows": len(df),
                "columns": list(df.columns),
                "file_path": resolved_path
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"Excel读取失败: {str(e)}"
            }

    async def _merge_excel(self, config: Dict, date_str: Optional[str] = None) -> Dict[str, Any]:
        daily_dir = self.resolver.get_upload_directory(date_str)
        files = self._get_excel_files_in_dir(daily_dir)

        public_dir = self.resolver.get_public_directory(date_str)
        public_files = []
        if not self.resolver.config.get("skip_public_in_merge") and os.path.exists(public_dir):
            public_files = self._get_excel_files_in_dir(public_dir)

        all_files = files + public_files

        if not all_files:
            return {
                "success": False,
                "message": f"目录下没有找到Excel文件: {daily_dir} 和 {public_dir}"
            }

        try:
            dfs = []
            exclude_patterns = config.get("exclude_patterns", ["total_", "output_", "deduped"])
            import re
            output_filename = config.get("output_filename", "")
            extra_excludes = []
            if output_filename:
                base_name = os.path.splitext(output_filename)[0]
                if base_name not in [p for p in exclude_patterns]:
                    extra_excludes.append(base_name)
            # 排除所有步骤可能产生的输出文件
            for step_type in ["merge_excel", "smart_dedup", "extract_columns",
                              "match_high_price", "match_ma20", "match_soe", "match_sector"]:
                auto_name = self.resolver.get_output_filename(step_type, date_str)
                if auto_name:
                    bn = os.path.splitext(auto_name)[0]
                    if bn not in exclude_patterns and bn not in extra_excludes:
                        extra_excludes.append(bn)

            for filepath in all_files:
                filename = os.path.basename(filepath)
                should_exclude = False
                all_patterns = exclude_patterns + extra_excludes
                for pattern in all_patterns:
                    if filename.startswith(pattern) or pattern in filename:
                        should_exclude = True
                        break
                if should_exclude:
                    continue

                is_public_file = check_is_public_file(filepath, public_dir)

                try:
                    # 质押：多 Sheet 遍历（含 public 文件），每 Sheet 单独检测表头并注入"来源"列
                    if self.workflow_type == "质押":
                        # 跳过历史的最终输出文件（如"5质押20260420.xlsx"）避免循环合并
                        if isinstance(filename, str) and filename.startswith("5质押"):
                            logger.info(f"质押：跳过最终输出文件: {filepath}")
                            continue
                        # 预先识别隐藏 sheet（state=hidden/veryHidden），合并时一律忽略
                        hidden_sheets = set()
                        try:
                            from openpyxl import load_workbook as _lwb
                            _wb = _lwb(filepath, read_only=True, data_only=True)
                            for _ws in _wb.worksheets:
                                if getattr(_ws, "sheet_state", "visible") != "visible":
                                    hidden_sheets.add(_ws.title)
                            _wb.close()
                        except Exception as e:
                            logger.warning(
                                f"质押：读取 sheet 可见性失败（忽略该检查）: {filepath}, {e}"
                            )
                        # 先尝试一次性读所有 sheet；失败则降级为逐 sheet 读取
                        sheet_map = None
                        try:
                            sheet_map = pd.read_excel(filepath, sheet_name=None)
                        except Exception as e:
                            logger.warning(
                                f"质押：一次性读全部 sheet 失败（{e}），降级为逐 sheet 读取: {filepath}"
                            )
                            try:
                                xl = pd.ExcelFile(filepath)
                                sheet_map = {}
                                for sn in xl.sheet_names:
                                    if sn in hidden_sheets:
                                        continue
                                    try:
                                        sheet_map[sn] = xl.parse(sn)
                                    except Exception as ee:
                                        logger.warning(f"质押：跳过损坏 sheet {sn}: {ee}")
                                        continue
                            except Exception as ee:
                                logger.warning(f"质押：文件彻底无法打开，跳过: {filepath}, {ee}")
                                continue
                        # 最终过滤掉隐藏的 sheet
                        if hidden_sheets and sheet_map:
                            before = len(sheet_map)
                            sheet_map = {k: v for k, v in sheet_map.items() if k not in hidden_sheets}
                            skipped = before - len(sheet_map)
                            if skipped > 0:
                                logger.info(
                                    f"质押：跳过 {skipped} 个隐藏 sheet ({sorted(hidden_sheets)[:5]}...): {filepath}"
                                )
                        known_col_names_pledge = {
                            "证券代码", "证券简称", "证券名称", "最新公告日",
                            "股权质押公告日期",
                        }
                        # 需要从源表彻底剔除的 4 类信息列（及其同义词变体）——
                        # 这些由 match_* 步骤权威填充，源表原始值一概不取
                        pledge_info_cols_to_drop = {
                            "百日新高", "百日最高价", "百日最高",
                            "站上20日线", "20日线", "20日均线", "20日均价",
                            "国央企", "国企", "国有企业",
                            "所属板块", "一级板块", "板块",
                        }
                        for sheet_name, df_sheet in sheet_map.items():
                            if df_sheet is None or len(df_sheet) == 0:
                                continue
                            df_parsed = self._detect_header_and_parse(
                                df_sheet, known_col_names_pledge,
                                f"{filepath}#{sheet_name}"
                            )
                            # 双行表头检测：若存在"质押比例"列且紧邻列名为空或重复后缀，
                            # 尝试用 openpyxl 读前两行单元格拼接列名（列名日期格式）
                            df_parsed = self._maybe_flatten_pledge_multiheader(
                                df_parsed, filepath, sheet_name
                            )
                            # 动态匹配"股权质押公告日期"前缀列（sheet 粒度，保证每个 sheet 独立生效）
                            pledge_date_cols = [
                                c for c in df_parsed.columns
                                if isinstance(c, str) and "股权质押公告日期" in c
                            ]
                            # 动态匹配原表的 3 个预判列（任意含该关键字的前缀列）
                            preset_inc_cols = [c for c in df_parsed.columns
                                               if isinstance(c, str) and "持续递增" in c]
                            preset_dec_cols = [c for c in df_parsed.columns
                                               if isinstance(c, str) and "持续递减" in c]
                            preset_event_cols = [c for c in df_parsed.columns
                                                 if isinstance(c, str) and "质押异动" in c]
                            df_parsed["来源"] = self._derive_pledge_source(filename, sheet_name)
                            df_parsed["_source_file"] = filename
                            # 质押：标记数据来源（public or 当日上传），smart_dedup 按此优先级去重
                            df_parsed["_is_public"] = 1 if is_public_file else 0
                            # 提前列过滤（质押白名单）
                            target_cols_pledge = (
                                ["序号", "证券代码", "证券简称", "证券名称",
                                 "最新公告日", "来源", "_source_file"]
                                + pledge_date_cols
                                + preset_inc_cols + preset_dec_cols + preset_event_cols
                            )
                            # 质押重构：
                            # 1) 移除 4 类信息列（及同义词变体），由 match_* 权威填充
                            # 2) 清理 pandas 产生的 Unnamed: 空列
                            # 3) 其余全部保留（含质押比例-xxx、股权质押公告日期-xxx 等）
                            cols_to_keep = [
                                c for c in df_parsed.columns
                                if not (isinstance(c, str) and (
                                    c.startswith("Unnamed") or c in pledge_info_cols_to_drop
                                ))
                            ]
                            if cols_to_keep and len(cols_to_keep) < len(df_parsed.columns):
                                df_parsed = df_parsed[cols_to_keep]
                            # target_cols_pledge 仅作为"必需列存在性"参考日志
                            missing_required = [c for c in target_cols_pledge if c not in df_parsed.columns]
                            if missing_required:
                                logger.debug(f"质押 sheet 缺少部分标准列（不阻塞）: {filepath}#{sheet_name} 缺={missing_required}")
                            dfs.append(df_parsed)
                            logger.info(f"读取 sheet: {filepath}#{sheet_name}, 行数: {len(df_parsed)}, 列数: {len(df_parsed.columns)}, 来源: {df_parsed['来源'].iloc[0] if len(df_parsed) else '-'}")
                        continue

                    if is_public_file:
                        public_skiprows = self.resolver.config.get("public_skiprows", 1)
                        df = self._read_public_file_cached(filepath, skiprows=public_skiprows)
                    elif self.workflow_type == "申报并购重组":
                        # 申报并购重组源文件第1行是分组头（如"申报进程日期"），直接跳过
                        df = pd.read_excel(filepath, skiprows=1)
                        if len(df) == 0:
                            continue
                    else:
                        df_all = pd.read_excel(filepath)
                        if len(df_all) > 0:
                            if self.resolver.config.get("skip_public_in_merge"):
                                # 涨幅排名等类型：直接使用第一行作为列头，不做表头检测
                                df = df_all
                            else:
                                known_col_names = {"证券代码", "证券简称", "最新公告日", "公告日期", "代码", "名称",
                                                   "首次公告日", "交易概述", "上市公告日", "更新日期", "受理日期", "重组类型",
                                                   "证券名称", "最新大股东减持公告日期", "发生日期"}
                                df = self._detect_header_and_parse(df_all, known_col_names, filepath)
                        else:
                            continue

                    # 减持叠加质押和大宗交易：Wind 风格列名含 \n 后缀，剥离以匹配标准列名
                    if self.workflow_type == "减持叠加质押和大宗交易":
                        df.columns = [c.split('\n')[0] if isinstance(c, str) else c for c in df.columns]

                    df["_source_file"] = filename
                    # 提前过滤到需要的列，减少 concat 内存（55列→4列）
                    # 涨幅排名等类型保留所有列
                    if not self.resolver.config.get("skip_public_in_merge"):
                        target_cols = ["序号", "证券代码", "证券简称", "最新公告日", "代码", "名称", "公告日期", "上市公告日", "更新日期",
                                       "证券名称", "最新大股东减持公告日期", "发生日期", "_source_file"]
                        available = [c for c in target_cols if c in df.columns]
                        if available:
                            df = df[available]
                    dfs.append(df)
                    logger.info(f"读取文件: {filepath}, 行数: {len(df)}")
                except Exception as e:
                    logger.warning(f"读取文件失败: {filepath}, 错误: {str(e)}")

            if not dfs:
                return {
                    "success": False,
                    "message": "没有可合并的数据"
                }

            merged_df = pd.concat(dfs, ignore_index=True)

            if self.workflow_type == "股权转让":
                column_mapping = {
                    "代码": "证券代码",
                    "名称": "证券简称",
                    "公告日期": "最新公告日"
                }
                merged_df = merged_df.rename(columns=column_mapping)
                logger.info(f"股权转让类型：列名映射完成 - {column_mapping}")

            if self.workflow_type == "增发实现" and "上市公告日" in merged_df.columns:
                if "最新公告日" in merged_df.columns:
                    merged_df["最新公告日"] = merged_df["最新公告日"].fillna(merged_df["上市公告日"])
                    merged_df = merged_df.drop(columns=["上市公告日"])
                else:
                    merged_df = merged_df.rename(columns={"上市公告日": "最新公告日"})
                logger.info(f"增发实现类型：上市公告日→最新公告日 映射完成")

            if self.workflow_type == "申报并购重组" and "更新日期" in merged_df.columns:
                # 源文件的"最新公告日"列多为"-"无效值，直接用"更新日期"替代
                if "最新公告日" in merged_df.columns:
                    merged_df = merged_df.drop(columns=["最新公告日"])
                merged_df = merged_df.rename(columns={"更新日期": "最新公告日"})
                logger.info(f"申报并购重组类型：更新日期→最新公告日 映射完成")

            if self.workflow_type == "减持叠加质押和大宗交易":
                # 清洗不规范字符：全角空格、不可见字符、前后空白
                for col in merged_df.columns:
                    if merged_df[col].dtype == object:
                        mask = merged_df[col].notna()
                        merged_df.loc[mask, col] = (merged_df.loc[mask, col]
                                                    .astype(str)
                                                    .str.replace('\u3000', ' ', regex=False)
                                                    .str.replace(r'[\x00-\x1f\x7f-\x9f]', '', regex=True)
                                                    .str.strip())

                # 列名映射：仅在目标列不存在时才做映射
                if "证券简称" not in merged_df.columns and "证券名称" in merged_df.columns:
                    merged_df = merged_df.rename(columns={"证券名称": "证券简称"})
                    logger.info("减持叠加质押和大宗交易：证券名称→证券简称 映射完成")
                if "最新公告日" not in merged_df.columns and "最新大股东减持公告日期" in merged_df.columns:
                    merged_df = merged_df.rename(columns={"最新大股东减持公告日期": "最新公告日"})
                    logger.info("减持叠加质押和大宗交易：最新大股东减持公告日期→最新公告日 映射完成")

            if self.workflow_type == "招投标":
                # 清洗不规范字符：全角空格、不可见字符、前后空白
                for col in merged_df.columns:
                    if merged_df[col].dtype == object:
                        mask = merged_df[col].notna()
                        merged_df.loc[mask, col] = (merged_df.loc[mask, col]
                                                    .astype(str)
                                                    .str.replace('\u3000', ' ', regex=False)
                                                    .str.replace(r'[\x00-\x1f\x7f-\x9f]', '', regex=True)
                                                    .str.strip())

                # 列名映射：仅在目标列不存在时才做映射
                if "证券简称" not in merged_df.columns and "证券名称" in merged_df.columns:
                    merged_df = merged_df.rename(columns={"证券名称": "证券简称"})
                    logger.info("招投标：证券名称→证券简称 映射完成")
                if "最新公告日" not in merged_df.columns and "发生日期" in merged_df.columns:
                    merged_df = merged_df.rename(columns={"发生日期": "最新公告日"})
                    logger.info("招投标：发生日期→最新公告日 映射完成")

            if self.workflow_type == "质押":
                # 清洗不规范字符
                for col in merged_df.columns:
                    if merged_df[col].dtype == object:
                        mask = merged_df[col].notna()
                        merged_df.loc[mask, col] = (merged_df.loc[mask, col]
                                                    .astype(str)
                                                    .str.replace('\u3000', ' ', regex=False)
                                                    .str.replace(r'[\x00-\x1f\x7f-\x9f]', '', regex=True)
                                                    .str.strip())

                # 证券名称 → 证券简称（仅在目标列不存在时）
                if "证券简称" not in merged_df.columns and "证券名称" in merged_df.columns:
                    merged_df = merged_df.rename(columns={"证券名称": "证券简称"})
                    logger.info("质押：证券名称→证券简称 映射完成")

                # 前缀含"股权质押公告日期" → 最新公告日
                if "最新公告日" not in merged_df.columns:
                    candidates = [c for c in merged_df.columns
                                  if isinstance(c, str) and "股权质押公告日期" in c]
                    if candidates:
                        primary = candidates[0]
                        for extra in candidates[1:]:
                            merged_df[primary] = merged_df[primary].fillna(merged_df[extra])
                            merged_df = merged_df.drop(columns=[extra])
                        merged_df = merged_df.rename(columns={primary: "最新公告日"})
                        logger.info(f"质押：{primary}→最新公告日 映射完成")

                # 若存在"股权质押公告日期"列 但 也存在"最新公告日"列，合并到最新公告日
                if "最新公告日" in merged_df.columns:
                    leftover = [c for c in merged_df.columns
                                if isinstance(c, str) and "股权质押公告日期" in c and c != "最新公告日"]
                    for extra in leftover:
                        merged_df["最新公告日"] = merged_df["最新公告日"].fillna(merged_df[extra])
                        merged_df = merged_df.drop(columns=[extra])

                # 预判列 rename：前缀列 → 标准列名（任一行有值即跳过后续计算）
                def _coalesce_rename(merged_df, prefix_keyword, target_name):
                    if target_name in merged_df.columns:
                        return merged_df
                    candidates = [c for c in merged_df.columns
                                  if isinstance(c, str) and prefix_keyword in c]
                    if not candidates:
                        return merged_df
                    primary = candidates[0]
                    for extra in candidates[1:]:
                        merged_df[primary] = merged_df[primary].fillna(merged_df[extra])
                        merged_df = merged_df.drop(columns=[extra])
                    return merged_df.rename(columns={primary: target_name})

                merged_df = _coalesce_rename(merged_df, "持续递增", "持续递增（一年内）")
                merged_df = _coalesce_rename(merged_df, "持续递减", "持续递减（一年内）")
                merged_df = _coalesce_rename(merged_df, "质押异动", "质押异动")

                # 过滤脏数据行（Choice 元信息 / 证券代码空 / "数据来源" 脚注）
                if "证券代码" in merged_df.columns:
                    _code_str = merged_df["证券代码"].astype(str).str.strip()
                    _bad_mask = (
                        merged_df["证券代码"].isna()
                        | (_code_str == "")
                        | _code_str.str.lower().eq("nan")
                        | _code_str.str.contains("数据来源", na=False, regex=False)
                        | _code_str.str.contains("妙想Choice", na=False, regex=False)
                        | _code_str.str.contains("Choice", na=False, regex=False)
                    )
                    bad_count = int(_bad_mask.sum())
                    if bad_count > 0:
                        logger.info(
                            f"质押：过滤脏数据行 {bad_count} 条（"
                            f"空/数据来源/Choice 等元信息行）"
                        )
                        merged_df = merged_df[~_bad_mask].reset_index(drop=True)

            # 删除内部辅助列
            if "_source_file" in merged_df.columns:
                merged_df = merged_df.drop(columns=["_source_file"])

            # skip_public_in_merge (涨幅排名): 跳过标准列提取、日期排序、序号处理，
            # 保留原始列结构供后续 ranking_sort 步骤使用
            if not self.resolver.config.get("skip_public_in_merge"):
                if self.workflow_type == "质押":
                    # 质押类型：保留源表全部业务列（质押比例-*/股权质押公告日期-*/大宗交易/总市值 等），
                    # 仅去除重复列名；4 类信息列已在 sheet 粒度提前过滤
                    merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
                else:
                    keep_columns = ["序号", "证券代码", "证券简称", "最新公告日"]
                    existing_columns = [col for col in keep_columns if col in merged_df.columns]
                    merged_df = merged_df[existing_columns]
                    # 去除重复列名（concat 不同格式文件 + rename 可能产生重复列）
                    merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]

                date_col = None
                for col in ["最新公告日", "公告日", "日期", "date"]:
                    if col in merged_df.columns:
                        date_col = col
                        break

                if date_col:
                    def parse_date(val):
                        if pd.isna(val):
                            return None
                        if isinstance(val, str):
                            val = val.strip()
                            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%d/%m/%Y", "%m/%d/%Y"]:
                                try:
                                    return pd.to_datetime(val, format=fmt)
                                except:
                                    continue
                        try:
                            return pd.to_datetime(val)
                        except:
                            return None

                    merged_df["_sort_date"] = merged_df[date_col].apply(parse_date)
                    merged_df = merged_df.sort_values("_sort_date", ascending=False)
                    merged_df = merged_df.drop(columns=["_sort_date"])
                    merged_df[date_col] = pd.to_datetime(merged_df[date_col], errors="coerce")
                    merged_df[date_col] = merged_df[date_col].dt.strftime("%Y-%m-%d")

                if "序号" in merged_df.columns:
                    merged_df["序号"] = pd.to_numeric(merged_df["序号"], errors='coerce')
                    merged_df = merged_df.dropna(subset=["序号"])
                    merged_df["序号"] = range(1, len(merged_df) + 1)
                else:
                    merged_df.insert(0, "序号", range(1, len(merged_df) + 1))

            output_filename = config.get("output_filename") or self.resolver.get_output_filename("merge_excel", date_str)
            output_path = os.path.join(daily_dir, output_filename)
            merged_df.to_excel(output_path, index=False)

            return {
                "success": True,
                "message": f"合并完成: {len(dfs)}个文件, 共{len(merged_df)}行",
                "data": clean_df_for_json(merged_df),
                "file_path": output_path,
                "rows": len(merged_df),
                "files_merged": len(dfs),
                "date_str": date_str or self.today,
                "_df": merged_df
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"合并Excel失败: {str(e)}"
            }

    async def _dedup(self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None) -> Dict[str, Any]:
        if df is None:
            return {
                "success": False,
                "message": "没有可处理的数据"
            }

        original_rows = len(df)
        df_deduped = df.drop_duplicates()

        return {
            "success": True,
            "message": f"去除重复行完成: {original_rows} -> {len(df_deduped)}",
            "data": clean_df_for_json(df_deduped),
            "original_rows": original_rows,
            "deduped_rows": len(df_deduped),
            "removed_rows": original_rows - len(df_deduped)
        }

    async def _smart_dedup(self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None) -> Dict[str, Any]:
        if df is None:
            return {
                "success": False,
                "message": "没有可处理的数据"
            }

        stock_code_col = config.get("stock_code_column")
        date_col = config.get("date_column")

        available_columns = [str(c) for c in df.columns.tolist()]

        if not stock_code_col:
            stock_code_candidates = ["证券代码", "股票代码", "代码", "code", "stock_code", "symbol"]
            for candidate in stock_code_candidates:
                for col in available_columns:
                    if candidate.lower() in col.lower():
                        stock_code_col = col
                        break
                if stock_code_col:
                    break

        if not date_col:
            date_candidates = ["最新公告日", "公告日", "更新日期", "日期", "date", "announcement_date", "report_date"]
            for candidate in date_candidates:
                for col in available_columns:
                    if candidate.lower() in col.lower():
                        date_col = col
                        break
                if date_col:
                    break

        if not stock_code_col:
            return {
                "success": False,
                "message": f"未找到证券代码列，请手动指定。可用列: {available_columns}"
            }

        if not date_col:
            return {
                "success": False,
                "message": f"未找到日期列，请手动指定。可用列: {available_columns}"
            }

        original_rows = len(df)

        try:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
        except Exception:
            pass

        # 质押类型：优先保留"本次上传"的行（_is_public=0），再"3 预判列任一非空"，再按最新公告日
        # 其他类型：保持原逻辑（按日期降序 + keep=first）
        if self.workflow_type == "质押":
            df = df.copy()
            # 保证 _is_public 列存在（来自 merge_excel 阶段，public=1 / 上传=0）
            if "_is_public" not in df.columns:
                df["_is_public"] = 0
            preset_cols = ["持续递增（一年内）", "持续递减（一年内）", "质押异动"]
            existing_preset = [c for c in preset_cols if c in df.columns]
            if existing_preset:
                def _has_preset(row):
                    for c in existing_preset:
                        v = row.get(c)
                        if v is None:
                            continue
                        s = str(v).strip()
                        if s != "" and s.lower() != "nan":
                            return 1
                    return 0
                df["_has_preset"] = df.apply(_has_preset, axis=1)
                # 优先级：_is_public 升序（上传=0 在前）→ _has_preset 降序 → date_col 降序
                df_sorted = df.sort_values(
                    by=["_is_public", "_has_preset", date_col],
                    ascending=[True, False, False],
                    kind="mergesort",  # 稳定排序
                )
                df_deduped = df_sorted.drop_duplicates(
                    subset=[stock_code_col], keep="first"
                )
                df_deduped = df_deduped.drop(columns=["_has_preset", "_is_public"])
            else:
                df_sorted = df.sort_values(
                    by=["_is_public", date_col],
                    ascending=[True, False],
                    kind="mergesort",
                )
                df_deduped = df_sorted.drop_duplicates(subset=[stock_code_col], keep="first")
                df_deduped = df_deduped.drop(columns=["_is_public"])
        else:
            df_sorted = df.sort_values(date_col, ascending=False)
            df_deduped = df_sorted.drop_duplicates(subset=[stock_code_col], keep="first")

        df_deduped = df_deduped[~df_deduped[stock_code_col].astype(str).str.upper().str.endswith('.NQ')]

        df_deduped[date_col] = pd.to_datetime(df_deduped[date_col], errors="coerce")
        df_deduped[date_col] = df_deduped[date_col].dt.strftime("%Y-%m-%d")

        df_deduped = df_deduped.sort_values(date_col, ascending=False).reset_index(drop=True)

        if "序号" in df_deduped.columns:
            df_deduped["序号"] = range(1, len(df_deduped) + 1)
        elif date_col in df_deduped.columns:
            df_deduped.insert(0, "序号", range(1, len(df_deduped) + 1))

        if "_source_file" in df_deduped.columns:
            df_deduped = df_deduped.drop(columns=["_source_file"])

        removed_rows = original_rows - len(df_deduped)

        daily_dir = self.resolver.get_upload_directory(date_str)
        deduped_filename = self.resolver.get_output_filename("smart_dedup", date_str, config.get("output_filename"))
        deduped_path = os.path.join(daily_dir, deduped_filename)
        df_deduped.to_excel(deduped_path, index=False)

        return {
            "success": True,
            "message": f"智能去重完成: {original_rows} -> {len(df_deduped)} (删除{removed_rows}行重复数据)",
            "data": clean_df_for_json(df_deduped),
            "file_path": deduped_path,
            "original_rows": original_rows,
            "deduped_rows": len(df_deduped),
            "removed_rows": removed_rows,
            "stock_code_column": stock_code_col,
            "date_column": date_col,
            "_df": df_deduped
        }

    async def _extract_columns_pledge(self, df: pd.DataFrame, output_path: str) -> Dict[str, Any]:
        """质押类型的 extract_columns：保留全部原始列，补齐必须列。

        - 不做白名单过滤
        - 保证 证券代码/证券简称/最新公告日/来源 四列存在（缺则补空列或从别名映射）
        - 最新公告日：缺失时找前缀含"股权质押公告日期"的第一列复制值（不删原列）
        - 证券简称：缺失时尝试 名称 / 证券名称
        - 不追加 持续递增/递减/质押异动（由 pledge_trend_analysis 负责）
        - 写出无样式 xlsx
        """
        if df is None:
            df = pd.DataFrame()
        df = df.copy()

        if "证券代码" not in df.columns:
            df["证券代码"] = ""

        if "证券简称" not in df.columns:
            for alias in ("名称", "证券名称"):
                if alias in df.columns:
                    df["证券简称"] = df[alias]
                    break
            else:
                df["证券简称"] = ""

        if "最新公告日" not in df.columns:
            pledge_date_cols = [c for c in df.columns if "股权质押公告日期" in str(c)]
            if pledge_date_cols:
                df["最新公告日"] = df[pledge_date_cols[0]]
            else:
                df["最新公告日"] = ""

        if "来源" not in df.columns:
            df["来源"] = "小盘"

        dirname = os.path.dirname(output_path)
        if dirname:
            os.makedirs(dirname, exist_ok=True)
        df.to_excel(output_path, index=False)
        return {
            "success": True,
            "message": f"质押 extract_columns 完成，保留 {len(df.columns)} 列",
            "data": clean_df_for_json(df),
            "columns": list(df.columns),
            "rows": len(df),
            "file_path": output_path,
            "_df": df,
        }

    async def _extract_columns(self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None) -> Dict[str, Any]:
        if self.workflow_type == "质押":
            # 重构：质押分支保留全列，不走白名单过滤
            if df is None:
                return {"success": False, "message": "没有可处理的数据"}
            daily_dir = self.resolver.get_upload_directory(date_str)
            output_name = self.resolver.get_output_filename("extract_columns", date_str) or "output_2.xlsx"
            output_path = os.path.join(daily_dir, output_name)
            return await self._extract_columns_pledge(df, output_path)

        if df is None:
            return {
                "success": False,
                "message": "没有可处理的数据"
            }

        available_columns = [str(c) for c in df.columns.tolist()]
        selected_columns = []

        columns_config = config.get("columns")
        if columns_config and len(columns_config) > 0:
            for col_name in columns_config:
                col_name_str = str(col_name)
                for avail_col in available_columns:
                    if col_name_str in avail_col or avail_col in col_name_str:
                        selected_columns.append(avail_col)
                        break
                else:
                    return {
                        "success": False,
                        "message": f"未找到列: {col_name}，可用列: {available_columns}"
                    }
        else:
            fixed_columns = ["序号", "证券代码", "证券简称", "最新公告日"]
            for fixed_col in fixed_columns:
                found = False
                for avail_col in available_columns:
                    if fixed_col in avail_col or avail_col in fixed_col:
                        selected_columns.append(avail_col)
                        found = True
                        break
                if not found:
                    for avail_col in available_columns:
                        if "序号" in avail_col and "证券" not in avail_col:
                            selected_columns.append(avail_col)
                            found = True
                            break

            if len(selected_columns) < len(fixed_columns):
                missing = []
                for i, col in enumerate(fixed_columns):
                    col_str = str(col)
                    found = any(col_str in str(c) or str(c) in col_str for c in available_columns)
                    if not found:
                        missing.append(col)
                if missing:
                    return {
                        "success": False,
                        "message": f"未找到以下列: {missing}，可用列: {available_columns}"
                    }

        df_extracted = df[selected_columns].copy()

        if '最新公告日' in df_extracted.columns:
            df_extracted = df_extracted.sort_values('最新公告日', ascending=False)

        if '序号' in df_extracted.columns:
            df_extracted['序号'] = range(1, len(df_extracted) + 1)
        elif selected_columns and '序号' in selected_columns[0]:
            df_extracted.iloc[:, 0] = range(1, len(df_extracted) + 1)

        if '最新公告日' in df_extracted.columns:
            df_extracted['最新公告日'] = pd.to_datetime(df_extracted['最新公告日']).dt.strftime('%Y-%m-%d')

        output_filename = config.get("output_filename")
        file_path = None
        if output_filename:
            daily_dir = self.resolver.get_upload_directory(date_str)
            file_path = os.path.join(daily_dir, output_filename)

            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active

            for r in dataframe_to_rows(df_extracted, index=False, header=True):
                ws.append(r)

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max(max_length + 4, 15)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)

        return {
            "success": True,
            "message": f"成功提取列: {selected_columns}",
            "data": clean_df_for_json(df_extracted),
            "columns": selected_columns,
            "file_path": file_path,
            "rows": len(df_extracted)
        }

    async def _export_excel(self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None) -> Dict[str, Any]:
        if df is None:
            return {
                "success": False,
                "message": "没有可导出的数据"
            }

        filename = config.get("output_filename") or self.resolver.get_output_filename("export_excel", date_str)
        daily_dir = self.resolver.get_upload_directory(date_str)
        filepath = os.path.join(daily_dir, filename)

        try:
            df.to_excel(filepath, index=False)
            auto_adjust_excel_width(filepath)
            return {
                "success": True,
                "message": f"成功导出Excel: {filepath}",
                "file_path": filepath,
                "rows": len(df),
                "date_str": date_str or self.today
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"Excel导出失败: {str(e)}"
            }

    async def _match_high_price(self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None) -> Dict[str, Any]:
        if df is None:
            return {
                "success": False,
                "message": "没有可匹配的数据"
            }

        source_path = config.get("source_dir") or self.resolver.get_match_source_directory("match_high_price", date_str)
        new_column_name = config.get("new_column_name", "百日新高")

        # 确保目录存在且有文件（自动从历史日期复制）
        if date_str:
            source_path = self.resolver.ensure_match_source_files("match_high_price", date_str)

        if not os.path.exists(source_path):
            return {
                "success": False,
                "message": f"目录不存在: {source_path}"
            }

        all_high_stocks = self._load_match_source(
            source_path,
            code_col_candidates=['股票代码', '股票代码.1', '证券代码'],
            name_col_candidates=['股票简称', '证券简称']
        )

        logger.info(f"从{source_path}目录共加载{len(all_high_stocks)}只新高股票")

        df[new_column_name] = df['证券代码'].apply(
            lambda code: match_stock_code_flexible(code, all_high_stocks)
        )

        matched_count = (df[new_column_name] != '').sum()
        logger.info(f"匹配完成，共匹配{matched_count}条记录")

        output_filename = config.get("output_filename") or self.resolver.get_output_filename("match_high_price", date_str)
        output_path = os.path.join(self._get_daily_dir(date_str), output_filename)
        df.to_excel(output_path, index=False)
        logger.info(f"结果已保存到: {output_path}")

        df_clean = df.fillna('')
        records = df_clean.head(100).to_dict('records')
        for record in records:
            for k, v in record.items():
                if isinstance(v, (float, int)) and (v != v or abs(v) == float('inf')):
                    record[k] = ''

        return {
            "success": True,
            "message": f"匹配完成，匹配{matched_count}条记录，已保存到{output_filename}",
            "data": records,
            "columns": df.columns.tolist(),
            "rows": len(df),
            "file_path": output_path,
            "_df": df
        }

    async def _match_ma20(self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None) -> Dict[str, Any]:
        if df is None:
            return {
                "success": False,
                "message": "没有可匹配的数据"
            }

        source_path = config.get("source_dir") or self.resolver.get_match_source_directory("match_ma20", date_str)
        # 质押类型：统一使用"站上20日线"作为最终列名（和 finalize 前 7 列对齐）
        default_col = "站上20日线" if self.workflow_type == "质押" else "20日均线"
        new_column_name = config.get("new_column_name", default_col)

        if date_str:
            source_path = self.resolver.ensure_match_source_files("match_ma20", date_str)

        if not os.path.exists(source_path):
            return {
                "success": False,
                "message": f"目录不存在: {source_path}"
            }

        all_ma20_stocks = self._load_match_source(
            source_path,
            code_col_candidates=['股票代码', '股票代码.1', '证券代码'],
            name_col_candidates=['股票简称', '证券简称']
        )

        logger.info(f"从{source_path}目录共加载{len(all_ma20_stocks)}只股票")

        try:
            if '证券代码' not in df.columns:
                return {"success": False, "message": f"输入数据缺少'证券代码'列, 可用列: {list(df.columns)}"}

            df = df.copy()
            df[new_column_name] = df['证券代码'].apply(
                lambda code: match_stock_code_flexible(code, all_ma20_stocks)
            )

            matched_count = (df[new_column_name] != '').sum()
            logger.info(f"匹配完成，共匹配{matched_count}条记录")

            output_filename = config.get("output_filename") or self.resolver.get_output_filename("match_ma20", date_str)
            output_path = os.path.join(self._get_daily_dir(date_str), output_filename)
            df.to_excel(output_path, index=False)
            logger.info(f"结果已保存到: {output_path}")
        except Exception as e:
            logger.error(f"匹配20日均线执行失败: {e}")
            return {"success": False, "message": f"匹配20日均线失败: {str(e)}"}

        df_clean = df.fillna('')
        records = df_clean.head(100).to_dict('records')
        for record in records:
            for k, v in record.items():
                if isinstance(v, (float, int)) and (v != v or abs(v) == float('inf')):
                    record[k] = ''

        return {
            "success": True,
            "message": f"匹配完成，匹配{matched_count}条记录，已保存到{output_filename}",
            "data": records,
            "columns": df.columns.tolist(),
            "rows": len(df),
            "file_path": output_path,
            "_df": df
        }

    async def _match_soe(self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None) -> Dict[str, Any]:
        if df is None:
            return {
                "success": False,
                "message": "没有可匹配的数据"
            }

        source_path = config.get("source_dir") or self.resolver.get_match_source_directory("match_soe", date_str)
        # 质押类型：统一使用"国央企"作为最终列名（和 finalize 前 7 列对齐）
        default_col = "国央企" if self.workflow_type == "质押" else "国企"
        new_column_name = config.get("new_column_name", default_col)

        if date_str:
            source_path = self.resolver.ensure_match_source_files("match_soe", date_str)

        if not os.path.exists(source_path):
            return {
                "success": False,
                "message": f"目录不存在: {source_path}"
            }

        all_soe_stocks = self._load_match_source(
            source_path,
            code_col_candidates=['股票代码.1', '股票代码', '证券代码'],
            name_col_candidates=['股票简称', '证券简称']
        )

        logger.info(f"从{source_path}目录共加载{len(all_soe_stocks)}只国企股票")

        df[new_column_name] = df['证券代码'].apply(
            lambda code: match_stock_code_flexible(code, all_soe_stocks)
        )

        matched_count = (df[new_column_name] != '').sum()
        logger.info(f"匹配完成，共匹配{matched_count}条记录")

        output_filename = config.get("output_filename") or self.resolver.get_output_filename("match_soe", date_str)
        output_path = os.path.join(self._get_daily_dir(date_str), output_filename)
        df.to_excel(output_path, index=False)
        logger.info(f"结果已保存到: {output_path}")

        df_clean = df.fillna('')
        records = df_clean.head(100).to_dict('records')
        for record in records:
            for k, v in record.items():
                if isinstance(v, (float, int)) and (v != v or abs(v) == float('inf')):
                    record[k] = ''

        return {
            "success": True,
            "message": f"匹配完成，匹配{matched_count}条记录，已保存到{output_filename}",
            "data": records,
            "columns": df.columns.tolist(),
            "rows": len(df),
            "file_path": output_path,
            "_df": df
        }

    async def _match_sector(self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None) -> Dict[str, Any]:
        if df is None:
            return {
                "success": False,
                "message": "没有可匹配的数据"
            }

        source_path = config.get("source_dir") or self.resolver.get_match_source_directory("match_sector", date_str)
        new_column_name = config.get("new_column_name", "所属板块")

        if date_str:
            source_path = self.resolver.ensure_match_source_files("match_sector", date_str)

        if not os.path.exists(source_path):
            return {
                "success": False,
                "message": f"目录不存在: {source_path}"
            }

        all_sector_stocks = self._load_match_source(
            source_path,
            code_col_candidates=['股票代码.1', '股票代码', '证券代码'],
            name_col_candidates=['所属板块', '所属一级板块', '一级板块', '板块', '所属二级板块', '二级板块']
        )

        logger.info(f"从{source_path}目录共加载{len(all_sector_stocks)}只股票")

        df[new_column_name] = df['证券代码'].apply(
            lambda code: match_stock_code_flexible(code, all_sector_stocks)
        )

        matched_count = (df[new_column_name] != '').sum()
        logger.info(f"匹配完成，共匹配{matched_count}条记录")

        output_filename = config.get("output_filename") or self.resolver.get_output_filename("match_sector", date_str)
        output_path = os.path.join(self._get_daily_dir(date_str), output_filename)
        df.to_excel(output_path, index=False)
        auto_adjust_excel_width(output_path)
        logger.info(f"结果已保存到: {output_path}")

        # 自动采集20日均线趋势数据（质押类型列名为"站上20日线"，其他为"20日均线"）
        ma20_col = None
        for candidate in ("站上20日线", "20日均线"):
            if candidate in df.columns:
                ma20_col = candidate
                break
        if ma20_col:
            try:
                from services.trend_service import save_trend_data
                # 质押类型：按"来源"(中大盘/小盘)拆成两个子类型独立存储
                if self.workflow_type == "质押" and "来源" in df.columns:
                    src_series = df["来源"].fillna("").astype(str).str.strip()
                    for src_label in ["中大盘", "小盘"]:
                        mask = src_series == src_label
                        sub_total = int(mask.sum())
                        if sub_total == 0:
                            continue
                        sub_count = int(
                            (df.loc[mask, ma20_col].fillna('').astype(str).str.strip() != '').sum()
                        )
                        sub_wt = f"质押({src_label})"
                        await save_trend_data(
                            metric_type='ma20',
                            workflow_type=sub_wt,
                            date_str=date_str or self.today,
                            count=sub_count,
                            total=sub_total,
                            source='auto',
                        )
                        logger.info(
                            f"自动采集MA20趋势: type={sub_wt}, count={sub_count}, total={sub_total}"
                        )
                elif (self.workflow_type or '并购重组') in ("并购重组", "股权转让", "招投标"):
                    from datetime import datetime as _dt_y
                    parent = self.workflow_type or '并购重组'
                    year_now = _dt_y.now().year
                    date_series = pd.to_datetime(df.get("最新公告日"), errors="coerce") \
                        if "最新公告日" in df.columns else pd.Series([pd.NaT] * len(df))
                    for year in (year_now, year_now - 1):
                        mask_year = date_series.dt.year == year
                        sub_total = int(mask_year.sum())
                        if sub_total == 0:
                            continue
                        sub_count = int(
                            (df.loc[mask_year, ma20_col].fillna('').astype(str).str.strip() != '').sum()
                        )
                        sub_wt = f"{parent}({year})"
                        await save_trend_data(
                            metric_type='ma20',
                            workflow_type=sub_wt,
                            date_str=date_str or self.today,
                            count=sub_count,
                            total=sub_total,
                            source='auto',
                        )
                        logger.info(
                            f"自动采集MA20趋势(年度): type={sub_wt}, count={sub_count}, total={sub_total}"
                        )
                else:
                    ma20_count = int((df[ma20_col].fillna('').astype(str).str.strip() != '').sum())
                    ma20_total = len(df)
                    await save_trend_data(
                        metric_type='ma20',
                        workflow_type=self.workflow_type or '并购重组',
                        date_str=date_str or self.today,
                        count=ma20_count,
                        total=ma20_total,
                        source='auto'
                    )
                    logger.info(
                        f"自动采集MA20趋势: type={self.workflow_type}, "
                        f"count={ma20_count}, total={ma20_total}"
                    )
            except Exception as e:
                logger.warning(f"自动采集MA20趋势失败: {e}")

        df_clean = df.fillna('')
        records = df_clean.head(100).to_dict('records')
        for record in records:
            for k, v in record.items():
                if isinstance(v, (float, int)) and (v != v or abs(v) == float('inf')):
                    record[k] = ''

        return {
            "success": True,
            "message": f"匹配完成，匹配{matched_count}条记录，已保存到{output_filename}",
            "data": records,
            "columns": df.columns.tolist(),
            "rows": len(df),
            "file_path": output_path,
            "_df": df
        }

    async def _condition_intersection(self, config: Dict, date_str: Optional[str] = None) -> Dict[str, Any]:
        """条件交集步骤：聚合所有工作流类型的最终数据，过滤后合并输出+交集选股池"""
        import zlib
        import json as json_mod
        from config.workflow_type_config import (
            WORKFLOW_TYPE_CONFIG, INTERSECTION_SOURCE_COLUMNS,
            INTERSECTION_COLUMN_RENAME, INTERSECTION_DISPLAY_COLUMNS
        )
        from core.database import AsyncSessionLocal
        from sqlalchemy import text

        date_str = config.get("date_str") or date_str or self.today
        filter_conditions = config.get("filter_conditions", [{"column": "百日新高", "enabled": True}])
        filter_logic = config.get("filter_logic", "AND")
        type_order = config.get("type_order", WORKFLOW_TYPE_CONFIG.get("条件交集", {}).get("default_type_order", []))
        # 兼容旧工作流：若 type_order 缺"质押"（2026-04 新增），自动追加到默认位置（6 号：减持后）
        if "质押" not in type_order:
            insert_at = len(type_order)
            for i, t in enumerate(type_order):
                if t == "减持叠加质押和大宗交易":
                    insert_at = i + 1
                    break
            type_order = list(type_order)
            type_order.insert(insert_at, "质押")
            logger.info(f"[条件交集] 旧工作流缺'质押'，自动追加到位置 {insert_at}：{type_order}")
        output_filename = config.get("output_filename") or f"7条件交集{date_str.replace('-', '')}.xlsx"
        workflow_id = config.get("_workflow_id")

        logger.info(f"[条件交集] 开始执行: date={date_str}, filters={filter_conditions}, logic={filter_logic}")

        # 1. 从 DB 获取各类型 final 数据
        type_dataframes = {}
        async with AsyncSessionLocal() as session:
            for wtype in type_order:
                # 并购重组 type 可能为空字符串，用 IN 参数化查询
                if wtype == "并购重组":
                    query = text("""
                        SELECT data_compressed FROM workflow_results
                        WHERE workflow_type IN (:wtype_empty, :wtype_cn)
                          AND date_str = :date_str AND step_type = 'final'
                        ORDER BY created_at DESC LIMIT 1
                    """)
                    params = {"date_str": date_str, "wtype_empty": "", "wtype_cn": "并购重组"}
                else:
                    query = text("""
                        SELECT data_compressed FROM workflow_results
                        WHERE workflow_type = :wtype
                          AND date_str = :date_str AND step_type = 'final'
                        ORDER BY created_at DESC LIMIT 1
                    """)
                    params = {"date_str": date_str, "wtype": wtype}

                result = await session.execute(query, params)
                row = result.fetchone()
                if row and row[0]:
                    try:
                        decompressed = zlib.decompress(row[0])
                        records = json_mod.loads(decompressed.decode("utf-8"))
                        df = pd.DataFrame(records)
                        type_dataframes[wtype] = df
                        logger.info(f"[条件交集] 读取 {wtype}: {len(df)}行")
                    except Exception as e:
                        logger.warning(f"[条件交集] 解压 {wtype} 数据失败: {e}")
                else:
                    logger.info(f"[条件交集] {wtype} 无数据，跳过")

        if not type_dataframes:
            return {"success": False, "message": "所有工作流类型均无数据，无法执行条件交集"}

        # 2. 对每个类型应用过滤条件
        filtered_dfs = {}
        for wtype, df in type_dataframes.items():
            filtered = self._apply_filters(df, filter_conditions, filter_logic)
            if len(filtered) > 0:
                filtered_dfs[wtype] = filtered
                logger.info(f"[条件交集] {wtype} 过滤后: {len(filtered)}行 (原{len(df)}行)")
            else:
                logger.info(f"[条件交集] {wtype} 过滤后无数据")

        if not filtered_dfs:
            return {"success": False, "message": "所有工作流类型过滤后均无数据"}

        # 3. 合并数据，按 type_order 顺序
        combined_rows = []
        warnings_list = []   # 冒泡到返回值的警告
        for wtype in type_order:
            if wtype not in filtered_dfs:
                continue
            df = filtered_dfs[wtype]
            # 提取标准列
            SECTOR_ALIASES = ['所属板块', '一级板块', '所属一级板块', '板块', '二级板块', '所属二级板块']
            # 质押类型 match_* 产出列名是权威名（"站上20日线"/"国央企"），和其他类型的
            # 源列名（"20日均线"/"国企"）不同。在此做反向 alias：查找 df 时如果源列名
            # 不存在，尝试其对应的权威名（INTERSECTION_COLUMN_RENAME 的 value）。
            REVERSE_RENAME_ALIASES = {
                "20日均线": ["20日均线", "站上20日线"],
                "国企": ["国企", "国央企"],
            }
            extracted = pd.DataFrame()
            for col in INTERSECTION_SOURCE_COLUMNS:
                if col in df.columns:
                    extracted[col] = df[col]
                elif col in REVERSE_RENAME_ALIASES:
                    alias_found = next((a for a in REVERSE_RENAME_ALIASES[col] if a in df.columns), None)
                    extracted[col] = df[alias_found] if alias_found else ""
                elif col in SECTOR_ALIASES:
                    found = next((a for a in SECTOR_ALIASES if a in df.columns), None)
                    extracted[col] = df[found] if found else ""
                else:
                    extracted[col] = ""

            # 资本运作行为列：质押类型按"来源"细分为"质押中大盘"/"质押小盘"
            if wtype == "质押":
                if "来源" in df.columns:
                    source_series = df["来源"].fillna("").astype(str).str.strip()
                    missing_mask = ~source_series.isin(["中大盘", "小盘"])
                    missing_count = int(missing_mask.sum())
                    if missing_count > 0:
                        sample_codes = []
                        if "证券代码" in df.columns:
                            sample_codes = df.loc[missing_mask, "证券代码"].head(5).tolist()
                        warnings_list.append(
                            f"质押类型 {missing_count}/{len(df)} 行'来源'字段缺失或非法，"
                            f"已兜底为'质押小盘'。示例: {sample_codes}"
                        )
                        logger.warning(f"[条件交集] {warnings_list[-1]}")
                    extracted["资本运作行为"] = source_series.apply(
                        lambda s: "质押中大盘" if s == "中大盘" else "质押小盘"
                    ).values
                else:
                    warnings_list.append("质押类型 final 数据完全缺少'来源'列，全部归入'质押小盘'")
                    logger.warning(f"[条件交集] {warnings_list[-1]}")
                    extracted["资本运作行为"] = "质押小盘"
            else:
                display_name = WORKFLOW_TYPE_CONFIG.get(wtype, {}).get("display_name", wtype)
                extracted["资本运作行为"] = display_name
            combined_rows.append(extracted)

        combined_df = pd.concat(combined_rows, ignore_index=True)

        # 去重合并：同证券代码多行 → 保留"最新公告日"那行，资本运作行为拼接为"A、B"
        if "证券代码" in combined_df.columns and len(combined_df) > 0:
            before = len(combined_df)
            # 原顺序索引，用于"最新公告日"相同时的 tie-breaker（先到先留）
            combined_df = combined_df.reset_index(drop=True)
            combined_df["_orig_idx"] = combined_df.index
            date_parsed = pd.to_datetime(combined_df.get("最新公告日"), errors="coerce")
            combined_df["_date_sort_key"] = date_parsed
            # 按 证券代码 分组，聚合"资本运作行为"为去重顿号拼接（保持首次出现顺序）
            def _merge_behaviors(series):
                seen = []
                for v in series:
                    s = str(v).strip() if v is not None and str(v).strip() != "" else ""
                    if not s:
                        continue
                    # 单个 cell 内已有顿号时分开处理
                    for part in s.split("、"):
                        part = part.strip()
                        if part and part not in seen:
                            seen.append(part)
                return "、".join(seen)

            behavior_map = combined_df.groupby("证券代码", sort=False)["资本运作行为"].apply(_merge_behaviors)

            # 选出每个代码的"保留行"：最新公告日最大，平局按 _orig_idx 最小（先到先留）
            # NaT 在 max 里会排到最末——用 fillna(pd.Timestamp.min) 保证至少能选出一行
            combined_df["_date_for_rank"] = combined_df["_date_sort_key"].fillna(pd.Timestamp.min)
            keep_idx = (
                combined_df.sort_values(
                    by=["证券代码", "_date_for_rank", "_orig_idx"],
                    ascending=[True, False, True],
                )
                .drop_duplicates(subset="证券代码", keep="first")
                .index
            )
            combined_df = combined_df.loc[keep_idx].copy()
            # 应用合并后的资本运作行为
            combined_df["资本运作行为"] = combined_df["证券代码"].map(behavior_map)
            # 按原顺序恢复
            combined_df = combined_df.sort_values("_orig_idx").drop(
                columns=["_orig_idx", "_date_sort_key", "_date_for_rank"]
            ).reset_index(drop=True)
            after = len(combined_df)
            if before != after:
                logger.info(f"[条件交集] 按证券代码合并: {before}行 → {after}行")

        # 按"最新公告日"降序排序后再重新编号（NaT 排到最末）
        if "最新公告日" in combined_df.columns and len(combined_df) > 0:
            _date_key = pd.to_datetime(combined_df["最新公告日"], errors="coerce")
            combined_df = combined_df.assign(_date_key=_date_key.fillna(pd.Timestamp.min)) \
                .sort_values("_date_key", ascending=False, kind="stable") \
                .drop(columns="_date_key") \
                .reset_index(drop=True)

        # 重新编号序号
        combined_df["序号"] = range(1, len(combined_df) + 1)

        # 4. Sheet2（选股池）= Sheet1 数据的完整复制（当前产品决策：不做代码交集）

        # 列名重映射（原始数据不变，只改最终输出的列名）
        combined_display = combined_df.rename(columns=INTERSECTION_COLUMN_RENAME)
        # 确保列顺序；临时屏蔽"序号"列（内部计算逻辑保留，不展示）
        final_columns = [c for c in INTERSECTION_DISPLAY_COLUMNS if c in combined_display.columns and c != "序号"]
        combined_display = combined_display[final_columns]

        # === 标注百日新高：按配置的每个周期，输出 2 列（次数 + 日期列表）===
        high_price_periods = config.get("high_price_periods", []) or []
        if high_price_periods and "证券代码" in combined_display.columns:
            today_codes = set(combined_display["证券代码"].astype(str).str.strip()) if date_str else set()
            period_new_columns = []
            from services.pool_cache import get_date_codes_map as _get_map
            for p in high_price_periods:
                p_start = (p.get("start") or "").strip()
                p_end = (p.get("end") or "").strip()
                if not p_start or not p_end or p_start > p_end:
                    continue
                # DB 查 [start, min(end, today-1)]；今天用 combined_df 自身补
                db_end = p_end
                if date_str and p_start <= date_str <= p_end:
                    try:
                        from datetime import datetime as _dt, timedelta as _td
                        db_end = (_dt.strptime(date_str, "%Y-%m-%d").date() - _td(days=1)).strftime("%Y-%m-%d")
                    except Exception:
                        db_end = p_end

                code_dates: Dict[str, List[str]] = {}
                if p_start <= db_end:
                    date_codes_map = await _get_map(p_start, db_end)
                    for ds, codes in date_codes_map.items():
                        for code in codes:
                            code_dates.setdefault(code, []).append(ds)
                # 当天加入
                if date_str and p_start <= date_str <= p_end:
                    for code in today_codes:
                        code_dates.setdefault(code, []).append(date_str)

                col_count = f"{p_start}至{p_end}期间百日新高次数"
                col_dates = f"{p_start}至{p_end}期间百日新高的日期"
                count_values = []
                date_values = []
                for code in combined_display["证券代码"].astype(str).str.strip():
                    dates = sorted(set(code_dates.get(code, [])), reverse=True)
                    count_values.append(len(dates))
                    date_values.append(",".join(dates) if dates else "")
                combined_display[col_count] = count_values
                combined_display[col_dates] = date_values
                period_new_columns.extend([col_count, col_dates])
                logger.info(f"[条件交集] 标注百日新高周期 {p_start}至{p_end}: 历史 pool 命中代码 {len(code_dates)} 个 (via cache)")

        # Sheet2（选股池）= Sheet1 数据的完整复制
        pool_df = combined_display.copy()

        # 5. 输出 Excel（双 Sheet）
        daily_dir = self._get_daily_dir(date_str)
        os.makedirs(daily_dir, exist_ok=True)
        output_path = os.path.join(daily_dir, output_filename)

        # 生成选股池名称
        try:
            from datetime import datetime as dt
            d = dt.strptime(date_str, "%Y-%m-%d")
            pool_name = f"{d.year}年{d.month:02d}月选股池"
        except Exception:
            pool_name = f"选股池_{date_str}"

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            combined_display.to_excel(writer, sheet_name="条件交集", index=False)
            pool_df.to_excel(writer, sheet_name=pool_name, index=False)

        auto_adjust_excel_width(output_path, all_sheets=True)

        # 窗口基线：优先取首个"标注百日新高周期"的 [start, end]；否则默认最近 100 天
        first_period = None
        if high_price_periods:
            for p in high_price_periods:
                s = (p.get("start") or "").strip()
                e = (p.get("end") or "").strip()
                if s and e and s <= e:
                    first_period = (s, e)
                    break
        baseline_codes = await self._find_baseline_codes_in_window(
            pool_name=pool_name,
            current_date_str=date_str,
            window=first_period,  # None → 默认 100 天
        )

        # 资本运作行为可能是多个类型顿号拼接；统一设宽为 70 保证不换行；全表居中
        # 同时对 Sheet2 "选股池" 的新增行高亮（证券代码/证券简称/最新公告日 3 列绿底）
        try:
            from openpyxl import load_workbook as _lw
            from openpyxl.styles import Alignment, PatternFill
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
            green_fill = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
            highlight_cols = {"证券代码", "证券简称", "最新公告日"}
            wb_w = _lw(output_path)
            for ws in wb_w.worksheets:
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                header_list = list(header_row)
                # 列宽
                for idx, h in enumerate(header_list, start=1):
                    if not isinstance(h, str):
                        continue
                    if h == "资本运作行为":
                        ws.column_dimensions[get_column_letter(idx)].width = 70
                    elif "期间百日新高的日期" in h:
                        # 日期列最长可能含 30+ 个日期（YYYY-MM-DD,）
                        ws.column_dimensions[get_column_letter(idx)].width = 80
                    elif "期间百日新高次数" in h:
                        ws.column_dimensions[get_column_letter(idx)].width = 30
                # 居中
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                        min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = center_align
                # Sheet1 "条件交集" 和 Sheet2 "选股池" 都做新增高亮
                if baseline_codes is not None and ws.title in ("条件交集", pool_name):
                    try:
                        code_col_idx = header_list.index("证券代码") + 1
                    except ValueError:
                        code_col_idx = None
                    if code_col_idx:
                        target_col_indices = [
                            header_list.index(c) + 1 for c in header_list
                            if c in highlight_cols
                        ]
                        for r_idx in range(2, ws.max_row + 1):
                            code_val = ws.cell(row=r_idx, column=code_col_idx).value
                            if code_val is None:
                                continue
                            if str(code_val).strip() not in baseline_codes:
                                for c_idx in target_col_indices:
                                    ws.cell(row=r_idx, column=c_idx).fill = green_fill
            wb_w.save(output_path)
        except Exception as e:
            logger.warning(f"[条件交集] 列宽/居中/高亮设置失败: {e}")
        logger.info(f"[条件交集] 输出文件: {output_path}")

        # 6. 保存选股池到 DB
        try:
            pool_records = pool_df.fillna("").to_dict("records")
            await self._save_stock_pool(
                name=pool_name,
                date_str=date_str,
                file_path=output_path,
                data=pool_records,
                total_stocks=len(pool_df),
                filter_conditions=filter_conditions,
                source_types=list(filtered_dfs.keys()),
                workflow_id=workflow_id
            )
            logger.info(f"[条件交集] 选股池已保存: {pool_name}, {len(pool_df)}条")
        except Exception as e:
            logger.error(f"[条件交集] 保存选股池失败: {e}")

        base_msg = f"条件交集完成: 合并{len(combined_display)}行, 选股池{len(pool_df)}条, 来源{len(filtered_dfs)}个类型"
        if warnings_list:
            base_msg += f"（{len(warnings_list)}条警告）"
        return {
            "success": True,
            "message": base_msg,
            "data": clean_df_for_json(combined_display),
            "columns": combined_display.columns.tolist(),
            "rows": len(combined_display),
            "file_path": output_path,
            "_df": combined_display,
            "pool_count": len(pool_df),
            "warnings": warnings_list,
        }

    async def _export_ma20_trend(self, config: Dict, date_str: Optional[str] = None) -> Dict[str, Any]:
        """导出20日均线趋势步骤：复用统计分析的趋势导出逻辑（含 Excel 折线图）"""
        from services.trend_service import get_trend_data, export_trend_excel_with_chart

        date_str = config.get("date_str") or date_str or self.today
        output_filename = config.get("output_filename") or f"10站上20日均线趋势{date_str}.xlsx"
        # 兼容用户在 UI 里留下的 {date} 占位
        if "{date}" in output_filename:
            output_filename = output_filename.replace("{date}", date_str)
        date_preset = config.get("date_preset")
        date_range_start = config.get("date_range_start")
        date_range_end = config.get("date_range_end")

        # preset 非 custom 时以 date_str 为锚点重算范围，保证跟数据日期同步（忽略 DB 里旧值）
        if date_preset and date_preset != "custom":
            from datetime import datetime as _dt
            from dateutil.relativedelta import relativedelta
            try:
                anchor = _dt.strptime(date_str, "%Y-%m-%d")
                if date_preset == "1m":
                    start = anchor - relativedelta(months=1)
                elif date_preset == "6m":
                    start = anchor - relativedelta(months=6)
                elif date_preset == "1y":
                    start = anchor - relativedelta(years=1)
                else:
                    start = None
                if start is not None:
                    date_range_start = start.strftime("%Y-%m-%d")
                    date_range_end = anchor.strftime("%Y-%m-%d")
            except Exception as e:
                logger.warning(f"[导出MA20趋势] preset 范围计算失败，回退到 config: {e}")

        logger.info(f"[导出MA20趋势] date={date_str}, preset={date_preset}, range={date_range_start}~{date_range_end}")

        # 获取趋势数据（已自动排除条件交集和导出20日均线趋势类型）
        data = await get_trend_data(
            metric_type="ma20",
            start_date=date_range_start,
            end_date=date_range_end
        )

        if not data:
            return {"success": False, "message": "指定时间范围内无趋势数据"}

        # 生成带图表的 Excel
        daily_dir = self._get_daily_dir(date_str)
        os.makedirs(daily_dir, exist_ok=True)
        output_path = os.path.join(daily_dir, output_filename)

        export_trend_excel_with_chart(data, output_path)
        # 注意：xlsxwriter 生成的文件不用 openpyxl 的 auto_adjust_excel_width，会破坏图表
        logger.info(f"[导出MA20趋势] 输出文件: {output_path}, {len(data)}条数据")

        return {
            "success": True,
            "message": f"20日均线趋势导出完成，{len(data)}条数据，含折线图",
            "data": data[:100],
            "columns": list(data[0].keys()) if data else [],
            "rows": len(data),
            "file_path": output_path,
        }

    async def _export_high_price_trend(self, config: Dict, date_str: Optional[str] = None) -> Dict[str, Any]:
        """导出百日新高总趋势：扫当日文件 count→写DB→拉历史→出图。"""
        from services.trend_service import (
            count_high_price_rows, save_trend_data, get_trend_data,
            export_trend_excel_with_chart,
        )

        date_str = config.get("date_str") or date_str or self.today
        output_filename = config.get("output_filename") or f"11百日新高趋势图{date_str}.xlsx"
        # 兼容用户在 UI 里留下的 {date} 占位
        if "{date}" in output_filename:
            output_filename = output_filename.replace("{date}", date_str)
        date_preset = config.get("date_preset")
        date_range_start = config.get("date_range_start")
        date_range_end = config.get("date_range_end")

        if date_preset and date_preset != "custom":
            from datetime import datetime as _dt
            from dateutil.relativedelta import relativedelta
            try:
                anchor = _dt.strptime(date_str, "%Y-%m-%d")
                if date_preset == "1m":
                    start = anchor - relativedelta(months=1)
                elif date_preset == "6m":
                    start = anchor - relativedelta(months=6)
                elif date_preset == "1y":
                    start = anchor - relativedelta(years=1)
                else:
                    start = None
                if start is not None:
                    date_range_start = start.strftime("%Y-%m-%d")
                    date_range_end = anchor.strftime("%Y-%m-%d")
            except Exception as e:
                logger.warning(f"[导出百日新高趋势] preset 范围计算失败，回退: {e}")

        count = count_high_price_rows(self.base_dir, date_str)
        if count > 0:
            ok = await save_trend_data(
                metric_type="high_price",
                workflow_type="百日新高总趋势",
                date_str=date_str,
                count=count,
                total=0,
                source="auto",
            )
            if not ok:
                logger.error(f"[导出百日新高趋势] {date_str} DB 写入失败，仍继续出图")
        else:
            logger.info(f"[导出百日新高趋势] {date_str} count=0 已跳过入库")

        data = await get_trend_data(
            metric_type="high_price",
            start_date=date_range_start,
            end_date=date_range_end,
        )

        output_dir = os.path.join(self.base_dir, "百日新高总趋势", date_str)
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_filename)
        export_trend_excel_with_chart(data, output_path, metric_type="high_price")
        logger.info(f"[导出百日新高趋势] 输出文件: {output_path}, 历史{len(data)}条，当日count={count}")

        return {
            "success": True,
            "message": f"百日新高总趋势导出完成：当日{count} + 历史{len(data)}条",
            "data": data[:100],
            "columns": list(data[0].keys()) if data else [],
            "rows": len(data),
            "file_path": output_path,
            "today_count": count,
        }

    def _apply_filters(self, df: pd.DataFrame, filter_conditions: list, filter_logic: str = "AND") -> pd.DataFrame:
        """应用过滤条件：非空即保留"""
        SECTOR_ALIASES = ['所属板块', '一级板块', '所属一级板块', '板块', '二级板块', '所属二级板块']

        enabled = [f for f in filter_conditions if f.get("enabled")]
        if not enabled:
            return df

        masks = []
        for f in enabled:
            col = f["column"]
            actual_col = None
            if col in df.columns:
                actual_col = col
            elif col in SECTOR_ALIASES:
                for alias in SECTOR_ALIASES:
                    if alias in df.columns:
                        actual_col = alias
                        break
            if actual_col is not None:
                mask = df[actual_col].fillna("").astype(str).str.strip() != ""
                masks.append(mask)

        if not masks:
            return df

        if filter_logic == "OR":
            combined = masks[0]
            for m in masks[1:]:
                combined = combined | m
        else:  # AND
            combined = masks[0]
            for m in masks[1:]:
                combined = combined & m

        return df[combined].copy()

    async def _find_baseline_codes_in_window(
        self,
        pool_name: str,
        current_date_str: str,
        window: Optional[tuple] = None,
    ) -> Optional[set]:
        """返回窗口期内所有历史 pool 的证券代码并集，用于"新增"高亮基线。

        走 pool_cache 进程内缓存（10 分钟刷新，入库时失效）。
        window: (start_str, end_str) 闭区间；None → 默认最近 100 天
        窗口严格 < current_date_str（排除当天）。
        找不到任何历史记录 → 返回 None（跳过高亮）。
        """
        from datetime import datetime as _dt, timedelta as _td
        from services.pool_cache import get_codes_union

        try:
            base_date = _dt.strptime(current_date_str, "%Y-%m-%d").date()
        except Exception as e:
            logger.warning(f"[条件交集-高亮] 解析 current_date_str 失败: {e}")
            return None

        yesterday = (base_date - _td(days=1)).strftime("%Y-%m-%d")
        if window is None:
            start_str = (base_date - _td(days=100)).strftime("%Y-%m-%d")
            end_str = yesterday
        else:
            start_str, end_str = window
            if end_str >= current_date_str:
                end_str = yesterday
        if start_str > end_str:
            logger.info(f"[条件交集-高亮] 窗口无效 ({start_str}~{end_str})，跳过高亮")
            return None

        codes_union = await get_codes_union(start_str, end_str)
        logger.info(
            f"[条件交集-高亮] baseline 窗口 {start_str}~{end_str} 代码 {len(codes_union)} 个 (via cache)"
        )
        return codes_union if codes_union else None

    async def _save_stock_pool(self, name: str, date_str: str, file_path: str,
                                data: list, total_stocks: int,
                                filter_conditions: list, source_types: list,
                                workflow_id: int = None):
        """保存选股池到数据库（atomic upsert by name + date_str）"""
        import json as json_mod
        from core.database import AsyncSessionLocal
        from sqlalchemy import text

        params = {
            "name": name,
            "workflow_id": workflow_id,
            "date_str": date_str,
            "file_path": file_path,
            "total_stocks": total_stocks,
            "data": json_mod.dumps(data, ensure_ascii=False, default=str),
            "filter_conditions": json_mod.dumps(filter_conditions, ensure_ascii=False),
            "source_types": json_mod.dumps(source_types, ensure_ascii=False),
        }

        async with AsyncSessionLocal() as session:
            await session.execute(text("""
                INSERT INTO stock_pools (name, workflow_id, date_str, file_path, total_stocks, data,
                    filter_conditions, source_types, is_active, created_at, updated_at)
                VALUES (:name, :workflow_id, :date_str, :file_path, :total_stocks, :data,
                    :filter_conditions, :source_types, 1, NOW(), NOW())
                AS new_row
                ON DUPLICATE KEY UPDATE
                    workflow_id = new_row.workflow_id,
                    file_path = new_row.file_path,
                    total_stocks = new_row.total_stocks,
                    data = new_row.data,
                    filter_conditions = new_row.filter_conditions,
                    source_types = new_row.source_types,
                    is_active = 1,
                    updated_at = NOW()
            """), params)
            await session.commit()
        # 失效缓存
        try:
            from services.pool_cache import invalidate as _inv
            _inv()
        except Exception:
            pass

    async def _ranking_sort(self, config: Dict, input_data: Optional[pd.DataFrame] = None,
                            date_str: Optional[str] = None) -> Dict[str, Any]:
        """涨幅排名排序步骤：排序→排名→历史对比→格式化输出→自动复制到public"""
        import shutil
        import re as re_mod
        from collections import OrderedDict
        from openpyxl.styles import PatternFill, Font, Alignment as openpyxl_Alignment

        date_str = config.get("date_str") or date_str or self.today
        logger.info(f"[涨幅排名] 开始执行: date={date_str}")

        # 1. 获取输入数据
        df = input_data
        if df is None or df.empty:
            return {"success": False, "message": "无输入数据，请先执行合并步骤"}

        cols = df.columns.tolist()
        if len(cols) < 2:
            return {"success": False, "message": f"输入数据至少需要2列，当前只有{len(cols)}列"}

        sector_col = cols[0]  # 板块名称（第1列）

        # 按"表头内容"识别三列（避免依赖列位置——有些源的列顺序是
        # [板块名, 本年初, 本月初, 今日]，有些是 [板块名, 今日, 本年初, 本月初]）：
        # - 今日涨跌幅：表头 **不含** "本年初" / "本月初"（通常含具体 YYYY-MM-DD 或"最新"）
        # - 年初涨跌幅：表头含 "本年初"
        # - 月初涨跌幅：表头含 "本月初"
        def _col_contains(col, key: str) -> bool:
            return key in str(col).replace("\n", "").replace("\r", "").replace(" ", "").replace("\t", "")

        ytd_col_raw = None  # 年初涨跌幅 原始列
        mtd_col_raw = None  # 月初涨跌幅 原始列
        today_col_raw = None  # 今日涨跌幅 原始列
        for c in cols[1:]:
            if ytd_col_raw is None and _col_contains(c, "本年初"):
                ytd_col_raw = c
                continue
            if mtd_col_raw is None and _col_contains(c, "本月初"):
                mtd_col_raw = c
                continue
            if today_col_raw is None:
                today_col_raw = c

        # 兜底：若未识别到今日列（例如只有 2 列输入），退回到第 2 列
        if today_col_raw is None:
            today_col_raw = cols[1]

        sort_col_raw = today_col_raw  # 排序主列——今日涨跌幅
        # 多行列名只取第一行（如 "成份区间涨跌幅(算术平均)\n[起始交易日期]..." → "成份区间涨跌幅(算术平均)"）
        # 按用户要求，展示名固定为"今日涨跌幅"（下游所有逻辑按 sort_col 变量处理，换名不影响）
        sort_col_display = "今日涨跌幅"
        def _col_head(c):
            return str(c).split("\n")[0].strip() if c is not None else None
        logger.info(
            "[涨幅排名] 板块列=%s, 今日列=%s, 年初列=%s, 月初列=%s",
            sector_col, _col_head(today_col_raw), _col_head(ytd_col_raw), _col_head(mtd_col_raw)
        )

        use_extended = ytd_col_raw is not None and mtd_col_raw is not None
        if use_extended:
            logger.info(f"[涨幅排名] 探测到年初列 + 月初列，启用扩展输出布局")
        else:
            logger.info(f"[涨幅排名] 未探测到「本年初」或「本月初」列，使用原输出布局")

        # 2. 提取板块名称和排序列，过滤空行，转数值排序
        work_df = df[[sector_col, sort_col_raw]].copy()
        work_df = work_df.rename(columns={sort_col_raw: sort_col_display})
        sort_col = sort_col_display

        # 2b. 扩展布局：同时抽年初 / 月初 原始值（保留原字符串，后续单独算数值排名）
        YTD_COL = "年初涨跌幅"
        MTD_COL = "月初涨跌幅"
        YTD_RANK_COL = "B列的数值升序排序结果"
        MTD_RANK_COL = "D列的数值升序排序结果"
        if use_extended:
            work_df[YTD_COL] = df[ytd_col_raw].values
            work_df[MTD_COL] = df[mtd_col_raw].values

        # 过滤板块名称为空/NaN/含"妙想Choice"的行
        sector_str = work_df[sector_col].astype(str).str.strip()
        work_df = work_df[
            work_df[sector_col].notna()
            & (sector_str != '')
            & (~sector_str.str.contains('妙想Choice', na=False, regex=False))
        ]
        work_df[sort_col] = pd.to_numeric(work_df[sort_col], errors='coerce')

        # 有效数字行：降序排列；非数字行（--等）：排到末尾
        valid_df = work_df.dropna(subset=[sort_col]).copy()
        valid_df[sort_col] = valid_df[sort_col].round(2)
        valid_df = valid_df.sort_values(by=sort_col, ascending=False).reset_index(drop=True)

        invalid_df = work_df[work_df[sort_col].isna()].copy()
        invalid_df[sort_col] = df.loc[invalid_df.index, sort_col_raw]  # 索引与 df 对齐（未 reset）

        work_df = pd.concat([valid_df, invalid_df], ignore_index=True)

        if work_df.empty:
            return {"success": False, "message": "过滤后无有效数据（所有行为空/妙想Choice或排序列无数值）"}

        # 2b2. 年初/月初数值列统一保留 2 位小数（非数字原样保留）
        if use_extended:
            def _round2_col(series: pd.Series) -> pd.Series:
                """可解析为数字的位置 round 到 2 位；其他保持原值。"""
                numeric = pd.to_numeric(series, errors='coerce')
                result = series.copy()
                mask = numeric.notna()
                result.loc[mask] = numeric.loc[mask].round(2)
                return result

            work_df[YTD_COL] = _round2_col(work_df[YTD_COL])
            work_df[MTD_COL] = _round2_col(work_df[MTD_COL])

        # 2c. 扩展布局：为 YTD / MTD 计算降序排名（大的排前，非数字按出现顺序占末尾固定位）
        if use_extended:
            def _compute_rank(col_name: str) -> list:
                """
                返回 work_df 每行对应的排名整数。
                有效数字 → 按降序排名（数值大排名小），method='min' 同值并列取最小名次
                非数字（-- / 空 / 非法）→ 取末尾固定位: valid_n+1, valid_n+2, ...（按 work_df 行顺序）
                """
                series = pd.to_numeric(work_df[col_name], errors='coerce')
                valid_mask = series.notna()
                ranks = series.rank(ascending=False, method='min')  # 含 NaN 行，结果也是 NaN
                valid_count = int(valid_mask.sum())
                result = []
                next_invalid = valid_count + 1
                for i in range(len(work_df)):
                    if valid_mask.iloc[i]:
                        result.append(int(ranks.iloc[i]))
                    else:
                        result.append(next_invalid)
                        next_invalid += 1
                return result

            work_df[YTD_RANK_COL] = _compute_rank(YTD_COL)
            work_df[MTD_RANK_COL] = _compute_rank(MTD_COL)

        # 3. 列名规范化函数 + 当日日期列名
        import re as re_mod
        _col_date_map = {}  # col_name → datetime, 用于后续 Excel 表头写入实际日期值

        def normalize_col_name(c):
            """Excel 日期对象/序列号/文本 → 'X月X日' 格式，同时记录原始日期"""
            dt_val = None
            if isinstance(c, (pd.Timestamp, datetime)):
                dt_val = c if isinstance(c, datetime) else c.to_pydatetime()
            elif isinstance(c, (int, float)) and 40000 <= c <= 50000:
                try:
                    dt_val = (pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(c))).to_pydatetime()
                except Exception:
                    pass
            elif isinstance(c, str):
                m = re_mod.match(r'^(\d+)月(\d+)日$', c)
                if m:
                    month, day = int(m.group(1)), int(m.group(2))
                    year = d.year if month <= d.month else d.year - 1
                    dt_val = datetime(year, month, day)

            if dt_val:
                name = f"{dt_val.month}月{dt_val.day}日"
                _col_date_map[name] = dt_val
                return name
            return str(c)

        d = datetime.strptime(date_str, "%Y-%m-%d")
        today_col = f"{d.month}月{d.day}日"
        _col_date_map[today_col] = d

        # 4. 查找上一个工作日的 public 文件（带 mtime 缓存）
        prev_file, prev_date = self.resolver.find_previous_public_file(date_str)
        prev_data = {}  # {板块名称: {top5_count, prev_rank, history}}
        prev_history_cols = []
        prev_all_sheets = None  # {sheet_name: DataFrame} 用于多 Sheet 输出

        if prev_file and os.path.exists(prev_file):
            try:
                # mtime 缓存: key=完整文件路径, 文件更新/覆盖后 mtime 变化自动失效
                # 缓存存储 (mtime, {sheet_name: DataFrame}) 格式，涨幅排名专用
                cache_key = f"{prev_file}::all_sheets"
                current_mtime = os.path.getmtime(prev_file)
                cached = _public_file_cache.get(cache_key)
                if cached and cached[0] == current_mtime:
                    prev_all_sheets = {k: v.copy() for k, v in cached[1].items()}
                    logger.info(f"[涨幅排名] 缓存命中: {os.path.basename(prev_file)} ({len(prev_all_sheets)} sheets)")
                else:
                    prev_all_sheets = pd.read_excel(prev_file, sheet_name=None, engine="openpyxl")
                    # 缓存淘汰：超过上限时移除最早的条目
                    if len(_public_file_cache) >= _PUBLIC_FILE_CACHE_MAX:
                        oldest_key = next(iter(_public_file_cache))
                        del _public_file_cache[oldest_key]
                    _public_file_cache[cache_key] = (current_mtime, prev_all_sheets)
                    logger.info(f"[涨幅排名] 加载并缓存: {os.path.basename(prev_file)} ({len(prev_all_sheets)} sheets)")

                # 用第一个 sheet 提取排名数据
                first_sheet_name = list(prev_all_sheets.keys())[0]
                prev_raw_df = prev_all_sheets[first_sheet_name]
                # 列名规范化
                prev_cols = [normalize_col_name(c) for c in prev_raw_df.columns.tolist()]
                prev_raw_df.columns = prev_cols
                # 列结构: [板块名称, 排序列, 迄今为止排进前5(次数), 上一日期列, ...]
                if len(prev_cols) >= 4:
                    prev_history_cols = prev_cols[3:]  # 从第4列开始都是历史日期列
                    for _, row in prev_raw_df.iterrows():
                        name = str(row[prev_cols[0]]).strip()
                        top5_count = int(row[prev_cols[2]]) if pd.notna(row[prev_cols[2]]) else 0
                        prev_rank = int(row[prev_cols[3]]) if pd.notna(row[prev_cols[3]]) else None
                        history = {c: row[c] for c in prev_history_cols}
                        prev_data[name] = {
                            "top5_count": top5_count,
                            "prev_rank": prev_rank,
                            "history": history,
                        }
                logger.info(f"[涨幅排名] 历史数据: {prev_file}, {len(prev_data)}个板块")
            except Exception as e:
                logger.warning(f"[涨幅排名] 读取上一工作日文件失败: {e}")

        # 5. 构建输出
        records = []
        for idx, row in work_df.iterrows():
            rank = idx + 1
            sector = str(row[sector_col]).strip()
            sort_val = row[sort_col]

            prev = prev_data.get(sector, {"top5_count": 0, "prev_rank": None, "history": {}})

            # 迄今为止排进前5(次数)
            if rank <= 5:
                new_top5 = prev["top5_count"] + 1
            else:
                new_top5 = prev["top5_count"]

            record = OrderedDict()
            record[sector_col] = sector
            if use_extended:
                # 扩展布局: 板块 | 年初 | B排名 | 月初 | D排名 | 今日 | Top5次数 | 日期列...
                record[YTD_COL] = row[YTD_COL]
                record[YTD_RANK_COL] = row[YTD_RANK_COL]
                record[MTD_COL] = row[MTD_COL]
                record[MTD_RANK_COL] = row[MTD_RANK_COL]
                record[sort_col] = sort_val
                record["迄今为止排进前5(次数)"] = new_top5
                record[today_col] = rank
            else:
                # 原布局: 板块 | 今日 | Top5次数 | 日期列...
                record[sort_col] = sort_val
                record["迄今为止排进前5(次数)"] = new_top5
                record[today_col] = rank

            # 附加历史列
            for hcol in prev_history_cols:
                record[hcol] = prev["history"].get(hcol, "")

            records.append(record)

        result_df = pd.DataFrame(records)

        # 6. 生成文件名
        # 从上一工作日文件最后一列提取起始日期
        start_date_str = ""
        all_cols = result_df.columns.tolist()
        if len(all_cols) > 4:
            last_col = str(all_cols[-1])  # 最后一列 = 最早的日期列，转字符串防止整数列名
            m = re_mod.search(r'(\d+)月(\d+)日', last_col)
            if m:
                start_date_str = f"{int(m.group(1)):02d}{int(m.group(2)):02d}"
        if not start_date_str:
            start_date_str = f"{d.month:02d}{d.day:02d}"

        date_no_hyphens = date_str.replace("-", "")
        default_filename = f"8涨幅排名{start_date_str}-{date_no_hyphens}.xlsx"
        output_filename = config.get("output_filename") or default_filename
        if not output_filename.endswith(".xlsx"):
            output_filename += ".xlsx"

        daily_dir = self.resolver.get_upload_directory(date_str)
        os.makedirs(daily_dir, exist_ok=True)
        output_path = os.path.join(daily_dir, output_filename)

        # 7. 保存 Excel（多 Sheet: 当日结果 + 上一工作日文件的所有 sheet）并格式化
        today_sheet = f"{d.month:02d}{d.day:02d}"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            result_df.to_excel(writer, sheet_name=today_sheet, index=False)

        wb = load_workbook(output_path)

        DEEP_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        DEEP_RED_FONT = Font(color="FFFFFF", bold=True)
        LIGHT_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        LIGHT_RED_FONT = Font(color="FFFFFF")
        GOLD_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        from services.ranking_format import apply_ranking_format

        # --- 当日 sheet ---
        ws_today = wb[today_sheet]
        data_row_count = len(result_df) + 1

        # prev_rank_by_sector：从 prev_data 提取板块 → 上一日排名（用于当日列浅红）
        prev_rank_by_sector = {
            name: (d.get("prev_rank") if isinstance(d, dict) else None)
            for name, d in prev_data.items()
        }
        apply_ranking_format(
            ws_today,
            prev_rank_by_sector=prev_rank_by_sector,
            date_col_date_map=_col_date_map,
        )

        wb.save(output_path)

        logger.info(f"[涨幅排名] 输出文件: {output_path}")

        # 8. 复制到当日 public 目录（不清空旧文件，同名覆盖）
        public_dir = self.resolver.get_public_directory(date_str)
        os.makedirs(public_dir, exist_ok=True)
        shutil.copy2(output_path, public_dir)
        logger.info(f"[涨幅排名] 已复制到 public: {public_dir}")

        # 9. 返回结果
        clean_df = result_df.fillna("").copy()
        preview = clean_df.head(100).to_dict("records")

        return {
            "success": True,
            "message": f"涨幅排名完成，{len(result_df)}个板块，前5名已标红",
            "data": preview,
            "columns": result_df.columns.tolist(),
            "rows": len(result_df),
            "file_path": output_path,
            "_df": result_df,
        }

    async def _pledge_trend_analysis(
        self, config: Dict, df: Optional[pd.DataFrame], date_str: Optional[str] = None
    ) -> Dict[str, Any]:
        """质押异动和趋势步骤：为每只股票查询东财 1 年质押历史，判定 3 列。

        Required: df 必须含 证券代码 / 最新公告日。锚点缺失任一行 → 整步骤失败。
        """
        from services.pledge_data_source import PledgeDataSource
        from services.pledge_trend import compute_trend
        from services.pledge_cache_cleanup import cleanup_expired_pledge_cache
        from core.redis_client import get_redis
        from utils.stock_code_normalizer import normalize_stock_code

        # 1. 校验输入
        if df is None or "证券代码" not in df.columns:
            return {"success": False, "message": "需要包含'证券代码'列的输入数据"}
        if "最新公告日" not in df.columns:
            return {"success": False, "message":
                "质押异动趋势步骤失败：输入数据缺少'最新公告日'列（或前缀'股权质押公告日期'列未被正确映射）"}

        df = df.copy()
        # 兜底过滤脏数据行（证券代码为空/数据来源脚注/Choice 元信息等）
        code_str = df["证券代码"].astype(str).str.strip()
        dirty_mask = (
            df["证券代码"].isna()
            | (code_str == "")
            | code_str.str.lower().eq("nan")
            | code_str.str.contains("数据来源", na=False, regex=False)
            | code_str.str.contains("Choice", na=False, regex=False)
        )
        dirty_count = int(dirty_mask.sum())
        if dirty_count > 0:
            logger.info(f"[质押异动趋势] 过滤脏数据行 {dirty_count} 条")
            df = df[~dirty_mask].reset_index(drop=True)
        if len(df) == 0:
            return {"success": False, "message": "质押异动趋势：过滤脏数据后无可处理行"}

        # 锚点严格校验：任何行缺失即整步骤失败
        anchor_series = df["最新公告日"].astype(str).str.strip()
        missing_mask = df["最新公告日"].isna() | (anchor_series == "") | (anchor_series.str.lower() == "nan")
        if missing_mask.any():
            missing_cnt = int(missing_mask.sum())
            sample_cols = [c for c in ["证券代码", "证券简称"] if c in df.columns]
            sample = df.loc[missing_mask, sample_cols].head(5).to_dict("records")
            return {"success": False, "message":
                f"质押异动趋势步骤失败：{missing_cnt} 行缺少'最新公告日'锚点，无法执行。示例: {sample}"}

        # 2. 读取配置
        trend_algo       = config.get("trend_algo", "mann_kendall")
        mk_p             = float(config.get("mk_pvalue", 0.05))
        b_rev            = int(config.get("b_max_reversals", 2))
        c_r2             = float(config.get("c_min_r2", 0.7))
        event_no_change  = float(config.get("event_no_change_threshold", 0.5))
        event_large     = float(config.get("event_large_threshold", 3.0))
        window_days      = int(config.get("window_days", 365))
        row_recency_days = int(config.get("row_recency_days", 30))

        # 3. 初始化数据源
        redis_cli = get_redis()
        ds = PledgeDataSource(redis_client=redis_cli, akshare_fallback=True)

        # 4. 逐股处理：预判列不存在则新建；存在则保留原值
        for preset in ["持续递增（一年内）", "持续递减（一年内）", "质押异动"]:
            if preset not in df.columns:
                df[preset] = ""
        # 计算行有效期的 cutoff（anchor 早于 cutoff 的行 skip）
        from datetime import datetime as _dt, timedelta as _td
        today = _dt.now().date()
        recency_cutoff = (today - _td(days=row_recency_days)).isoformat()

        # total = 实际要查询的行数（扣除 skip），input_total = 输入行数
        stats = {
            "total": 0, "input_total": len(df),
            "ok": 0, "empty": 0, "fail": 0,
            "skipped_preset": 0, "skipped_old": 0,
            "by_source": {"eastmoney": 0, "cache": 0, "akshare": 0, "empty": 0},
            "by_result": {
                "持续递增": 0, "持续递减": 0, "无趋势": 0,
                "小幅转增": 0, "小幅转减": 0, "大幅激增": 0, "大幅骤减": 0,
                "本次质押趋势无变化": 0, "空": 0,
            },
        }
        fail_samples: List[Dict[str, str]] = []

        import re as _re_mod

        def _nonempty(v):
            if v is None:
                return False
            s = str(v).strip()
            return s != "" and s.lower() != "nan"

        for idx, row in df.iterrows():
            raw_code = str(row["证券代码"])
            symbol = normalize_stock_code(raw_code)
            # 取纯数字部分（东财 SECURITY_CODE 是 6 位数字）
            m = _re_mod.search(r'(\d{6})', symbol)
            numeric = m.group(1) if m else symbol
            anchor = str(row["最新公告日"]).strip()[:10]

            # 跳过规则 1：原表 3 列任一非空 → 整行跳过
            has_preset = (
                _nonempty(row.get("持续递增（一年内）"))
                or _nonempty(row.get("持续递减（一年内）"))
                or _nonempty(row.get("质押异动"))
            )
            if has_preset:
                stats["skipped_preset"] += 1
                logger.info(
                    f"[质押异动趋势] {idx+1}/{len(df)} {symbol} {row.get('证券简称','')} "
                    f"原表已有值 → skip（递增={row.get('持续递增（一年内）') or '-'} "
                    f"递减={row.get('持续递减（一年内）') or '-'} "
                    f"异动={row.get('质押异动') or '-'}）"
                )
                continue

            # 跳过规则 2：行日期早于 row_recency_days 前 → 整行跳过
            if anchor < recency_cutoff:
                stats["skipped_old"] += 1
                logger.info(
                    f"[质押异动趋势] {idx+1}/{len(df)} {symbol} {row.get('证券简称','')} "
                    f"锚点 {anchor} 早于 {recency_cutoff} → skip（超出{row_recency_days}天行有效期）"
                )
                continue

            try:
                stats["total"] += 1
                records, source = ds.get_history(numeric, anchor, window_days)
                stats["by_source"][source] = stats["by_source"].get(source, 0) + 1
                result = compute_trend(
                    records, anchor, trend_algo, mk_p, b_rev, c_r2,
                    event_no_change, event_large
                )
                df.at[idx, "持续递增（一年内）"] = result["持续递增（一年内）"]
                df.at[idx, "持续递减（一年内）"] = result["持续递减（一年内）"]
                df.at[idx, "质押异动"] = result["质押异动"]
                if records:
                    stats["ok"] += 1
                else:
                    stats["empty"] += 1
                if result["持续递增（一年内）"] == "Y":
                    stats["by_result"]["持续递增"] += 1
                elif result["持续递减（一年内）"] == "Y":
                    stats["by_result"]["持续递减"] += 1
                else:
                    stats["by_result"]["无趋势"] += 1
                event_key = result["质押异动"] or "空"
                stats["by_result"][event_key] = stats["by_result"].get(event_key, 0) + 1

                logger.info(
                    f"[质押异动趋势] {idx+1}/{len(df)} {symbol} {row.get('证券简称','')} "
                    f"锚点={anchor} 源={source} 历史{len(records)}条 → "
                    f"递增={result['持续递增（一年内）'] or '-'} "
                    f"递减={result['持续递减（一年内）'] or '-'} "
                    f"异动={result['质押异动'] or '-'}"
                )
            except Exception as e:
                stats["fail"] += 1
                if len(fail_samples) < 10:
                    fail_samples.append({"symbol": symbol, "error": str(e)[:120]})
                logger.warning(f"[质押异动趋势] {symbol} 失败: {e}")

        # 5. 覆盖最终输出（复用 match_sector 的文件名）
        user_filename = config.get("output_filename")
        output_filename = self.resolver.get_output_filename(
            "match_sector", date_str, user_filename
        )
        output_path = os.path.join(self._get_daily_dir(date_str), output_filename)
        df.to_excel(output_path, index=False)
        try:
            auto_adjust_excel_width(output_path)
        except Exception as e:
            logger.warning(f"[质押异动趋势] 设置列宽失败: {e}")

        # 7. 缓存清理兜底
        try:
            cleanup_expired_pledge_cache(redis_cli, max_age_days=370)
        except Exception as e:
            logger.warning(f"[质押异动趋势] 缓存清理失败（不影响主流程）: {e}")

        summary = (
            f"质押异动趋势: 输入 {stats['input_total']} 只 · "
            f"实际查询 {stats['total']} 只 · "
            f"成功 {stats['ok']} · 无历史 {stats['empty']} · 失败 {stats['fail']} · "
            f"跳过(原表已有) {stats['skipped_preset']} · "
            f"跳过(超出{row_recency_days}天) {stats['skipped_old']}   ｜   "
            f"数据源: 东财 {stats['by_source']['eastmoney']} · "
            f"缓存 {stats['by_source']['cache']} · "
            f"降级 AkShare {stats['by_source']['akshare']} · "
            f"空 {stats['by_source']['empty']}"
        )
        logger.info(f"[质押异动趋势] {summary}")

        df_clean = df.fillna("").copy()
        records_preview = df_clean.head(100).to_dict("records")
        return {
            "success": True,
            "message": summary,
            "stats": stats,
            "fail_samples": fail_samples,
            "data": records_preview,
            "columns": df.columns.tolist(),
            "rows": len(df),
            "file_path": output_path,
            "_df": df,
        }


workflow_executor = WorkflowExecutor()
