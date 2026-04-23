"""7_1 样式选股池 Excel 解析器（简化版）。

假设每段子表头含"资本运作行为"列（用户已手工规范）。按列名识别，不做位置映射。

支持两种 sheet 样式：
- 纵向多段（"0420 型"）：一段接一段，每段以"证券代码"子表头起始。
- 横向并排多段（"0318 型"）：分组标题行下面是多段并排子表头。

关键字段识别（按 header 文本）：
- 证券代码 / 代码
- 证券简称 / 证券名称 / 简称 / 名称
- 最新公告日 / 公告日 / 发生日期 / 股权质押公告日期 / 最新大股东减持公告日期
- 百日新高
- 站上20日线 / 占20日均线
- 所属板块
- 国央企
- 资本运作行为

某段以"无"代替"证券代码"等表头 → 整段跳过。
缺少的字段（如某段没有"国央企"）→ 该字段留空。
"""

from __future__ import annotations

import logging
import re
from datetime import datetime, date
from typing import Dict, List, Optional, Any

import pandas as pd
from openpyxl import load_workbook

from utils.beijing_time import BEIJING_TZ

logger = logging.getLogger(__name__)


# 字段 → 可能的 header 文本（第一个字符串命中为主）
FIELD_HEADER_KEYWORDS: Dict[str, List[str]] = {
    "证券代码": ["证券代码", "代码"],
    "证券简称": ["证券简称", "证券名称", "简称", "名称"],
    "最新公告日": ["最新公告日", "股权质押公告日期", "最新大股东减持公告日期", "发生日期", "公告日期", "公告日"],
    "百日新高": ["百日新高"],
    "站上20日线": ["站上20日线", "占20日均线", "20日均线"],
    "所属板块": ["所属板块", "一级板块", "板块"],
    "国央企": ["国央企", "国企"],
    "资本运作行为": ["资本运作行为"],
}

OUTPUT_FIELDS = list(FIELD_HEADER_KEYWORDS.keys())


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).replace("\n", "").strip()


def _is_code_header_row(row: List[Any]) -> bool:
    """子表头行：至少一个单元格归一化后等于'证券代码'或'代码'。"""
    for c in row:
        n = _norm(c)
        if n == "证券代码" or n == "代码":
            return True
    return False


def _is_skip_segment_header(row_slice: List[Any]) -> bool:
    """段子表头内只有'无'，没有'证券代码/代码' → 该段为空，跳过。"""
    has_code = False
    has_nothing = False
    for c in row_slice:
        n = _norm(c)
        if n == "证券代码" or n == "代码":
            has_code = True
        if n == "无":
            has_nothing = True
    return has_nothing and not has_code


def _map_columns(header_slice: List[Any], col_start: int) -> Dict[str, int]:
    """对给定段（列范围 header_slice 从 col_start 开始）返回字段 → 绝对列 idx。"""
    result: Dict[str, int] = {}
    for rel_idx, cell in enumerate(header_slice):
        h = _norm(cell)
        if not h:
            continue
        for field, keywords in FIELD_HEADER_KEYWORDS.items():
            if field in result:
                continue
            # 精确匹配 or 含关键字
            if h in keywords:
                result[field] = col_start + rel_idx
                break
            # 模糊：header 包含关键字（例"股权质押公告日期[截止日期]最新" 命中 "股权质押公告日期"）
            for kw in keywords:
                if kw in h:
                    result[field] = col_start + rel_idx
                    break
            if field in result:
                break
    return result


def _parse_sheet_name_to_date(sheet_name: str, year: int) -> Optional[date]:
    m = re.fullmatch(r"(\d{2})(\d{2})", sheet_name.strip())
    if not m:
        return None
    month, day = int(m.group(1)), int(m.group(2))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _normalize_date_value(v: Any) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)) and 30000 <= v <= 60000:
        try:
            return (pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(v))).strftime("%Y-%m-%d")
        except Exception:
            return ""
    s = str(v).strip()
    if not s:
        return ""
    try:
        return pd.to_datetime(s).strftime("%Y-%m-%d")
    except Exception:
        return ""


def _extract_records(data_rows: List[List[Any]], field_to_col: Dict[str, int]) -> List[Dict[str, str]]:
    """按字段-列映射抽取数据行。代码非空 + 形状合理才入库。"""
    records: List[Dict[str, str]] = []
    code_col = field_to_col.get("证券代码")
    if code_col is None:
        return records
    for row in data_rows:
        if code_col >= len(row):
            continue
        code = _norm(row[code_col])
        if not code:
            continue
        if not re.fullmatch(r"[\dA-Za-z\.]{3,15}", code):
            continue
        if code == "证券代码" or code == "代码":
            continue
        rec: Dict[str, str] = {}
        for field in OUTPUT_FIELDS:
            col = field_to_col.get(field)
            val = ""
            if col is not None and col < len(row):
                v = row[col]
                if field == "最新公告日":
                    val = _normalize_date_value(v)
                else:
                    val = _norm(v)
            rec[field] = val
        records.append(rec)
    return records


# ------------------------- 纵向 -------------------------

def _parse_vertical_sheet(raw_rows: List[List[Any]]) -> List[Dict[str, str]]:
    """纵向：找所有"证券代码"子表头，相邻表头之间是数据段。"""
    all_recs: List[Dict[str, str]] = []
    header_idxs: List[int] = [i for i, r in enumerate(raw_rows) if _is_code_header_row(r)]
    if not header_idxs:
        return all_recs
    for n, start_i in enumerate(header_idxs):
        end_i = header_idxs[n + 1] if n + 1 < len(header_idxs) else len(raw_rows)
        header = raw_rows[start_i]
        data_rows = raw_rows[start_i + 1:end_i]
        # 段实际列范围：header 最远非空列
        col_end = -1
        for col in range(len(header) - 1, -1, -1):
            if _norm(header[col]) != "":
                col_end = col
                break
        if col_end < 0:
            continue
        field_to_col = _map_columns(header[:col_end + 1], 0)
        all_recs.extend(_extract_records(data_rows, field_to_col))
    return all_recs


# ------------------------- 横向 -------------------------

def _is_group_title_row(row: List[Any]) -> bool:
    """分组标题行：至少 2 个非空单元格，且没有一个是"证券代码"/"代码"/"证券简称"等关键字段表头。

    另外：单元格值应看起来像"段名"（短文本、不是日期、不是代码形状）——
    用简单启发：非空 cell 值不含 '.'（排除代码）、不全是数字（排除日期 serial）、
    不以 'YYYY-' 或 '20' 开头（排除日期字符串）。
    """
    nonempty_vals: List[str] = []
    for c in row:
        n = _norm(c)
        if not n:
            continue
        # 命中任意已知字段表头 → 不是分组标题
        for field_keys in FIELD_HEADER_KEYWORDS.values():
            if n in field_keys:
                return False
        nonempty_vals.append(n)
    if len(nonempty_vals) < 2:
        return False
    # 每个值都必须"像段名"：不含 '.'、不是日期/代码形状、长度 2-30
    for v in nonempty_vals:
        if "." in v:
            return False
        if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", v):
            return False
        if re.fullmatch(r"\d+", v) and len(v) >= 4:
            return False
        if len(v) > 30 or len(v) < 2:
            return False
    return True


def _row_is_group_title_with_next_header(raw_rows: List[List[Any]], i: int) -> bool:
    """分组标题行 + 下一行必须是子表头（含"证券代码"/"代码"）。收紧判定避免误把数据行识别成分组。"""
    if i + 1 >= len(raw_rows):
        return False
    return _is_group_title_row(raw_rows[i]) and _is_code_header_row(raw_rows[i + 1])


def _parse_horizontal_sheet(raw_rows: List[List[Any]]) -> List[Dict[str, str]]:
    """横向：分组标题行 → 下一行是并排子表头 → 后面是数据行直到下一个分组/子表头。

    各段列范围 = [本段在分组标题行里的列位置, 下一段的列位置 - 1]。
    某段子表头全空（只含'无'）→ 跳过。
    数据块后可能跟随独立子表头行（非 group_title），继续用 vertical 逻辑解析剩余部分。
    """
    all_recs: List[Dict[str, str]] = []
    max_col = max((len(r) for r in raw_rows), default=0)
    i = 0
    while i < len(raw_rows):
        if not _row_is_group_title_with_next_header(raw_rows, i):
            # 独立的子表头行（非 group_title 前导）→ 可能是多段横向子表头，
            # 找出其中所有"证券代码/代码"起始列，按多段逻辑提取到下一个分组/子表头为止
            if _is_code_header_row(raw_rows[i]):
                header_row = raw_rows[i]
                seg_starts = [ci for ci, c in enumerate(header_row)
                              if _norm(c) in ("证券代码", "代码")]
                data_end2 = len(raw_rows)
                for j in range(i + 1, len(raw_rows)):
                    if _row_is_group_title_with_next_header(raw_rows, j):
                        data_end2 = j
                        break
                    if _is_code_header_row(raw_rows[j]):
                        data_end2 = j
                        break
                data_rows2 = raw_rows[i + 1:data_end2]
                for n, seg_col in enumerate(seg_starts):
                    next_col = seg_starts[n + 1] if n + 1 < len(seg_starts) else max_col
                    col_end = next_col - 1
                    header_slice = header_row[seg_col:col_end + 1]
                    if _is_skip_segment_header(header_slice):
                        continue
                    actual_end = seg_col
                    for col in range(col_end, seg_col - 1, -1):
                        if col < len(header_row) and _norm(header_row[col]) != "":
                            actual_end = col
                            break
                    field_to_col = _map_columns(header_row[seg_col:actual_end + 1], seg_col)
                    all_recs.extend(_extract_records(data_rows2, field_to_col))
                i = data_end2
                continue
            i += 1
            continue
        title_row = raw_rows[i]
        # 段列位置：title_row 里非空 cell 的列 idx
        seg_cols = [ci for ci, c in enumerate(title_row) if _norm(c)]
        if len(seg_cols) < 2:
            i += 1
            continue
        header_idx = i + 1
        if header_idx >= len(raw_rows):
            break
        header_row = raw_rows[header_idx]
        # 数据行：从 header_idx+1 到下一个真正的分组标题 or 首列独立子表头 or 末尾。
        data_end = len(raw_rows)
        for j in range(header_idx + 1, len(raw_rows)):
            if _row_is_group_title_with_next_header(raw_rows, j):
                data_end = j
                break
            # 首列（seg_cols[0]）出现"证券代码"独立子表头 → 新大段开始
            row_j = raw_rows[j]
            if seg_cols and seg_cols[0] < len(row_j) and _norm(row_j[seg_cols[0]]) in ("证券代码", "代码"):
                data_end = j
                break
        data_rows = raw_rows[header_idx + 1:data_end]

        for n, seg_col in enumerate(seg_cols):
            next_col = seg_cols[n + 1] if n + 1 < len(seg_cols) else max_col
            col_end = next_col - 1
            header_slice = header_row[seg_col:col_end + 1]
            if _is_skip_segment_header(header_slice):
                continue
            if not any(_norm(c) in ("证券代码", "代码") for c in header_slice):
                continue
            # 缩到 header 实际非空列
            actual_end = seg_col
            for col in range(col_end, seg_col - 1, -1):
                if col < len(header_row) and _norm(header_row[col]) != "":
                    actual_end = col
                    break
            field_to_col = _map_columns(header_row[seg_col:actual_end + 1], seg_col)
            all_recs.extend(_extract_records(data_rows, field_to_col))

        i = data_end
    return all_recs


def _detect_style(raw_rows: List[List[Any]]) -> str:
    """前 30 行如发现"分组标题行 + 下一行是证券代码子表头"组合 → horizontal，否则 vertical。"""
    for i in range(min(30, len(raw_rows))):
        if _row_is_group_title_with_next_header(raw_rows, i):
            return "horizontal"
    return "vertical"


# ------------------------- 合并 -------------------------

def _merge_records(records: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """同证券代码合并：保留最新公告日那行的其他字段，资本运作行为去重顿号拼接。"""
    if not records:
        return []
    df = pd.DataFrame(records)
    if "证券代码" not in df.columns:
        return records
    df["_orig"] = range(len(df))
    df["_date"] = pd.to_datetime(df.get("最新公告日"), errors="coerce").fillna(pd.Timestamp.min)

    def _merge_beh(series):
        seen: List[str] = []
        for v in series:
            s = _norm(v)
            if not s:
                continue
            for part in s.split("、"):
                p = part.strip()
                if p and p not in seen:
                    seen.append(p)
        return "、".join(seen)

    beh_map = df.groupby("证券代码", sort=False)["资本运作行为"].apply(_merge_beh)
    keep = (
        df.sort_values(by=["证券代码", "_date", "_orig"], ascending=[True, False, True])
        .drop_duplicates(subset="证券代码", keep="first")
        .index
    )
    merged = df.loc[keep].copy()
    merged["资本运作行为"] = merged["证券代码"].map(beh_map)
    merged = merged.sort_values("_orig").drop(columns=["_orig", "_date"]).reset_index(drop=True)
    return merged.fillna("").to_dict("records")


# ------------------------- 入口 -------------------------

def parse_pool_xlsx(file_path: str, min_date: str = "2026-03-18") -> List[Dict[str, Any]]:
    """解析整个 xlsx。返回每个 >= min_date 的 sheet 一个 dict，按 date_str 降序。"""
    out: List[Dict[str, Any]] = []
    try:
        min_d = datetime.strptime(min_date, "%Y-%m-%d").date()
    except Exception:
        raise ValueError(f"min_date 格式错误: {min_date}")
    year = datetime.now(BEIJING_TZ).year
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            d = _parse_sheet_name_to_date(sn, year)
            if d is None:
                logger.info(f"[pool_xlsx] 跳过不匹配 MMDD 的 sheet: {sn}")
                continue
            if d < min_d:
                logger.info(f"[pool_xlsx] 跳过早于 {min_date} 的 sheet: {sn} ({d})")
                continue
            ws = wb[sn]
            raw_rows: List[List[Any]] = [list(row) for row in ws.iter_rows(values_only=True)]
            style = _detect_style(raw_rows)
            if style == "horizontal":
                records = _parse_horizontal_sheet(raw_rows)
            else:
                records = _parse_vertical_sheet(raw_rows)
            merged = _merge_records(records)
            out.append({
                "sheet_name": sn,
                "date_str": d.strftime("%Y-%m-%d"),
                "style": style,
                "record_count": len(merged),
                "raw_row_count": len(records),
                "records": merged,
            })
    finally:
        wb.close()
    out.sort(key=lambda x: x["date_str"], reverse=True)
    return out
