import logging
import re as _re
import pandas as pd
from typing import Optional, List, Dict, Any
from sqlalchemy import text
from core.database import AsyncSessionLocal
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import xlsxwriter

logger = logging.getLogger(__name__)


async def save_trend_data(
    metric_type: str,
    workflow_type: str,
    date_str: str,
    count: int,
    total: int,
    source: str = "manual"
) -> bool:
    ratio = round(count / total, 4) if total > 0 else 0.0
    try:
        async with AsyncSessionLocal() as session:
            await session.execute(text("""
                INSERT INTO trend_statistics (metric_type, workflow_type, date_str, count, total, ratio, source, created_at, updated_at)
                VALUES (:metric_type, :workflow_type, :date_str, :count, :total, :ratio, :source, NOW(), NOW())
                ON DUPLICATE KEY UPDATE
                    count = VALUES(count),
                    total = VALUES(total),
                    ratio = VALUES(ratio),
                    source = VALUES(source),
                    updated_at = NOW()
            """), {
                "metric_type": metric_type,
                "workflow_type": workflow_type,
                "date_str": date_str,
                "count": count,
                "total": total,
                "ratio": ratio,
                "source": source,
            })
            await session.commit()
            logger.info(f"趋势数据已保存: {metric_type}/{workflow_type}/{date_str} count={count} total={total} ratio={ratio}")
            return True
    except Exception as e:
        logger.error(f"保存趋势数据失败: {e}")
        return False


async def get_trend_data(
    metric_type: str,
    workflow_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> List[Dict[str, Any]]:
    try:
        async with AsyncSessionLocal() as session:
            query = "SELECT id, metric_type, workflow_type, date_str, count, total, ratio, source, created_at FROM trend_statistics WHERE metric_type = :metric_type AND workflow_type NOT IN ('条件交集', '导出20日均线趋势')"
            params: Dict[str, Any] = {"metric_type": metric_type}

            if workflow_type:
                query += " AND workflow_type = :workflow_type"
                params["workflow_type"] = workflow_type
            if start_date:
                query += " AND date_str >= :start_date"
                params["start_date"] = start_date
            if end_date:
                query += " AND date_str <= :end_date"
                params["end_date"] = end_date

            query += " ORDER BY workflow_type, date_str ASC"
            result = await session.execute(text(query), params)
            rows = result.fetchall()

            return [
                {
                    "id": r[0], "metric_type": r[1], "workflow_type": r[2],
                    "date_str": r[3], "count": r[4], "total": r[5],
                    "ratio": r[6], "source": r[7],
                    "created_at": r[8].strftime("%Y-%m-%d %H:%M:%S") if r[8] else None
                }
                for r in rows
            ]
    except Exception as e:
        logger.error(f"查询趋势数据失败: {e}")
        return []


async def delete_trend_data(record_id: int) -> bool:
    try:
        async with AsyncSessionLocal() as session:
            await session.execute(text("DELETE FROM trend_statistics WHERE id = :id"), {"id": record_id})
            await session.commit()
            return True
    except Exception as e:
        logger.error(f"删除趋势数据失败: {e}")
        return False


def _find_dual_columns_sheet(file_path: str, left_keywords: List[str], right_keywords: List[str]) -> Optional[str]:
    """扫描 Excel 所有可见 sheet，找第一个符合"日期 + 左关键字 + 右关键字"双行表头的 sheet。

    left/right 关键字是 "or" 语义：只要前 3 行里有任一 left_keywords 的词命中、
    且有任一 right_keywords 的词命中、且出现"日期"或"占比"，即判定为目标 sheet。
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        return None
    target = None
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            head_rows = []
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                head_rows.append(row)
                if len(head_rows) >= 3:
                    break
            flat = "|".join(str(c) for r in head_rows for c in r if c is not None)
            has_left = any(k in flat for k in left_keywords)
            has_right = any(k in flat for k in right_keywords)
            if has_left and has_right and ("日期" in flat or "占比" in flat):
                target = sn
                break
    finally:
        wb.close()
    return target


def _find_pledge_side_by_side_sheet(file_path: str) -> Optional[str]:
    """向后兼容 thin wrapper。"""
    return _find_dual_columns_sheet(file_path, ["中大盘"], ["小盘"])


def _parse_dual_columns_excel(
    file_path: str,
    left_workflow_type: str,
    right_workflow_type: str,
    left_keywords: List[str],
    right_keywords: List[str],
    log_tag: str = "双列并排",
) -> List[Dict[str, Any]]:
    """通用"并排双列"Excel 解析：日期 | 左(数量/占比) | 右(数量/占比)。

    表头两行：Row1 含左/右分组关键字（例 中大盘/小盘 或 2026/2025）；
    Row2 含 占20均线数量/占比。每行生成 2 条记录（分别标 left/right_workflow_type）。
    """
    target_sheet = _find_dual_columns_sheet(file_path, left_keywords, right_keywords)
    if target_sheet is None:
        raise ValueError(
            f"未找到符合格式的 sheet（需要双行表头：日期 | {left_keywords[0]}-数量/占比 | {right_keywords[0]}-数量/占比）"
        )
    df = pd.read_excel(file_path, sheet_name=target_sheet, header=[0, 1])
    logger.info(f"[{log_tag}] 使用 sheet: {target_sheet}")

    date_col = None
    left_count_col = None
    left_ratio_col = None
    right_count_col = None
    right_ratio_col = None

    def _norm(s):
        return str(s).replace('\n', '').replace(' ', '').strip()

    def _match_any(s: str, keywords: List[str]) -> bool:
        return any(k in s for k in keywords)

    for c in df.columns:
        l0 = _norm(c[0])
        l1 = _norm(c[1])
        if ("日期" in l0 or "日期" in l1) and date_col is None:
            date_col = c
            continue
        is_left = _match_any(l0, left_keywords) or _match_any(l1, left_keywords)
        is_right = _match_any(l0, right_keywords) or _match_any(l1, right_keywords)
        is_count = "数量" in l1 or "数量" in l0
        is_ratio = "占比" in l1 or "占比" in l0 or "比例" in l1 or "比例" in l0
        if is_left and is_count:
            left_count_col = c
        elif is_left and is_ratio:
            left_ratio_col = c
        elif is_right and is_count:
            right_count_col = c
        elif is_right and is_ratio:
            right_ratio_col = c

    if not date_col:
        raise ValueError(f"未找到日期列，可用列: {list(df.columns)}")
    if not (left_count_col or right_count_col):
        raise ValueError(
            f"未找到'{left_keywords[0]}'或'{right_keywords[0]}'的数量列，可用列: {list(df.columns)}"
        )

    def _to_ratio_float(v):
        if pd.isna(v):
            return None
        if isinstance(v, str):
            v2 = v.strip().rstrip('%')
            try:
                f = float(v2)
                if f > 1:
                    f /= 100
                return f
            except ValueError:
                return None
        if isinstance(v, (int, float)):
            f = float(v)
            if f > 1:
                f /= 100
            return f
        return None

    # 自动识别日期年份：如果原表只给"3月11日"，需要推断年份；就近规则：
    # 基准 = 今天；若"M月D日"的日期晚于今天则视为去年，否则今年。
    from datetime import datetime as _dt
    base_year = _dt.now().year
    today = _dt.now().date()
    re_cnd = _re.compile(r'(\d+)\s*月\s*(\d+)\s*日')

    def _parse_date(val):
        if pd.isna(val):
            return None
        if isinstance(val, (int, float)) and 30000 <= val <= 60000:
            try:
                return (pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(val))).strftime("%Y-%m-%d")
            except Exception:
                return None
        if isinstance(val, str):
            s = val.strip()
            m = re_cnd.match(s)
            if m:
                month, day = int(m.group(1)), int(m.group(2))
                for y in (base_year, base_year - 1):
                    try:
                        d = _dt(y, month, day).date()
                        if d <= today:
                            return d.strftime("%Y-%m-%d")
                    except ValueError:
                        continue
                return None
        try:
            return pd.to_datetime(val).strftime("%Y-%m-%d")
        except Exception:
            return None

    def _emit(row, cnt_col, ratio_col, wt_label):
        if cnt_col is None:
            return None
        cnt_v = row[cnt_col]
        if not pd.notna(cnt_v):
            return None
        try:
            cnt = int(float(str(cnt_v).strip()))
            ratio_f = _to_ratio_float(row[ratio_col]) if ratio_col is not None else None
            if ratio_f and ratio_f > 0:
                total = round(cnt / ratio_f)
                return {
                    "workflow_type": wt_label,
                    "date_str": None,
                    "count": cnt,
                    "total": int(total),
                    "ratio": round(cnt / total, 4) if total > 0 else 0.0,
                }
        except (ValueError, TypeError):
            return None
        return None

    records = []
    for _, row in df.iterrows():
        date_str = _parse_date(row[date_col])
        if not date_str:
            continue
        for rec_col_pair, wt_label in (
            ((left_count_col, left_ratio_col), left_workflow_type),
            ((right_count_col, right_ratio_col), right_workflow_type),
        ):
            rec = _emit(row, rec_col_pair[0], rec_col_pair[1], wt_label)
            if rec is not None:
                rec["date_str"] = date_str
                records.append(rec)
    return records


def _parse_pledge_side_by_side(file_path: str) -> List[Dict[str, Any]]:
    """向后兼容 thin wrapper：解析质押中大盘+小盘双列。"""
    return _parse_dual_columns_excel(
        file_path=file_path,
        left_workflow_type="质押(中大盘)",
        right_workflow_type="质押(小盘)",
        left_keywords=["中大盘"],
        right_keywords=["小盘"],
        log_tag="质押双列并排",
    )


def parse_excel_for_trend(file_path: str, workflow_type: str) -> List[Dict[str, Any]]:
    """解析 Excel 文件为趋势数据记录列表，供前端预览确认"""
    # 特殊类型：质押并排双列
    if workflow_type == "质押(双列并排)":
        return _parse_pledge_side_by_side(file_path)
    # 年度(X)：本年 + 上年 并排双列
    if isinstance(workflow_type, str) and workflow_type.startswith("年度(") and workflow_type.endswith(")"):
        parent = workflow_type[len("年度("):-1]
        from datetime import datetime as _dt_y
        Y = _dt_y.now().year
        return _parse_dual_columns_excel(
            file_path=file_path,
            left_workflow_type=f"{parent}({Y})",
            right_workflow_type=f"{parent}({Y - 1})",
            left_keywords=[str(Y), f"{Y}至今", "本年", "今年"],
            right_keywords=[str(Y - 1), "上年", "去年"],
            log_tag=f"年度双列并排({parent})",
        )
    df = pd.read_excel(file_path)

    # 归一化列名：去除空白、换行后缀
    col_map = {}
    for col in df.columns:
        clean = str(col).split('\n')[0].strip()
        col_map[col] = clean
    df = df.rename(columns=col_map)

    # 映射标准列名
    rename = {}
    for col in df.columns:
        cl = col.lower().strip()
        if cl in ("日期", "date", "数据日期"):
            rename[col] = "日期"
        elif "占20均线" in col or "站20日均线" in col or "20日均线数量" in cl:
            rename[col] = "站20日均线数量"
        elif "占比" in col or "比例" in col:
            rename[col] = "占比"
    df = df.rename(columns=rename)

    if "日期" not in df.columns:
        raise ValueError(f"未找到日期列，可用列: {list(df.columns)}")
    if "站20日均线数量" not in df.columns:
        raise ValueError(f"未找到数量列(占20均线数量/站20日均线数量)，可用列: {list(df.columns)}")

    records = []
    for _, row in df.iterrows():
        date_val = row["日期"]
        if pd.isna(date_val):
            continue

        # 解析日期（兼容 Excel 序列号、datetime 对象、字符串）
        try:
            if isinstance(date_val, (int, float)) and 30000 <= date_val <= 60000:
                # Excel 序列号 → 日期
                date_str = (pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(date_val))).strftime("%Y-%m-%d")
            else:
                date_str = pd.to_datetime(date_val).strftime("%Y-%m-%d")
        except Exception:
            continue

        count = int(float(str(row["站20日均线数量"]).strip())) if pd.notna(row["站20日均线数量"]) else 0

        # 从占比反算总量
        total = 0
        if "占比" in df.columns and pd.notna(row.get("占比")):
            ratio_val = row["占比"]
            if isinstance(ratio_val, str):
                ratio_val = ratio_val.strip().rstrip('%')
                try:
                    ratio_float = float(ratio_val)
                    # 判断是百分比值还是小数值
                    if ratio_float > 1:
                        ratio_float = ratio_float / 100
                    total = round(count / ratio_float) if ratio_float > 0 else 0
                except ValueError:
                    pass
            elif isinstance(ratio_val, (int, float)):
                ratio_float = float(ratio_val)
                if ratio_float > 1:
                    ratio_float = ratio_float / 100
                total = round(count / ratio_float) if ratio_float > 0 else 0

        if "总量" in df.columns and pd.notna(row.get("总量")):
            total = int(float(str(row["总量"]).strip()))

        ratio = round(count / total, 4) if total > 0 else 0.0

        records.append({
            "workflow_type": workflow_type,
            "date_str": date_str,
            "count": count,
            "total": total,
            "ratio": ratio,
        })

    return records


async def batch_save_trend_data(metric_type: str, records: List[Dict[str, Any]], source: str = "excel") -> int:
    if not records:
        return 0
    try:
        async with AsyncSessionLocal() as session:
            for rec in records:
                count = rec["count"]
                total = rec["total"]
                ratio = round(count / total, 4) if total > 0 else 0.0
                await session.execute(text("""
                    INSERT INTO trend_statistics (metric_type, workflow_type, date_str, count, total, ratio, source, created_at, updated_at)
                    VALUES (:metric_type, :wt, :ds, :count, :total, :ratio, :source, NOW(), NOW())
                    ON DUPLICATE KEY UPDATE
                        count = VALUES(count), total = VALUES(total), ratio = VALUES(ratio),
                        source = VALUES(source), updated_at = NOW()
                """), {
                    "metric_type": metric_type, "wt": rec["workflow_type"], "ds": rec["date_str"],
                    "count": count, "total": total, "ratio": ratio, "source": source,
                })
            await session.commit()
            logger.info(f"批量保存趋势数据: {len(records)}条")
            return len(records)
    except Exception as e:
        logger.error(f"批量保存趋势数据失败: {e}")
        return 0


def export_trend_excel_with_chart(data: List[Dict], file_path: str, single_type: str = None):
    """导出趋势数据到 Excel（xlsxwriter）：每个工作流类型的数据 + 双Y轴折线图，兼容 WPS 和 Excel

    质押类型特殊处理：DB 里的 '质押(中大盘)' 和 '质押(小盘)' 两个子类型在导出时
    合并成一个 '质押' 逻辑组，同一表格并列两套数据列，共享 1 个双曲线图表。
    """
    from config.workflow_type_config import WORKFLOW_TYPE_CONFIG

    # 按工作流类型分组；质押所有变种统一归入 pledge_sub 两个桶（不入 type_groups）
    # 年度子类型（并购重组(YYYY) / 股权转让(YYYY) / 招投标(YYYY)）折叠到父类型下，
    # 带 _year_label=YYYY 字符串，导出时用双线图展示 Y vs Y-1
    # 历史遗留的 '质押' 裸记录会被丢弃（旧版本单线图不再导出，改由下方合并双曲线图覆盖）
    type_groups = {}
    pledge_sub = {"中大盘": [], "小盘": []}
    yearly_parents = ("并购重组", "股权转让", "招投标")
    import re as _re_mod
    _yearly_re = _re_mod.compile(r'^(并购重组|股权转让|招投标)\((\d{4})\)$')
    for d in data:
        wt = d["workflow_type"]
        if wt == "质押(中大盘)":
            pledge_sub["中大盘"].append(d)
            continue
        if wt == "质押(小盘)":
            pledge_sub["小盘"].append(d)
            continue
        if wt == "质押":
            continue
        m = _yearly_re.match(wt or "")
        if m:
            parent, year_str = m.group(1), m.group(2)
            if parent not in type_groups:
                type_groups[parent] = []
            type_groups[parent].append({**d, "_year_label": year_str})
            continue
        if wt not in type_groups:
            type_groups[wt] = []
        type_groups[wt].append(d)

    # 从 config 提取排序前缀和显示名
    def get_type_sort_key(wt):
        cfg = WORKFLOW_TYPE_CONFIG.get(wt, WORKFLOW_TYPE_CONFIG.get("", {}))
        tpl = cfg.get("naming", {}).get("output_template", "")
        # 提取开头数字如 "1并购重组{date}.xlsx" → 1
        import re
        m = re.match(r'^(\d+)', tpl)
        return int(m.group(1)) if m else 99

    def get_type_prefix(wt):
        cfg = WORKFLOW_TYPE_CONFIG.get(wt, WORKFLOW_TYPE_CONFIG.get("", {}))
        tpl = cfg.get("naming", {}).get("output_template", "")
        import re
        m = re.match(r'^(\d+)', tpl)
        return m.group(1) if m else ""

    type_names = sorted(type_groups.keys(), key=get_type_sort_key)

    wb = xlsxwriter.Workbook(file_path)
    ws = wb.add_worksheet("站上20日均线趋势")

    # 格式
    title_fmt = wb.add_format({'bold': True, 'font_size': 13, 'align': 'center', 'valign': 'vcenter'})
    header_fmt = wb.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    data_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
    pct_fmt = wb.add_format({'border': 1, 'num_format': '0.00', 'align': 'center', 'valign': 'vcenter'})

    # 列宽
    ws.set_column('A:A', 10)
    ws.set_column('B:B', 16)
    ws.set_column('C:C', 10)
    ws.set_column('D:D', 12)
    ws.set_column('E:E', 14)

    if not type_names:
        ws.write(0, 0, "暂无数据")
        wb.close()
        return

    current_row = 0
    pledge_written = False

    def _write_pledge_block(start_row: int) -> int:
        """写入【5质押】合并表格 + 双曲线图，返回写入后的 current_row。"""
        if not (pledge_sub["中大盘"] or pledge_sub["小盘"]):
            return start_row
        zdp_items = sorted(pledge_sub["中大盘"], key=lambda x: x["date_str"])
        xp_items = sorted(pledge_sub["小盘"], key=lambda x: x["date_str"])
        all_dates = sorted(set([d["date_str"] for d in zdp_items] + [d["date_str"] for d in xp_items]))
        zdp_map = {d["date_str"]: d for d in zdp_items}
        xp_map = {d["date_str"]: d for d in xp_items}

        cur = start_row
        ws.merge_range(cur, 0, cur, 8, "【5质押】", title_fmt)
        cur += 1
        headers = ["日期", "中大盘数量", "中大盘总量", "中大盘占比(%)",
                   "小盘数量", "小盘总量", "小盘占比(%)", "完整日期"]
        for col_idx, h in enumerate(headers):
            ws.write(cur, col_idx, h, header_fmt)
        header_row_local = cur
        cur += 1
        data_start_row = cur
        for ds in all_dates:
            try:
                parts = ds.split('-')
                short_date = f"{int(parts[1])}/{int(parts[2])}"
            except Exception:
                short_date = ds
            zd = zdp_map.get(ds)
            xd = xp_map.get(ds)
            ws.write(cur, 0, short_date, data_fmt)
            if zd:
                ws.write(cur, 1, zd["count"], data_fmt)
                ws.write(cur, 2, zd["total"], data_fmt)
                ws.write(cur, 3, round(zd["ratio"] * 100, 2) if zd.get("ratio") else 0, pct_fmt)
            else:
                for c in (1, 2, 3):
                    ws.write(cur, c, "", data_fmt)
            if xd:
                ws.write(cur, 4, xd["count"], data_fmt)
                ws.write(cur, 5, xd["total"], data_fmt)
                ws.write(cur, 6, round(xd["ratio"] * 100, 2) if xd.get("ratio") else 0, pct_fmt)
            else:
                for c in (4, 5, 6):
                    ws.write(cur, c, "", data_fmt)
            ws.write(cur, 7, ds, data_fmt)
            cur += 1
        data_end_row = cur - 1

        chart_local = wb.add_chart({'type': 'line'})
        chart_local.set_title({'name': "5质押 - 站上20日均线占比趋势（中大盘 vs 小盘）"})
        chart_local.set_size({'width': 720, 'height': 380})
        chart_local.add_series({
            'name': '中大盘占比(%)',
            'categories': ['站上20日均线趋势', data_start_row, 0, data_end_row, 0],
            'values': ['站上20日均线趋势', data_start_row, 3, data_end_row, 3],
            'line': {'width': 2.5, 'color': '#409EFF'},
            'marker': {'type': 'circle', 'size': 4, 'fill': {'color': '#409EFF'}},
        })
        chart_local.add_series({
            'name': '小盘占比(%)',
            'categories': ['站上20日均线趋势', data_start_row, 0, data_end_row, 0],
            'values': ['站上20日均线趋势', data_start_row, 6, data_end_row, 6],
            'line': {'width': 2.5, 'color': '#E6A23C'},
            'marker': {'type': 'circle', 'size': 4, 'fill': {'color': '#E6A23C'}},
        })
        chart_local.set_y_axis({'name': '占比(%)', 'num_format': '0.00'})
        num_points = len(all_dates)
        x_axis_opts = {'name': '日期', 'label_position': 'low'}
        if num_points > 15:
            interval = max(1, num_points // 15)
            x_axis_opts['interval_unit'] = interval
            x_axis_opts['num_font'] = {'rotation': -45, 'size': 9}
        chart_local.set_x_axis(x_axis_opts)
        ws.insert_chart(header_row_local - 1, 10, chart_local)

        chart_height_rows = 22
        data_rows_used = len(all_dates) + 2
        return header_row_local - 1 + max(chart_height_rows, data_rows_used) + 2

    def _write_yearly_block(parent: str, items: list, start_row: int) -> int:
        """写入年度父类型（并购/股权转让/招投标）的双线块：Y vs Y-1"""
        if not items:
            return start_row
        # 按 _year_label 分组 → 取最新 2 个年份（倒序）
        from collections import defaultdict as _dd
        by_year = _dd(list)
        for it in items:
            by_year[str(it.get("_year_label", ""))].append(it)
        years_sorted = sorted([y for y in by_year.keys() if y], reverse=True)
        if not years_sorted:
            return start_row
        # 最多画两个年份
        years_to_draw = years_sorted[:2]
        Y = years_to_draw[0]
        Y1 = years_to_draw[1] if len(years_to_draw) > 1 else None

        y_items = sorted(by_year[Y], key=lambda x: x["date_str"])
        y1_items = sorted(by_year[Y1], key=lambda x: x["date_str"]) if Y1 else []
        all_dates = sorted(set([d["date_str"] for d in y_items] + [d["date_str"] for d in y1_items]))
        y_map = {d["date_str"]: d for d in y_items}
        y1_map = {d["date_str"]: d for d in y1_items}

        prefix = get_type_prefix(parent)
        table_title = f"【{prefix}{parent}】" if prefix else f"【{parent}】"

        cur = start_row
        ws.merge_range(cur, 0, cur, 8, table_title, title_fmt)
        cur += 1
        if Y1:
            headers = ["日期",
                       f"{Y}数量", f"{Y}总量", f"{Y}占比(%)",
                       f"{Y1}数量", f"{Y1}总量", f"{Y1}占比(%)",
                       "完整日期"]
        else:
            headers = ["日期", f"{Y}数量", f"{Y}总量", f"{Y}占比(%)", "完整日期"]
        for col_idx, h in enumerate(headers):
            ws.write(cur, col_idx, h, header_fmt)
        header_row_local = cur
        cur += 1
        data_start_row = cur
        for ds in all_dates:
            try:
                parts = ds.split('-')
                short_date = f"{int(parts[1])}/{int(parts[2])}"
            except Exception:
                short_date = ds
            yd = y_map.get(ds)
            yd1 = y1_map.get(ds)
            ws.write(cur, 0, short_date, data_fmt)
            if yd:
                ws.write(cur, 1, yd["count"], data_fmt)
                ws.write(cur, 2, yd["total"], data_fmt)
                ws.write(cur, 3, round(yd["ratio"] * 100, 2) if yd.get("ratio") else 0, pct_fmt)
            else:
                for c in (1, 2, 3):
                    ws.write(cur, c, "", data_fmt)
            if Y1:
                if yd1:
                    ws.write(cur, 4, yd1["count"], data_fmt)
                    ws.write(cur, 5, yd1["total"], data_fmt)
                    ws.write(cur, 6, round(yd1["ratio"] * 100, 2) if yd1.get("ratio") else 0, pct_fmt)
                else:
                    for c in (4, 5, 6):
                        ws.write(cur, c, "", data_fmt)
                ws.write(cur, 7, ds, data_fmt)
            else:
                ws.write(cur, 4, ds, data_fmt)
            cur += 1
        data_end_row = cur - 1

        chart_local = wb.add_chart({'type': 'line'})
        chart_title_local = f"{prefix}{parent} - 站上20日均线占比趋势（{Y} vs {Y1})" if Y1 else f"{prefix}{parent} - 站上20日均线占比趋势（{Y})"
        chart_local.set_title({'name': chart_title_local})
        chart_local.set_size({'width': 720, 'height': 380})
        chart_local.add_series({
            'name': str(Y),
            'categories': ['站上20日均线趋势', data_start_row, 0, data_end_row, 0],
            'values': ['站上20日均线趋势', data_start_row, 3, data_end_row, 3],
            'line': {'width': 2.5, 'color': '#409EFF'},
            'marker': {'type': 'circle', 'size': 4, 'fill': {'color': '#409EFF'}},
        })
        if Y1:
            chart_local.add_series({
                'name': str(Y1),
                'categories': ['站上20日均线趋势', data_start_row, 0, data_end_row, 0],
                'values': ['站上20日均线趋势', data_start_row, 6, data_end_row, 6],
                'line': {'width': 2.5, 'color': '#E6A23C'},
                'marker': {'type': 'circle', 'size': 4, 'fill': {'color': '#E6A23C'}},
            })
        chart_local.set_y_axis({'name': '占比(%)', 'num_format': '0.00'})
        num_points = len(all_dates)
        x_axis_opts = {'name': '日期', 'label_position': 'low'}
        if num_points > 15:
            interval = max(1, num_points // 15)
            x_axis_opts['interval_unit'] = interval
            x_axis_opts['num_font'] = {'rotation': -45, 'size': 9}
        chart_local.set_x_axis(x_axis_opts)
        ws.insert_chart(header_row_local - 1, 10, chart_local)

        chart_height_rows = 22
        data_rows_used = len(all_dates) + 2
        return header_row_local - 1 + max(chart_height_rows, data_rows_used) + 2

    for wt in type_names:
        items = sorted(type_groups[wt], key=lambda x: x["date_str"])
        if not items:
            continue

        # 在 prefix > 5 的第一个类型前插入【5质押】块
        if not pledge_written and get_type_sort_key(wt) > 5:
            current_row = _write_pledge_block(current_row)
            pledge_written = True

        # 年度父类型：折叠后的 items 含 _year_label，走双线块
        if wt in yearly_parents and items and any("_year_label" in it for it in items):
            current_row = _write_yearly_block(wt, items, current_row)
            continue

        prefix = get_type_prefix(wt)
        # 表格标题（不含后缀）
        table_title = f"【{prefix}{wt}】" if prefix else f"【{wt}】"
        # 图表标题（保留完整）
        chart_title = f"{prefix}{wt} - 站上20日均线占比趋势"

        # 标题（合并 A~E 列）
        ws.merge_range(current_row, 0, current_row, 4, table_title, title_fmt)
        current_row += 1

        # 表头
        for col_idx, h in enumerate(["日期", "站20均线数量", "总量", "占比(%)", "完整日期"]):
            ws.write(current_row, col_idx, h, header_fmt)
        header_row = current_row
        current_row += 1

        # 数据
        data_start_row = current_row
        for item in items:
            # 图表横坐标用短格式 M/D
            ds = item["date_str"]
            try:
                parts = ds.split('-')
                short_date = f"{int(parts[1])}/{int(parts[2])}"
            except Exception:
                short_date = ds
            ws.write(current_row, 0, short_date, data_fmt)
            ws.write(current_row, 1, item["count"], data_fmt)
            ws.write(current_row, 2, item["total"], data_fmt)
            ws.write(current_row, 3, round(item["ratio"] * 100, 2) if item.get("ratio") else 0, pct_fmt)
            ws.write(current_row, 4, ds, data_fmt)  # 完整日期备查
            current_row += 1
        data_end_row = current_row - 1

        # 图表：仅占比(%)折线
        chart = wb.add_chart({'type': 'line'})
        chart.set_title({'name': chart_title})
        chart.set_size({'width': 620, 'height': 360})
        chart.set_legend({'none': True})

        chart.add_series({
            'name': '占比(%)',
            'categories': ['站上20日均线趋势', data_start_row, 0, data_end_row, 0],
            'values': ['站上20日均线趋势', data_start_row, 3, data_end_row, 3],
            'line': {'width': 2.5, 'color': '#409EFF'},
            'marker': {'type': 'circle', 'size': 4, 'fill': {'color': '#409EFF'}},
        })

        chart.set_y_axis({'name': '占比(%)', 'num_format': '0.00'})

        # X轴：日期过多时间隔显示标签，保证最多约15个可见标签
        num_points = len(items)
        x_axis_opts = {'name': '日期', 'label_position': 'low'}
        if num_points > 15:
            interval = max(1, num_points // 15)
            x_axis_opts['interval_unit'] = interval
            x_axis_opts['label_position'] = 'low'
            x_axis_opts['num_font'] = {'rotation': -45, 'size': 9}
        chart.set_x_axis(x_axis_opts)

        ws.insert_chart(header_row - 1, 8, chart)

        # 间距
        chart_height_rows = 22
        data_rows_used = len(items) + 2
        current_row = header_row - 1 + max(chart_height_rows, data_rows_used) + 2

    # 若遍历完所有类型都没触发插入（所有 prefix ≤ 5），兜底追加在末尾
    if not pledge_written:
        current_row = _write_pledge_block(current_row)
        pledge_written = True

    wb.close()
    logger.info(
        f"趋势Excel已导出(xlsxwriter): {file_path}, {len(data)}条, "
        f"{len(type_names)}个类型"
        f"{'；质押(中大盘 '+str(len(pledge_sub['中大盘']))+' / 小盘 '+str(len(pledge_sub['小盘']))+')' if pledge_sub['中大盘'] or pledge_sub['小盘'] else ''}"
    )
