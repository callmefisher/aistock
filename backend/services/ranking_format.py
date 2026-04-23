"""涨幅排名 Excel 格式化统一工具。

工作流(`_ranking_sort`) 和 统计分析导出(`statistics_api.download_result`) 共用。
识别两种布局：
 - 老布局：板块 | 今日涨跌幅 | 迄今为止排进前5(次数) | 日期列...
 - 扩展布局：板块 | 年初 | B排名 | 月初 | D排名 | 今日涨跌幅 | 次数 | 日期列...

格式规则：
 - 金色头 = 次数列
 - Top5 深红 (#C00000): 日期列 + （扩展时）年初/月初排名列
 - 6-10 名浅红 (#FF0000):（扩展时）年初/月初排名列
 - 当日列（日期第1列）浅红：非 Top5 且排名提升（相比上日列）
"""
from typing import List, Optional
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


DEEP_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
DEEP_RED_FONT = Font(color="FFFFFF", bold=True)
LIGHT_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
LIGHT_RED_FONT = Font(color="FFFFFF")
# 6-10 名专用橙色（按用户截图 #E69138）
RANK_6_10_FILL = PatternFill(start_color="E69138", end_color="E69138", fill_type="solid")
RANK_6_10_FONT = Font(color="FFFFFF", bold=True)
GOLD_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")


def _find_col(headers: List[str], predicate) -> int:
    """返回 1-based 列号；找不到返回 -1。"""
    for i, h in enumerate(headers):
        if predicate(h):
            return i + 1
    return -1


def detect_ranking_layout(headers: List[str]) -> dict:
    """按列头识别布局，返回各关键列位置 (1-based)。
    未识别的位置为 -1。"""
    top5_count_col = _find_col(headers, lambda h: '迄今为止排进前5' in h)
    ytd_rank_col = _find_col(headers, lambda h: h == 'B列的数值升序排序结果')
    mtd_rank_col = _find_col(headers, lambda h: h == 'D列的数值升序排序结果')
    is_extended = top5_count_col > 0 and ytd_rank_col > 0 and mtd_rank_col > 0

    # 次数列 = 金色列；老布局=col3，扩展=top5_count_col；找不到降级到 3
    gold_col = top5_count_col if top5_count_col > 0 else 3
    # 日期列起点 = 次数列 + 1
    date_start_col = (top5_count_col + 1) if top5_count_col > 0 else 4

    return {
        "is_extended": is_extended,
        "gold_col": gold_col,
        "date_start_col": date_start_col,
        "ytd_rank_col": ytd_rank_col,
        "mtd_rank_col": mtd_rank_col,
        "top5_count_col": top5_count_col,
    }


def _cell_int(cell):
    try:
        return int(cell.value) if cell.value is not None else None
    except (ValueError, TypeError):
        return None


def apply_ranking_format(
    ws,
    prev_rank_by_sector: Optional[dict] = None,
    date_col_date_map: Optional[dict] = None,
    ref_date=None,
):
    """对已写入数据的 sheet 应用涨幅排名格式。

    参数：
     - prev_rank_by_sector: {板块名称: 上一工作日排名} —— 用于"当日列排名提升浅红"。
       None 时改用 sheet 自身的"上日列"值比对（供统计分析导出路径使用）。
     - date_col_date_map: {"4月22日": datetime(2026,4,22), ...} —— 日期列头写入实际日期值。
     - ref_date: datetime，用于把"4月22日"字符串解析成含年份的日期对象（统计分析路径使用）。
       与 date_col_date_map 二选一即可。
    """
    import re as re_mod
    from datetime import datetime as dt

    max_row = ws.max_row
    max_col = ws.max_column
    headers = [str(ws.cell(row=1, column=c).value or '') for c in range(1, max_col + 1)]
    layout = detect_ranking_layout(headers)
    gold_col = layout["gold_col"]
    date_start_col = layout["date_start_col"]
    ytd_rank_col = layout["ytd_rank_col"]
    mtd_rank_col = layout["mtd_rank_col"]

    # 表头：加粗居中；日期列头写入日期值
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        header = str(cell.value or '')
        if date_col_date_map and header in date_col_date_map:
            cell.value = date_col_date_map[header]
            cell.number_format = 'm"月"d"日"'
        else:
            m = re_mod.match(r'^(\d+)月(\d+)日$', header)
            if m and ref_date is not None:
                month, day = int(m.group(1)), int(m.group(2))
                year = ref_date.year if month <= ref_date.month else ref_date.year - 1
                cell.value = dt(year, month, day)
                cell.number_format = 'm"月"d"日"'
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if 1 <= gold_col <= max_col:
        ws.cell(row=1, column=gold_col).fill = GOLD_FILL
        ws.cell(row=1, column=gold_col).font = Font(bold=True, size=11)

    # 列宽
    wide_cols = {2, gold_col}
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 35 if col_idx in wide_cols else 25

    # 所有数据单元格居中
    center = Alignment(horizontal="center", vertical="center")
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center

    # Top5 深红：日期列 + （扩展时）年初/月初排名列
    top5_cols = set(range(date_start_col, max_col + 1))
    rank_rank_cols: set = set()  # 年初/月初排名列（需要 6-10 名浅红）
    if layout["is_extended"]:
        rank_rank_cols = {ytd_rank_col, mtd_rank_col}
        top5_cols.update(rank_rank_cols)
    for col_idx in sorted(top5_cols):
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = _cell_int(cell)
            if val is not None:
                if val <= 5:
                    cell.fill = DEEP_RED
                    cell.font = DEEP_RED_FONT
                elif col_idx in rank_rank_cols and 6 <= val <= 10:
                    cell.fill = RANK_6_10_FILL
                    cell.font = RANK_6_10_FONT

    # 当日列（日期第1列）浅红：非 Top5 且排名提升
    if max_col >= date_start_col:
        today_idx = date_start_col
        for row_idx in range(2, max_row + 1):
            cell_today = ws.cell(row=row_idx, column=today_idx)
            rank_today = _cell_int(cell_today)
            if rank_today is None or rank_today <= 5:
                continue
            # 两种方式找上一日排名：优先 prev_rank_by_sector；否则读"上日列"
            rank_prev = None
            if prev_rank_by_sector is not None:
                sector_name = str(ws.cell(row=row_idx, column=1).value or '').strip()
                rank_prev = prev_rank_by_sector.get(sector_name)
            elif max_col >= date_start_col + 1:
                rank_prev = _cell_int(ws.cell(row=row_idx, column=date_start_col + 1))
            if rank_prev is not None and rank_today < rank_prev:
                cell_today.fill = LIGHT_RED
                cell_today.font = LIGHT_RED_FONT

    # 自动过滤
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
