import pandas as pd
import os
from typing import Dict, List, Optional, Any
from datetime import datetime
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class ExcelProcessor:
    def __init__(self, output_dir: str = "/app/data/excel"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def save_excel(
        self,
        df: pd.DataFrame,
        filename: str,
        sheet_name: str = "Sheet1",
        apply_formatting: bool = True
    ) -> Dict[str, Any]:
        try:
            filepath = os.path.join(self.output_dir, filename)
            
            df.to_excel(filepath, sheet_name=sheet_name, index=False)
            
            if apply_formatting:
                self._apply_formatting(filepath, sheet_name)
            
            return {
                "success": True,
                "file_path": filepath,
                "rows": len(df),
                "columns": len(df.columns),
                "message": f"成功保存Excel文件：{filepath}"
            }
        except Exception as e:
            logger.error(f"Excel保存失败: {str(e)}")
            return {
                "success": False,
                "file_path": None,
                "message": f"Excel保存失败: {str(e)}"
            }
    
    def _apply_formatting(self, filepath: str, sheet_name: str):
        try:
            wb = load_workbook(filepath)
            ws = wb[sheet_name]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
        except Exception as e:
            logger.warning(f"格式化失败: {str(e)}")
    
    def read_excel(self, filepath: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        try:
            if not os.path.exists(filepath):
                return {
                    "success": False,
                    "data": None,
                    "message": f"文件不存在：{filepath}"
                }
            
            df = pd.read_excel(filepath, sheet_name=sheet_name or 0)
            
            return {
                "success": True,
                "data": df,
                "rows": len(df),
                "columns": list(df.columns),
                "message": f"成功读取Excel，共{len(df)}行"
            }
        except Exception as e:
            logger.error(f"Excel读取失败: {str(e)}")
            return {
                "success": False,
                "data": None,
                "message": f"Excel读取失败: {str(e)}"
            }
    
    def merge_excels(
        self,
        filepaths: List[str],
        output_filename: str,
        merge_type: str = "vertical"
    ) -> Dict[str, Any]:
        try:
            dfs = []
            for filepath in filepaths:
                result = self.read_excel(filepath)
                if result['success']:
                    dfs.append(result['data'])
            
            if not dfs:
                return {
                    "success": False,
                    "message": "没有可合并的数据"
                }
            
            if merge_type == "vertical":
                merged_df = pd.concat(dfs, ignore_index=True)
            else:
                merged_df = dfs[0]
                for df in dfs[1:]:
                    merged_df = pd.merge(merged_df, df, how='outer')
            
            return self.save_excel(merged_df, output_filename)
        except Exception as e:
            logger.error(f"Excel合并失败: {str(e)}")
            return {
                "success": False,
                "message": f"Excel合并失败: {str(e)}"
            }
    
    def generate_stock_pool(
        self,
        df: pd.DataFrame,
        task_name: str,
        additional_info: Optional[Dict] = None
    ) -> Dict[str, Any]:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"stock_pool_{task_name}_{timestamp}.xlsx"
            
            summary_data = {
                "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "任务名称": task_name,
                "股票总数": len(df),
                "列数": len(df.columns)
            }
            
            if additional_info:
                summary_data.update(additional_info)
            
            summary_df = pd.DataFrame([summary_data])
            
            filepath = os.path.join(self.output_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='汇总信息', index=False)
                df.to_excel(writer, sheet_name='选股池', index=False)
            
            self._apply_formatting(filepath, '汇总信息')
            self._apply_formatting(filepath, '选股池')
            
            return {
                "success": True,
                "file_path": filepath,
                "filename": filename,
                "total_stocks": len(df),
                "message": f"成功生成选股池：{filename}"
            }
        except Exception as e:
            logger.error(f"选股池生成失败: {str(e)}")
            return {
                "success": False,
                "message": f"选股池生成失败: {str(e)}"
            }
    
    def get_column_info(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        columns_info = []
        for col in df.columns:
            col_info = {
                "name": col,
                "dtype": str(df[col].dtype),
                "non_null_count": int(df[col].notna().sum()),
                "null_count": int(df[col].isna().sum()),
                "unique_count": int(df[col].nunique()),
                "sample_values": df[col].dropna().head(5).tolist()
            }
            columns_info.append(col_info)
        return columns_info
