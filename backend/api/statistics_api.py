from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask
from typing import Optional
from api.auth import get_current_user
from models.models import User
from services import workflow_result_service as svc
import tempfile
import os
import json

router = APIRouter()


@router.get("/results/grouped")
async def get_results_grouped(
    current_user: User = Depends(get_current_user)
):
    grouped = await svc.get_results_grouped()
    return {"success": True, "data": grouped}


@router.get("/results/ranking-analysis")
async def get_ranking_analysis(
    result_id: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """返回涨幅排名完整数据供板块涨幅分析Tab使用"""
    grouped = await svc.get_results_grouped()
    ranking_list = grouped.get("涨幅排名", [])
    if not ranking_list:
        return {"success": True, "data": None, "available": []}

    available = [
        {"id": r["id"], "date_str": r["date_str"], "workflow_name": r["workflow_name"]}
        for r in ranking_list
    ]

    valid_ids = {r["id"] for r in ranking_list}
    target_id = result_id if result_id and result_id in valid_ids else ranking_list[0]["id"]
    full = await svc.get_result_full(target_id)
    if not full:
        return {"success": True, "data": None, "available": available}

    return {"success": True, "data": full, "available": available}


@router.get("/results/{result_id}/preview")
async def get_result_preview(
    result_id: int,
    current_user: User = Depends(get_current_user)
):
    result = await svc.get_result_preview(result_id)
    if not result:
        raise HTTPException(status_code=404, detail="结果不存在")
    return {"success": True, "data": result}


@router.get("/results/{result_id}/full")
async def get_result_full(
    result_id: int,
    current_user: User = Depends(get_current_user)
):
    result = await svc.get_result_full(result_id)
    if not result:
        raise HTTPException(status_code=404, detail="结果不存在")
    return {"success": True, "data": result}


@router.delete("/results/{result_id}")
async def delete_result(
    result_id: int,
    current_user: User = Depends(get_current_user)
):
    ok = await svc.delete_result(result_id)
    if not ok:
        raise HTTPException(status_code=500, detail="删除失败")
    return {"success": True, "message": "已删除"}


@router.get("/results/{result_id}/download")
async def download_result(
    result_id: int,
    current_user: User = Depends(get_current_user)
):
    """下载格式化 Excel，涨幅排名类型应用专属格式"""
    result = await svc.get_result_full(result_id)
    if not result:
        raise HTTPException(status_code=404, detail="结果不存在")

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(result['data'])
    filename = result.get('source_filename') or f"{result['workflow_name']}_{result['date_str']}.xlsx"
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()

    # 质押类型：直接复制 public 目录下的 final 文件（含双 sheet / 条件格式 / 列宽 / 筛选），
    # 然后用 stock_pools 全局最大公告日重做"最新公告日"红标。
    if result.get('workflow_type') == '质押':
        import shutil
        from services.path_resolver import get_resolver
        from core.config import settings
        base_dir = os.path.join(settings.DATA_DIR, "excel") if hasattr(settings, "DATA_DIR") else None
        # 兜底：尝试两种常见 base_dir
        for candidate in (base_dir, "/app/data/excel", "/Users/xiayanji/qbox/aistock/data/excel"):
            if candidate and os.path.isdir(candidate):
                base_dir = candidate
                break
        resolver = get_resolver(base_dir, "质押")
        public_dir = resolver.get_public_directory(result.get('date_str'))
        # 优先按 source_filename 命中；否则取 public 目录下第一个 xlsx
        src_path = None
        sf = result.get('source_filename')
        if sf:
            cand = os.path.join(public_dir, sf)
            if os.path.isfile(cand):
                src_path = cand
        if not src_path and os.path.isdir(public_dir):
            for f in os.listdir(public_dir):
                if f.lower().endswith('.xlsx'):
                    src_path = os.path.join(public_dir, f)
                    break
        if not src_path:
            raise HTTPException(status_code=404, detail=f"质押 public 文件不存在: {public_dir}")

        shutil.copy2(src_path, tmp_path)

        # 复用 finalize 端相同的 baseline + 着色逻辑（workflow_results 表 DB baseline）
        from services.workflow_executor import (
            load_pledge_announcement_max_ts,
            apply_announcement_date_coloring,
        )
        max_ts = load_pledge_announcement_max_ts(result.get('date_str'))
        if max_ts is not None:
            wb = load_workbook(tmp_path)
            apply_announcement_date_coloring(wb, max_ts)
            wb.save(tmp_path)

        return FileResponse(
            path=tmp_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            background=BackgroundTask(os.unlink, tmp_path),
        )

    df.to_excel(tmp_path, index=False, engine="openpyxl")

    # 涨幅排名: 应用专属格式（与工作流导出共享 apply_ranking_format）
    if result.get('workflow_type') == '涨幅排名':
        from datetime import datetime as dt
        from services.ranking_format import apply_ranking_format

        wb = load_workbook(tmp_path)
        ws = wb.active

        date_str = result.get('date_str', '')
        try:
            ref_date = dt.strptime(date_str, '%Y-%m-%d')
        except Exception:
            ref_date = dt.now()

        # 统计分析路径下 prev_rank 取 sheet 第 (date_start+1) 列的值（共享函数内部处理），
        # 所以这里 prev_rank_by_sector 传 None 即可
        apply_ranking_format(ws, prev_rank_by_sector=None, ref_date=ref_date)

        wb.save(tmp_path)

    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        background=BackgroundTask(os.unlink, tmp_path),
    )
