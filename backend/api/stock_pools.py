from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Body
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime
import json
import os
import logging
import tempfile
import pandas as pd
from core.database import get_async_db
from models.models import StockPool
from api.auth import get_current_user
from models.models import User
from services.pool_xlsx_importer import parse_pool_xlsx

router = APIRouter()


class StockPoolResponse(BaseModel):
    id: int
    name: str
    task_id: Optional[int] = None
    workflow_id: Optional[int] = None
    date_str: str = ""
    file_path: Optional[str] = None
    total_stocks: Optional[int] = None
    filter_conditions: Optional[list] = None
    source_types: Optional[list] = None
    is_active: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


@router.get("/", response_model=dict)
async def list_stock_pools(
    skip: int = 0,
    limit: int = 30,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user)
):
    # 按数据日期降序（最新在前），日期相同则 id 降序
    from sqlalchemy import func
    total_result = await db.execute(select(func.count()).select_from(StockPool))
    total = total_result.scalar() or 0
    result = await db.execute(
        select(StockPool)
        .order_by(StockPool.date_str.desc(), StockPool.id.desc())
        .offset(skip)
        .limit(limit)
    )
    items = result.scalars().all()
    return {
        "items": [StockPoolResponse.model_validate(p).model_dump() for p in items],
        "total": total,
        "skip": skip,
        "limit": limit,
    }


@router.get("/{stock_pool_id}", response_model=StockPoolResponse)
async def get_stock_pool(
    stock_pool_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user)
):
    result = await db.execute(
        select(StockPool).where(StockPool.id == stock_pool_id)
    )
    stock_pool = result.scalar_one_or_none()
    if not stock_pool:
        raise HTTPException(status_code=404, detail="选股池不存在")
    return stock_pool


@router.get("/{stock_pool_id}/data")
async def get_stock_pool_data(
    stock_pool_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user)
):
    """返回选股池完整数据"""
    result = await db.execute(
        select(StockPool).where(StockPool.id == stock_pool_id)
    )
    stock_pool = result.scalar_one_or_none()
    if not stock_pool:
        raise HTTPException(status_code=404, detail="选股池不存在")

    data = stock_pool.data or []
    # data 可能是 JSON 字符串或已解析的 list
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except Exception:
            data = []

    return {
        "id": stock_pool.id,
        "name": stock_pool.name,
        "date_str": stock_pool.date_str or "",
        "total_stocks": stock_pool.total_stocks or 0,
        "source_types": stock_pool.source_types or [],
        "filter_conditions": stock_pool.filter_conditions or [],
        "data": data
    }


@router.get("/{stock_pool_id}/download")
async def download_stock_pool(
    stock_pool_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user)
):
    result = await db.execute(
        select(StockPool).where(StockPool.id == stock_pool_id)
    )
    stock_pool = result.scalar_one_or_none()

    if not stock_pool:
        raise HTTPException(status_code=404, detail="选股池不存在")

    # 原始文件存在，直接返回
    if stock_pool.file_path and os.path.exists(stock_pool.file_path):
        return FileResponse(
            path=stock_pool.file_path,
            filename=os.path.basename(stock_pool.file_path),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # 文件不存在，从 DB 中的 data 字段重新生成 Excel + 新增行高亮
    data = stock_pool.data or []
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except Exception:
            data = []

    if not data:
        raise HTTPException(status_code=404, detail="文件不存在且无数据可导出")

    # 找基线：走 pool_cache，遍历 < date_str 所有日期并集
    baseline_codes = None
    try:
        from services.pool_cache import get_codes_before
        codes_union = await get_codes_before(stock_pool.date_str)
        if codes_union:
            baseline_codes = codes_union
    except Exception:
        baseline_codes = None

    date_part = (stock_pool.date_str or "").replace("-", "")
    filename = f"{stock_pool.name}-{date_part}.xlsx" if stock_pool.name else f"7条件交集{date_part}.xlsx"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        df = pd.DataFrame(data)
        df.to_excel(tmp_path, index=False, engine="openpyxl")

        # 应用居中 + 列宽 + 高亮
        from openpyxl import load_workbook as _lw
        from openpyxl.styles import Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        wb = _lw(tmp_path)
        ws = wb.active
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        green_fill = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
        highlight_cols = {"证券代码", "证券简称", "最新公告日"}
        headers = [c.value for c in ws[1]]
        # 列宽：资本运作行为 70；百日新高的日期 80；次数 30
        for idx, h in enumerate(headers, start=1):
            if not isinstance(h, str):
                continue
            if h == "资本运作行为":
                ws.column_dimensions[get_column_letter(idx)].width = 70
            elif "期间百日新高的日期" in h:
                ws.column_dimensions[get_column_letter(idx)].width = 80
            elif "期间百日新高次数" in h:
                ws.column_dimensions[get_column_letter(idx)].width = 30
        # 所有 cell 居中
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_align
        # 新增行高亮
        if baseline_codes is not None and "证券代码" in headers:
            code_col_idx = headers.index("证券代码") + 1
            target_col_indices = [headers.index(c) + 1 for c in headers if c in highlight_cols]
            for r_idx in range(2, ws.max_row + 1):
                code_val = ws.cell(row=r_idx, column=code_col_idx).value
                if code_val is None:
                    continue
                if str(code_val).strip() not in baseline_codes:
                    for c_idx in target_col_indices:
                        ws.cell(row=r_idx, column=c_idx).fill = green_fill
        wb.save(tmp_path)
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"生成Excel失败: stock_pool_id={stock_pool_id}, error={e}")
        os.unlink(tmp_path)
        raise HTTPException(status_code=500, detail="生成Excel失败")

    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        background=BackgroundTask(os.unlink, tmp_path),
    )


@router.delete("/{stock_pool_id}")
async def delete_stock_pool(
    stock_pool_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user)
):
    result = await db.execute(
        select(StockPool).where(StockPool.id == stock_pool_id)
    )
    stock_pool = result.scalar_one_or_none()

    if not stock_pool:
        raise HTTPException(status_code=404, detail="选股池不存在")

    if stock_pool.file_path and os.path.exists(stock_pool.file_path):
        os.remove(stock_pool.file_path)

    await db.delete(stock_pool)
    await db.commit()
    try:
        from services.pool_cache import invalidate as _inv
        _inv()
    except Exception:
        pass
    return {"message": "选股池已删除"}


# ==================== 批量导入（从 7_1 样式 xlsx） ====================

_IMPORT_CACHE_DIR = os.path.join(tempfile.gettempdir(), "stock_pool_imports")
os.makedirs(_IMPORT_CACHE_DIR, exist_ok=True)
_logger = logging.getLogger(__name__)


class ImportConfirmReq(BaseModel):
    token: str
    sheet_name: str


@router.post("/import/upload")
async def upload_import_xlsx(
    file: UploadFile = File(...),
    min_date: str = Query("2026-03-18"),
    current_user: User = Depends(get_current_user),
):
    """上传 xlsx 并解析所有合格 sheet，返回预览 + token。"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="只支持 .xlsx/.xls 文件")

    import uuid
    token = uuid.uuid4().hex[:16]
    xlsx_path = os.path.join(_IMPORT_CACHE_DIR, f"{token}.xlsx")
    try:
        content = await file.read()
        with open(xlsx_path, "wb") as f:
            f.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"保存上传文件失败: {e}")

    try:
        parsed = parse_pool_xlsx(xlsx_path, min_date=min_date)
    except Exception as e:
        _logger.exception("[pool_import] 解析失败")
        raise HTTPException(status_code=400, detail=f"解析失败: {e}")

    # 缓存 parsed 结果 json
    cache_path = os.path.join(_IMPORT_CACHE_DIR, f"{token}.json")
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(parsed, f, ensure_ascii=False, default=str)
    except Exception as e:
        _logger.exception("[pool_import] 缓存解析结果失败")
        raise HTTPException(status_code=500, detail=f"缓存失败: {e}")

    # 返回精简预览（不含 records 内容，只含元信息）
    preview = [
        {
            "sheet_name": s["sheet_name"],
            "date_str": s["date_str"],
            "style": s["style"],
            "record_count": s["record_count"],
            "raw_row_count": s["raw_row_count"],
        }
        for s in parsed
    ]
    return {"token": token, "sheets": preview}


@router.get("/import/sheet")
async def get_import_sheet_records(
    token: str = Query(...),
    sheet_name: str = Query(...),
    current_user: User = Depends(get_current_user),
):
    """返回某 sheet 的完整 records，用于前端预览数据详情。"""
    cache_path = os.path.join(_IMPORT_CACHE_DIR, f"{token}.json")
    if not os.path.exists(cache_path):
        raise HTTPException(status_code=404, detail="token 不存在或已过期")
    with open(cache_path, "r", encoding="utf-8") as f:
        parsed = json.load(f)
    for s in parsed:
        if s["sheet_name"] == sheet_name:
            return s
    raise HTTPException(status_code=404, detail=f"sheet {sheet_name} 不存在")


@router.post("/import/confirm")
async def confirm_import_sheet(
    req: ImportConfirmReq,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user),
):
    """将指定 sheet 的 records 入库（upsert by name+date_str）。
    name 按"{Y}年{MM}月选股池"自动生成。"""
    cache_path = os.path.join(_IMPORT_CACHE_DIR, f"{req.token}.json")
    if not os.path.exists(cache_path):
        raise HTTPException(status_code=404, detail="token 不存在或已过期")
    with open(cache_path, "r", encoding="utf-8") as f:
        parsed = json.load(f)

    target = None
    for s in parsed:
        if s["sheet_name"] == req.sheet_name:
            target = s
            break
    if target is None:
        raise HTTPException(status_code=404, detail=f"sheet {req.sheet_name} 不存在")

    date_str = target["date_str"]
    records = target["records"] or []
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        pool_name = f"{dt.year}年{dt.month:02d}月选股池"
    except Exception:
        pool_name = f"选股池_{date_str}"

    params = {
        "name": pool_name,
        "workflow_id": None,
        "date_str": date_str,
        "file_path": "",
        "total_stocks": len(records),
        "data": json.dumps(records, ensure_ascii=False, default=str),
        "filter_conditions": json.dumps([], ensure_ascii=False),
        "source_types": json.dumps(["xlsx_import"], ensure_ascii=False),
    }
    await db.execute(text("""
        INSERT INTO stock_pools (name, workflow_id, date_str, file_path, total_stocks, data,
            filter_conditions, source_types, is_active, created_at, updated_at)
        VALUES (:name, :workflow_id, :date_str, :file_path, :total_stocks, :data,
            :filter_conditions, :source_types, 1, NOW(), NOW())
        AS new_row
        ON DUPLICATE KEY UPDATE
            workflow_id = new_row.workflow_id,
            file_path = new_row.file_path,
            total_stocks = new_row.total_stocks,
            data = new_row.data,
            filter_conditions = new_row.filter_conditions,
            source_types = new_row.source_types,
            is_active = 1,
            updated_at = NOW()
    """), params)
    await db.commit()
    try:
        from services.pool_cache import invalidate as _inv
        _inv()
    except Exception:
        pass
    _logger.info(f"[pool_import] 入库: {pool_name} / {date_str} / {len(records)}条")
    return {
        "success": True,
        "sheet_name": req.sheet_name,
        "date_str": date_str,
        "pool_name": pool_name,
        "imported_count": len(records),
    }
