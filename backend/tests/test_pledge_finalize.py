import os
import tempfile
import pandas as pd
import pytest
from openpyxl import load_workbook
from services.workflow_executor import WorkflowExecutor


@pytest.fixture
def executor():
    with tempfile.TemporaryDirectory() as base:
        ex = WorkflowExecutor(base_dir=base, workflow_type="质押")
        yield ex


class TestFinalizeLayout:
    def test_column_order_with_all_prefix_cols(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"],
            "证券简称": ["平安银行"],
            "最新公告日": ["2026-04-20"],
            "百日新高": ["是"],
            "站上20日线": ["是"],
            "国央企": ["是"],
            "所属板块": ["金融"],
            "来源": ["中大盘"],
            "质押比例-20260118": [0.10],
            "质押比例-20260304": [0.12],
            "额外列": ["foo"],
        })
        output_path = tmp_path / "5质押20260420.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        assert wb.sheetnames == ["中大盘20260420", "小盘20260420"]
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        assert header[0] == "序号"
        assert header[1:8] == ["证券代码", "证券简称", "最新公告日", "百日新高", "站上20日线", "国央企", "所属板块"]
        assert "来源" not in header
        assert "额外列" in header
        assert "质押比例-20260118" in header

    def test_split_by_source(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ", "000002.SZ"],
            "证券简称": ["平安银行", "万科A"],
            "最新公告日": ["2026-04-20", "2026-04-20"],
            "来源": ["中大盘", "小盘"],
        })
        output_path = tmp_path / "5质押20260420.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        big = wb["中大盘20260420"]
        small = wb["小盘20260420"]
        assert big.max_row == 2
        assert big.cell(2, 2).value == "000001.SZ"
        assert small.max_row == 2
        assert small.cell(2, 2).value == "000002.SZ"

    def test_empty_source_keeps_header_only(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"],
            "证券简称": ["平安银行"],
            "最新公告日": ["2026-04-20"],
            "来源": ["中大盘"],
        })
        output_path = tmp_path / "5质押20260420.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        small = wb["小盘20260420"]
        assert small.max_row == 1

    def test_missing_prefix_cols_filled_empty(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"],
            "证券简称": ["平安银行"],
            "最新公告日": ["2026-04-20"],
            "来源": ["中大盘"],
        })
        output_path = tmp_path / "5质押20260420.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        assert "百日新高" in header
        assert "站上20日线" in header
        assert "国央企" in header
        assert "所属板块" in header

    def test_drops_original_xuhao(self, executor, tmp_path):
        df = pd.DataFrame({
            "序号": [99],
            "证券代码": ["000001.SZ"],
            "证券简称": ["平安银行"],
            "最新公告日": ["2026-04-20"],
            "来源": ["中大盘"],
        })
        output_path = tmp_path / "5质押20260420.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        assert ws.cell(2, 1).value == 1  # 新序号 1 起

    def test_non_default_index_no_source(self, executor, tmp_path):
        """df 有非默认索引 + 无来源列：不应 crash。"""
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"],
            "证券简称": ["A"],
            "最新公告日": ["2026-04-20"],
        }, index=[99])
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        # 无来源 → 全部归小盘
        small = wb["小盘20260420"]
        assert small.max_row == 2
        assert small.cell(2, 2).value == "000001.SZ"

    def test_duplicate_column_names(self, executor, tmp_path):
        """df 含重复列名：不应 crash，重复列自动追加 _N 后缀。"""
        # 构造一个带重复列名的 DataFrame
        df = pd.DataFrame([["000001.SZ", "A", "2026-04-20", "中大盘", "x", "y"]])
        df.columns = ["证券代码", "证券简称", "最新公告日", "来源", "重复列", "重复列"]
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        # 两个"重复列"都保留（一个原名一个带后缀）
        assert header.count("重复列") == 1
        assert "重复列_1" in header

    def test_drops_source_info_cols_exact(self, executor, tmp_path):
        """精确权威名（百日新高/站上20日线/国央企/所属板块）在 finalize 阶段**保留**——
        因为 merge 已经过滤源表的原始值，finalize 看到的权威名列都是 match_* 的产出。
        即使 match_* 未跑（值为空字符串），列仍保留。
        """
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"],
            "证券简称": ["A"],
            "最新公告日": ["2026-04-20"],
            "来源": ["中大盘"],
            # 模拟 merge 后、match_* 未跑：权威名列为空（由 _extract_columns_pledge 或
            # _reorder_pledge_columns prefix fill 补的）
            "百日新高": [""],
            "站上20日线": [""],
            "国央企": [""],
            "所属板块": [""],
            "质押比例-20260420": [0.10],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        for col_name in ("百日新高", "站上20日线", "国央企", "所属板块"):
            assert col_name in header  # 固定前 7 列保留
        # 质押比例列也应保留
        assert "质押比例-20260420" in header

    def test_drops_source_info_cols_synonyms(self, executor, tmp_path):
        """源表带 4 类信息列的同义词变体：也应被丢弃，不出现在剩余列里。"""
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"],
            "证券简称": ["A"],
            "最新公告日": ["2026-04-20"],
            "来源": ["中大盘"],
            # 同义词变体
            "百日最高价": ["源值"],       # → 百日新高
            "20日均线": ["源值"],         # → 站上20日线
            "20日线": ["源值2"],          # → 站上20日线（另一个变体）
            "国企": ["源值"],             # → 国央企
            "一级板块": ["源值"],         # → 所属板块
            "板块": ["源值2"],            # → 所属板块（另一个变体）
            # 非 4 类信息列：应保留
            "质押比例-20260420": [0.10],
            "额外业务列": ["保留"],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        for forbidden in ("百日最高价", "20日均线", "20日线", "国企", "一级板块", "板块"):
            assert forbidden not in header, f"{forbidden} 应被丢弃"
        # 质押比例和额外业务列应保留
        assert "质押比例-20260420" in header
        assert "额外业务列" in header

    def test_keeps_non_info_columns_intact(self, executor, tmp_path):
        """源表大量业务列（非 4 类信息）：应全部保留在剩余列区域，按源序。"""
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"],
            "证券简称": ["A"],
            "最新公告日": ["2026-04-20"],
            "来源": ["中大盘"],
            "股权质押公告日期-20260420": ["2026-04-20"],
            "质押比例-20260118": [0.08],
            "质押比例-20260304": [0.10],
            "质押比例-20260420": [0.12],
            "大股东名称": ["张三"],
            "质押方": ["某证券"],
            "备注": ["..."],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        for c in ("股权质押公告日期-20260420", "质押比例-20260118", "质押比例-20260304",
                  "质押比例-20260420", "大股东名称", "质押方", "备注"):
            assert c in header, f"{c} 应保留"

    def test_flatten_multiheader_with_newline(self, executor, tmp_path):
        """源表列名含 \\n（如'质押比例\\n[截止日期]2025-04-01'）应被规整为'质押比例2025-04-01'。"""
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"],
            "证券名称": ["A"],
            "股权质押公告日期\n[截止日期]最新": ["2026-04-18"],
            "质押比例\n[截止日期]2025-04-01": [10.56],
            "质押比例\n[截止日期]2025-05-01": [12.57],
            "质押比例\n[截止日期]2026-04-01": [11.58],
            "质押比例\n[截止日期]最新": [12.59],
            "总市值\n[单位]亿元": [153.09],
        })
        # 调展平方法（fake filepath，不会走 openpyxl 分支因为单行表头已含日期）
        df2 = executor._maybe_flatten_pledge_multiheader(df, "fake.xlsx", "fake_sheet")
        assert "质押比例2025-04-01" in df2.columns
        assert "质押比例2025-05-01" in df2.columns
        assert "质押比例2026-04-01" in df2.columns
        assert "质押比例最新" in df2.columns
        # 非质押比例列原样保留
        assert "证券代码" in df2.columns
        assert "总市值\n[单位]亿元" in df2.columns

    def test_final_output_styling(self, executor, tmp_path):
        """最终输出：两 sheet 都含 autofilter、列宽固定、居中对齐。"""
        df = pd.DataFrame({
            "证券代码": ["000001.SZ", "000002.SZ"],
            "证券简称": ["A", "B"],
            "最新公告日": ["2026-04-20", "2026-04-20"],
            "来源": ["中大盘", "小盘"],
            "质押比例-20260420": [0.10, 0.20],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        for sn in ("中大盘20260420", "小盘20260420"):
            ws = wb[sn]
            # autofilter
            assert ws.auto_filter.ref is not None, f"{sn} 没有 autofilter"
            # 列宽
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                width = ws.column_dimensions[col_letter].width
                assert width is not None and width >= 20, f"{sn} col {col_letter} 宽度过小: {width}"
            # 居中（至少一个数据单元格）
            if ws.max_row >= 2:
                cell = ws.cell(2, 2)  # 证券代码列第一行数据
                assert cell.alignment is not None
                assert cell.alignment.horizontal == "center", f"{sn} 未居中: {cell.alignment.horizontal}"


RED_COLOR = "FFFFA7A7"
GREEN_COLOR = "FFC6EFCE"
NEW_ROW_RED = "FFFFC7CE"  # 最新公告日"新行"浅红标记


def _get_fill(ws, row, col):
    c = ws.cell(row, col)
    if c.fill and c.fill.start_color:
        return (c.fill.start_color.rgb or "").upper()
    return ""


class TestPledgeRatioColoring:
    def test_ratio_increase_red(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
            "质押比例-20260118": [0.10],
            "质押比例-20260304": [0.15],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        col_right = header.index("质押比例-20260304") + 1
        assert RED_COLOR in _get_fill(ws, 2, col_right)

    def test_ratio_decrease_green(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
            "质押比例-20260118": [0.15],
            "质押比例-20260304": [0.10],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        col_right = header.index("质押比例-20260304") + 1
        assert GREEN_COLOR in _get_fill(ws, 2, col_right)

    def test_ratio_equal_no_color(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
            "质押比例-20260118": [0.10],
            "质押比例-20260304": [0.10],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        col_right = header.index("质押比例-20260304") + 1
        fill = _get_fill(ws, 2, col_right)
        assert RED_COLOR not in fill and GREEN_COLOR not in fill

    def test_ratio_either_empty_skipped(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ", "000002.SZ"],
            "证券简称": ["A", "B"],
            "最新公告日": ["2026-04-20", "2026-04-20"],
            "来源": ["中大盘", "中大盘"],
            "质押比例-20260118": [None, 0.10],
            "质押比例-20260304": [0.15, None],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        col_right = header.index("质押比例-20260304") + 1
        for row in (2, 3):
            fill = _get_fill(ws, row, col_right)
            assert RED_COLOR not in fill and GREEN_COLOR not in fill

    def test_ratio_leftmost_never_colored(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
            "质押比例-20260118": [0.10],
            "质押比例-20260304": [0.15],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        col_left = header.index("质押比例-20260118") + 1
        fill = _get_fill(ws, 2, col_left)
        assert RED_COLOR not in fill and GREEN_COLOR not in fill

    def test_ratio_with_percent_string(self, executor, tmp_path):
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
            "质押比例-20260118": ["10.0%"],
            "质押比例-20260304": ["15.0%"],
        })
        output_path = tmp_path / "out.xlsx"
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        col_right = header.index("质押比例-20260304") + 1
        assert RED_COLOR in _get_fill(ws, 2, col_right)


class TestPledgeFirstAppearance:
    def _write_public_baseline(self, public_dir, code, date_val, date_str):
        """写一份历史 public 文件，含某代码的某公告日。"""
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(f"中大盘{date_str}")
        ws.append(["序号", "证券代码", "证券简称", "最新公告日"])
        ws.append([1, code, "A", date_val])
        ws2 = wb.create_sheet(f"小盘{date_str}")
        ws2.append(["序号", "证券代码", "证券简称", "最新公告日"])
        wb.save(str(public_dir / f"5质押{date_str}.xlsx"))

    def test_new_code_green(self, executor, tmp_path):
        """public 为空（baseline=None）→ 不标红（保守行为，避免首次运行全红）。"""
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
        })
        output_path = tmp_path / "out.xlsx"
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        col_date = header.index("最新公告日") + 1
        # 无 baseline → 不标
        assert NEW_ROW_RED not in _get_fill(ws, 2, col_date)

    def test_existing_code_newer_date_green(self, executor, tmp_path):
        """public 有基准，新公告日 > baseline → 标红。"""
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        self._write_public_baseline(public_dir, "000001.SZ", "2026-03-01", "20260301")
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
        })
        output_path = tmp_path / "out.xlsx"
        executor._finalize_pledge_output(df, "20260420", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260420"]
        header = [c.value for c in ws[1]]
        col_date = header.index("最新公告日") + 1
        assert NEW_ROW_RED in _get_fill(ws, 2, col_date)

    def test_existing_code_same_or_older_not_green(self, executor, tmp_path):
        """公告日 == baseline → 不标红（只看 >，不看 >=）。"""
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        self._write_public_baseline(public_dir, "000001.SZ", "2026-04-20", "20260420")
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
        })
        output_path = tmp_path / "out.xlsx"
        executor._finalize_pledge_output(df, "20260421", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["中大盘20260421"]
        header = [c.value for c in ws[1]]
        col_date = header.index("最新公告日") + 1
        fill = _get_fill(ws, 2, col_date)
        assert NEW_ROW_RED not in fill

    def test_baseline_merges_both_sheets(self, executor, tmp_path):
        """按 sheet 分基准：public 小盘 sheet 为空 → 小盘 baseline=None → 不标红。
        中大盘 sheet 的数据不影响小盘的判定。"""
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        from openpyxl import Workbook
        wb_pub = Workbook()
        wb_pub.remove(wb_pub.active)
        ws_big = wb_pub.create_sheet("中大盘20260301")
        ws_big.append(["序号", "证券代码", "证券简称", "最新公告日"])
        ws_big.append([1, "000002.SZ", "B", "2026-04-20"])  # 中大盘 baseline=2026-04-20
        ws_small = wb_pub.create_sheet("小盘20260301")
        ws_small.append(["序号", "证券代码", "证券简称", "最新公告日"])  # 小盘 baseline=None
        wb_pub.save(str(public_dir / "5质押20260301.xlsx"))
        df = pd.DataFrame({
            "证券代码": ["000002.SZ"], "证券简称": ["B"],
            "最新公告日": ["2026-04-20"], "来源": ["小盘"],
        })
        output_path = tmp_path / "out.xlsx"
        executor._finalize_pledge_output(df, "20260421", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws = wb["小盘20260421"]
        header = [c.value for c in ws[1]]
        col_date = header.index("最新公告日") + 1
        # 小盘 baseline=None → 不标（新行为）
        assert NEW_ROW_RED not in _get_fill(ws, 2, col_date)

    def test_global_baseline_only_newer_dates_green(self, executor, tmp_path):
        """中大盘 baseline 仅对中大盘行生效：公告日 > baseline 才标红，= 不标，< 不标。"""
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        from openpyxl import Workbook
        wb_pub = Workbook()
        wb_pub.remove(wb_pub.active)
        ws = wb_pub.create_sheet("中大盘20260410")
        ws.append(["序号", "证券代码", "证券简称", "最新公告日"])
        ws.append([1, "000001.SZ", "A", "2026-04-10"])
        ws.append([2, "000002.SZ", "B", "2026-04-05"])
        ws2 = wb_pub.create_sheet("小盘20260410")
        ws2.append(["序号", "证券代码", "证券简称", "最新公告日"])
        wb_pub.save(str(public_dir / "5质押20260410.xlsx"))

        df = pd.DataFrame({
            "证券代码": ["000003.SZ", "000001.SZ", "000004.SZ"],
            "证券简称": ["C", "A", "D"],
            "最新公告日": ["2026-04-21", "2026-04-09", "2026-04-10"],
            "来源": ["中大盘", "中大盘", "中大盘"],
        })
        output_path = tmp_path / "out.xlsx"
        executor._finalize_pledge_output(df, "20260421", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        ws_big = wb["中大盘20260421"]
        header = [c.value for c in ws_big[1]]
        col_date = header.index("最新公告日") + 1
        # 2026-04-21 > baseline 2026-04-10 → 标红
        assert NEW_ROW_RED in _get_fill(ws_big, 2, col_date)
        # 2026-04-09 < baseline → 不标
        assert NEW_ROW_RED not in _get_fill(ws_big, 3, col_date)
        # 2026-04-10 == baseline → 不标
        assert NEW_ROW_RED not in _get_fill(ws_big, 4, col_date)

    def test_per_sheet_baseline_isolation(self, executor, tmp_path):
        """中大盘和小盘基准独立：小盘行用小盘 baseline，不受中大盘 baseline 影响。"""
        public_dir = tmp_path / "public"
        public_dir.mkdir()
        from openpyxl import Workbook
        wb_pub = Workbook()
        wb_pub.remove(wb_pub.active)
        ws_big = wb_pub.create_sheet("中大盘20260420")
        ws_big.append(["序号", "证券代码", "证券简称", "最新公告日"])
        ws_big.append([1, "000001.SZ", "A", "2026-04-20"])
        ws_small = wb_pub.create_sheet("小盘20260420")
        ws_small.append(["序号", "证券代码", "证券简称", "最新公告日"])
        ws_small.append([1, "000099.SZ", "X", "2026-04-05"])
        wb_pub.save(str(public_dir / "5质押20260420.xlsx"))

        df = pd.DataFrame({
            "证券代码": ["000010.SZ", "000020.SZ"],
            "证券简称": ["M", "S"],
            "最新公告日": ["2026-04-10", "2026-04-10"],
            "来源": ["中大盘", "小盘"],
        })
        output_path = tmp_path / "out.xlsx"
        executor._finalize_pledge_output(df, "20260421", str(output_path), str(public_dir))
        wb = load_workbook(str(output_path))
        # 中大盘：2026-04-10 < 中大盘 baseline 2026-04-20 → 不标
        ws_big_new = wb["中大盘20260421"]
        h_big = [c.value for c in ws_big_new[1]]
        dc_big = h_big.index("最新公告日") + 1
        assert NEW_ROW_RED not in _get_fill(ws_big_new, 2, dc_big), "中大盘行不该标红（< 中大盘基准）"
        # 小盘：2026-04-10 > 小盘 baseline 2026-04-05 → 标红
        ws_small_new = wb["小盘20260421"]
        h_small = [c.value for c in ws_small_new[1]]
        dc_small = h_small.index("最新公告日") + 1
        assert NEW_ROW_RED in _get_fill(ws_small_new, 2, dc_small), "小盘行该标红（> 小盘基准）"


class TestFinalizePledgeIfNeeded:
    def test_non_pledge_noop(self, tmp_path):
        ex = WorkflowExecutor(base_dir=str(tmp_path), workflow_type="并购重组")
        result = ex.finalize_pledge_if_needed(
            last_output_path=str(tmp_path / "nonexistent.xlsx"),
            date_str="20260420",
        )
        assert result is False

    def test_pledge_runs_finalize_and_syncs(self, tmp_path, monkeypatch):
        ex = WorkflowExecutor(base_dir=str(tmp_path), workflow_type="质押")
        daily_dir = tmp_path / "data" / "excel" / "质押" / "20260420"
        daily_dir.mkdir(parents=True)
        public_dir = tmp_path / "data" / "excel" / "质押" / "public"
        public_dir.mkdir(parents=True)
        df = pd.DataFrame({
            "证券代码": ["000001.SZ"], "证券简称": ["A"],
            "最新公告日": ["2026-04-20"], "来源": ["中大盘"],
        })
        # 用最终文件名命名（模拟 match_sector 老代码已写出的 final 文件）
        last_output = daily_dir / "5质押20260420.xlsx"
        df.to_excel(str(last_output), index=False)

        monkeypatch.setattr(ex.resolver, "get_daily_dir", lambda d=None: str(daily_dir))
        monkeypatch.setattr(ex.resolver, "get_public_directory", lambda d=None: str(public_dir))

        result = ex.finalize_pledge_if_needed(
            last_output_path=str(last_output),
            date_str="20260420",
        )
        assert result is True
        # finalize 必须覆盖 last_output_path 本身（而非另起新文件），
        # 这样 run_workflow 的 DB 保存、前端下载看到的就是 finalize 后的 2 sheet 文件
        assert last_output.exists()
        from openpyxl import load_workbook as _lwb
        wb = _lwb(str(last_output))
        assert wb.sheetnames == ["中大盘20260420", "小盘20260420"]
        # public 同步（保持同名）
        assert (public_dir / "5质押20260420.xlsx").exists()

    def test_finalize_overwrites_match_sector_single_sheet_output(self, tmp_path, monkeypatch):
        """match_sector 写出的单 sheet final 文件，含来源列 + 原始列 + 未权威填充的信息列，
        finalize 必须：读全部 sheet 合并；覆盖原文件为双 sheet；去来源列；保留质押比例；丢 4 类信息列。"""
        ex = WorkflowExecutor(base_dir=str(tmp_path), workflow_type="质押")
        daily_dir = tmp_path / "data" / "excel" / "质押" / "20260420"
        daily_dir.mkdir(parents=True)
        public_dir = tmp_path / "data" / "excel" / "质押" / "public"
        public_dir.mkdir(parents=True)
        df = pd.DataFrame({
            "序号": [1, 2],
            "证券代码": ["000001.SZ", "600000.SH"],
            "证券简称": ["A", "B"],
            "最新公告日": ["2026-04-20", "2026-04-19"],
            "来源": ["中大盘", "小盘"],
            "百日新高": ["是", ""],
            "20日均线": ["是", "否"],           # 同义词变体 → 应丢弃
            "国企": ["否", "是"],                # 同义词变体 → 应丢弃
            "所属板块": ["金融", "银行"],
            "质押比例-20260118": [0.10, 0.15],
            "质押比例-20260420": [0.12, 0.10],
            "股权质押公告日期-20260420": ["2026-04-20", "2026-04-19"],
            "额外业务列": ["x", "y"],
        })
        last_output = daily_dir / "5质押20260420.xlsx"
        df.to_excel(str(last_output), index=False)

        monkeypatch.setattr(ex.resolver, "get_daily_dir", lambda d=None: str(daily_dir))
        monkeypatch.setattr(ex.resolver, "get_public_directory", lambda d=None: str(public_dir))

        result = ex.finalize_pledge_if_needed(
            last_output_path=str(last_output),
            date_str="20260420",
        )
        assert result is True

        from openpyxl import load_workbook as _lwb
        wb = _lwb(str(last_output))
        # Bug 1 修：双 sheet
        assert wb.sheetnames == ["中大盘20260420", "小盘20260420"]
        ws_big = wb["中大盘20260420"]
        header_big = [c.value for c in ws_big[1]]
        # Bug 2 修：来源列不出现
        assert "来源" not in header_big
        # Bug 3 修：质押比例列保留
        assert "质押比例-20260118" in header_big
        assert "质押比例-20260420" in header_big
        assert "股权质押公告日期-20260420" in header_big
        assert "额外业务列" in header_big
        # 4 类信息列同义词变体已丢弃
        assert "20日均线" not in header_big
        assert "国企" not in header_big
        # 中大盘 sheet 只含中大盘来源的行
        assert ws_big.max_row == 2
        assert ws_big.cell(2, 2).value == "000001.SZ"
        # 小盘 sheet 只含小盘来源的行
        ws_small = wb["小盘20260420"]
        assert ws_small.max_row == 2
        assert ws_small.cell(2, 2).value == "600000.SH"
