import os
import tempfile
import shutil
import pytest
import pandas as pd
from openpyxl import load_workbook
from unittest.mock import patch

from services.workflow_executor import WorkflowExecutor


@pytest.fixture
def tmp_base():
    d = tempfile.mkdtemp()
    # 准备好涨幅排名需要的目录结构
    os.makedirs(os.path.join(d, "涨幅排名", "2026-04-22"), exist_ok=True)
    os.makedirs(os.path.join(d, "涨幅排名", "2026-04-22", "public"), exist_ok=True)
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _build_input_df(include_ytd_mtd: bool = True, n_rows: int = 6):
    """构造 ranking_sort 的 input_data（merge_excel 的输出）。
    列 0=板块名称, 列 1=今日涨跌幅(多行), 之后随意列，含"本年初"/"本月初"多行列名。"""
    cols = {
        "板块名称": [f"板块{i+1}" for i in range(n_rows)],
        "成份区间涨跌幅(算术平均)\n[起始交易日期] 2026-04-22\n[终止交易日期] 2026-04-22": [
            5.0, 3.0, 1.0, -1.0, -3.0, -5.0
        ][:n_rows],
    }
    if include_ytd_mtd:
        cols["成份区间涨跌幅(算术平均)\n[起始交易日期] 本年初\n[终止交易日期] 2026-04-22"] = [
            39.59, 27.0, -3.09, 32.0, "--", ""
        ][:n_rows]
        cols["成份区间涨跌幅(算术平均)\n[起始交易日期] 本月初\n[终止交易日期] 2026-04-22"] = [
            10.0, 20.0, 5.0, 15.0, 25.0, None
        ][:n_rows]
    return pd.DataFrame(cols)


@pytest.mark.asyncio
async def test_extended_layout_basic(tmp_base):
    """启用扩展布局：列顺序 + YTD/MTD 排名 + 文件结构"""
    exe = WorkflowExecutor(base_dir=tmp_base, workflow_type="涨幅排名")
    df = _build_input_df(include_ytd_mtd=True, n_rows=6)

    with patch.object(exe.resolver, "find_previous_public_file", return_value=(None, None)):
        res = await exe._ranking_sort(
            {"date_str": "2026-04-22"}, df, "2026-04-22"
        )

    assert res["success"] is True
    cols = res["columns"]
    # 顺序：板块 | 年初 | B排名 | 月初 | D排名 | 今日 | 次数 | 日期列
    assert cols[0] == "板块名称"
    assert cols[1] == "年初涨跌幅"
    assert cols[2] == "B列的数值升序排序结果"
    assert cols[3] == "月初涨跌幅"
    assert cols[4] == "D列的数值升序排序结果"
    assert cols[5] == "今日涨跌幅"
    assert cols[6] == "迄今为止排进前5(次数)"
    assert cols[7] == "4月22日"

    # 打开文件验证排名数字正确
    wb = load_workbook(res["file_path"], data_only=True)
    ws = wb.active
    # 读出所有行数据（row1 表头）
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    # 第二列=年初值, 第三列=年初排名
    # 原值 [39.59, 27, -3.09, 32, "--", ""] → 降序排名 39.59(1) 32(2) 27(3) -3.09(4) + 2 个 invalid (5, 6)
    # 但注意 rows 按今日涨跌幅降序排过 (5,3,1,-1,-3,-5)，板块1-6 行顺序不变
    ytd_values = [r[1] for r in data_rows]
    ytd_ranks = [r[2] for r in data_rows]

    # 板块1 年初=39.59 → rank=1; 板块2 年初=27 → rank=3; 板块3 年初=-3.09 → rank=4;
    # 板块4 年初=32 → rank=2; 板块5 年初="--" → rank=5; 板块6 年初="" → rank=6
    # 今日排序后顺序：5,3,1,-1,-3,-5 对应板块1,2,3,4,5,6
    assert ytd_ranks[0] == 1  # 板块1
    assert ytd_ranks[1] == 3  # 板块2
    assert ytd_ranks[2] == 4  # 板块3
    assert ytd_ranks[3] == 2  # 板块4
    assert ytd_ranks[4] == 5  # 板块5
    assert ytd_ranks[5] == 6  # 板块6


@pytest.mark.asyncio
async def test_legacy_layout_when_no_ytd_mtd(tmp_base):
    """没有本年初/本月初列：走老逻辑，保持原输出格式"""
    exe = WorkflowExecutor(base_dir=tmp_base, workflow_type="涨幅排名")
    df = _build_input_df(include_ytd_mtd=False, n_rows=3)

    with patch.object(exe.resolver, "find_previous_public_file", return_value=(None, None)):
        res = await exe._ranking_sort(
            {"date_str": "2026-04-22"}, df, "2026-04-22"
        )

    assert res["success"] is True
    cols = res["columns"]
    # 老布局：板块 | 今日 | 次数 | 日期
    assert cols[0] == "板块名称"
    assert cols[1] == "今日涨跌幅"
    assert cols[2] == "迄今为止排进前5(次数)"
    assert cols[3] == "4月22日"
    # 没有年初/月初
    assert "年初涨跌幅" not in cols
    assert "月初涨跌幅" not in cols


@pytest.mark.asyncio
async def test_extended_layout_only_ytd_falls_back_to_legacy(tmp_base):
    """只有年初没有月初：回退老逻辑"""
    exe = WorkflowExecutor(base_dir=tmp_base, workflow_type="涨幅排名")
    df = _build_input_df(include_ytd_mtd=False, n_rows=3)
    df["成份区间涨跌幅(算术平均)\n[起始交易日期] 本年初"] = [1.0, 2.0, 3.0]

    with patch.object(exe.resolver, "find_previous_public_file", return_value=(None, None)):
        res = await exe._ranking_sort(
            {"date_str": "2026-04-22"}, df, "2026-04-22"
        )

    assert res["success"] is True
    assert "年初涨跌幅" not in res["columns"]


@pytest.mark.asyncio
async def test_column_order_year_month_today(tmp_base):
    """源文件列顺序为 [板块, 本年初, 本月初, 今日] 时正确识别三列

    回归 bug1：之前硬编码 cols[1] 为今日列，真实数据中 cols[1] 是"本年初"
    时会把年初值当今日排序，颜色标记全错。
    """
    exe = WorkflowExecutor(base_dir=tmp_base, workflow_type="涨幅排名")
    df = pd.DataFrame({
        "板块名称": ["A", "B", "C"],
        "成份区间涨跌幅(算术平均)\n[起始交易日期]本年初\n[截止交易日期]最新": [-13.35, 11.44, 38.78],
        "成份区间涨跌幅(算术平均)\n[起始交易日期]本月初\n[截止交易日期]最新": [0.49, 11.36, -3.98],
        "成份区间涨跌幅(算术平均)\n[起始交易日期]2026-04-23\n[截止交易日期]2026-04-23": [2.84, 2.37, 2.26],
    })

    with patch.object(exe.resolver, "find_previous_public_file", return_value=(None, None)):
        res = await exe._ranking_sort(
            {"date_str": "2026-04-23"}, df, "2026-04-23"
        )

    assert res["success"] is True
    cols = res["columns"]
    assert cols[0] == "板块名称"
    data_rows = res["data"]
    sectors_ordered = [r["板块名称"] for r in data_rows]
    # 今日涨跌幅 [2.84, 2.37, 2.26] 降序 → A, B, C
    assert sectors_ordered == ["A", "B", "C"]
    today_values = [r["今日涨跌幅"] for r in data_rows]
    assert today_values == [2.84, 2.37, 2.26]
    ytd_values = [r["年初涨跌幅"] for r in data_rows]
    assert ytd_values == [-13.35, 11.44, 38.78]
    mtd_values = [r["月初涨跌幅"] for r in data_rows]
    assert mtd_values == [0.49, 11.36, -3.98]


@pytest.mark.asyncio
async def test_values_rounded_to_2_decimals(tmp_base):
    """今日/年初/月初三列都保留 2 位小数；非数字（如 '--'）原样保留"""
    exe = WorkflowExecutor(base_dir=tmp_base, workflow_type="涨幅排名")
    df = pd.DataFrame({
        "板块名称": ["P1", "P2", "P3"],
        "成份区间涨跌幅(算术平均)\n[起始交易日期]本年初": [1.23456, -7.891234, 0.1],
        "成份区间涨跌幅(算术平均)\n[起始交易日期]本月初": [2.99999, "--", 3.005],
        "成份区间涨跌幅(算术平均)\n[起始交易日期]2026-04-23": [5.5555, 4.4444, 3.3333],
    })

    with patch.object(exe.resolver, "find_previous_public_file", return_value=(None, None)):
        res = await exe._ranking_sort(
            {"date_str": "2026-04-23"}, df, "2026-04-23"
        )

    assert res["success"] is True
    data_rows = res["data"]

    today_values = [r["今日涨跌幅"] for r in data_rows]
    assert today_values == [5.56, 4.44, 3.33]

    ytd_values = [r["年初涨跌幅"] for r in data_rows]
    assert ytd_values == [1.23, -7.89, 0.1]

    mtd_values = [r["月初涨跌幅"] for r in data_rows]
    assert mtd_values[0] == 3.0
    assert mtd_values[1] == "--"
    assert mtd_values[2] in (3.0, 3.01)
