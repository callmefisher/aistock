import os
import tempfile
import shutil
import pytest
from openpyxl import load_workbook

from services.trend_service import export_trend_excel_with_chart


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp()
    yield d
    shutil.rmtree(d, ignore_errors=True)


def test_high_price_基本结构(tmp_dir):
    data = [
        {"workflow_type": "百日新高总趋势", "date_str": "2026-04-01", "count": 99, "total": 0, "ratio": 0},
        {"workflow_type": "百日新高总趋势", "date_str": "2026-04-02", "count": 105, "total": 0, "ratio": 0},
        {"workflow_type": "百日新高总趋势", "date_str": "2026-04-03", "count": 120, "total": 0, "ratio": 0},
    ]
    fp = os.path.join(tmp_dir, "t.xlsx")
    export_trend_excel_with_chart(data, fp, metric_type="high_price")
    assert os.path.exists(fp)
    wb = load_workbook(fp, data_only=True)
    assert "百日新高总趋势" in wb.sheetnames
    ws = wb["百日新高总趋势"]
    rows = list(ws.iter_rows(values_only=True))
    # Row0: 标题(可能合并), Row1: 表头, Row2+ 数据
    headers = rows[1]
    assert headers[0] == "日期"
    assert headers[1] == "数量"
    assert headers[2] == "完整日期"
    # 无占比/总量
    assert "占比(%)" not in headers
    assert "总量" not in headers
    # 数据正确
    assert rows[2][1] == 99
    assert rows[3][1] == 105
    assert rows[4][1] == 120
    wb.close()


def test_high_price_空数据(tmp_dir):
    fp = os.path.join(tmp_dir, "t.xlsx")
    export_trend_excel_with_chart([], fp, metric_type="high_price")
    assert os.path.exists(fp)
    wb = load_workbook(fp, data_only=True)
    ws = wb["百日新高总趋势"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "暂无数据"
    wb.close()


def test_high_price_x轴间隔(tmp_dir):
    """>15 点时应设 interval_unit；我们只验证不报错 + 数据行数正确"""
    data = [
        {"workflow_type": "百日新高总趋势",
         "date_str": f"2026-04-{i:02d}", "count": 100 + i, "total": 0, "ratio": 0}
        for i in range(1, 21)
    ]
    fp = os.path.join(tmp_dir, "t.xlsx")
    export_trend_excel_with_chart(data, fp, metric_type="high_price")
    assert os.path.exists(fp)
    wb = load_workbook(fp, data_only=True)
    ws = wb["百日新高总趋势"]
    rows = list(ws.iter_rows(values_only=True))
    # 标题1 + 表头1 + 数据20
    assert len(rows) >= 22
    wb.close()


def test_ma20_回归不受影响(tmp_dir):
    """确认默认 metric_type='ma20' 仍走原逻辑"""
    data = [
        {"workflow_type": "并购重组", "date_str": "2026-04-01",
         "count": 100, "total": 5000, "ratio": 0.02},
    ]
    fp = os.path.join(tmp_dir, "t.xlsx")
    export_trend_excel_with_chart(data, fp)
    assert os.path.exists(fp)
    wb = load_workbook(fp, data_only=True)
    assert "站上20日均线趋势" in wb.sheetnames
    wb.close()
