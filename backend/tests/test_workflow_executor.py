import pytest
import os
import tempfile
import pandas as pd
from datetime import datetime

from services.workflow_executor import WorkflowExecutor


class TestWorkflowExecutor:
    """工作流执行器单元测试"""

    @pytest.fixture
    def temp_excel_dir(self):
        """创建临时Excel目录"""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.fixture
    def executor(self, temp_excel_dir):
        """创建工作流执行器实例"""
        return WorkflowExecutor(base_dir=temp_excel_dir)

    @pytest.fixture
    def sample_excel_file(self, temp_excel_dir):
        """创建示例Excel文件"""
        df = pd.DataFrame({
            "证券代码": ["002128.SZ", "600519.SH", "000001.SZ"],
            "证券简称": ["露天煤业", "贵州茅台", "平安银行"],
            "最新公告日": ["2026-04-01", "2026-04-09", "2026-04-05"]
        })
        filepath = os.path.join(temp_excel_dir, "sample.xlsx")
        df.to_excel(filepath, index=False)
        return filepath

    @pytest.fixture
    def sample_excel_with_date(self, temp_excel_dir):
        """创建带序号的示例Excel文件"""
        df = pd.DataFrame({
            "序号": [1, 2, 3],
            "证券代码": ["002128.SZ", "600519.SH", "000001.SZ"],
            "证券简称": ["露天煤业", "贵州茅台", "平安银行"],
            "最新公告日": ["2026-04-01", "2026-04-09", "2026-04-05"]
        })
        filepath = os.path.join(temp_excel_dir, "sample.xlsx")
        df.to_excel(filepath, index=False)
        return filepath

    def test_get_daily_dir(self, executor, temp_excel_dir):
        """测试获取当日目录"""
        daily_dir = executor._get_daily_dir()
        assert daily_dir == temp_excel_dir

    def test_resolve_path_absolute(self, executor):
        """测试绝对路径解析"""
        abs_path = "/absolute/path/file.xlsx"
        resolved = executor._resolve_path(abs_path)
        assert resolved == abs_path

    def test_resolve_path_relative(self, executor, temp_excel_dir):
        """测试相对路径解析"""
        rel_path = "file.xlsx"
        resolved = executor._resolve_path(rel_path)
        expected = os.path.join(temp_excel_dir, "file.xlsx")
        assert resolved == expected

    def test_get_excel_files_in_dir(self, executor, sample_excel_file):
        """测试获取目录下的Excel文件"""
        daily_dir = os.path.dirname(sample_excel_file)
        files = executor._get_excel_files_in_dir(daily_dir)
        assert len(files) >= 1
        assert any("sample.xlsx" in f for f in files)

    def test_import_excel_success(self, executor, sample_excel_file):
        """测试成功导入Excel"""
        import asyncio
        config = {"file_path": sample_excel_file}
        result = asyncio.run(executor._import_excel(config))

        assert result["success"] is True
        assert result["rows"] == 3
        assert "data" in result

    def test_import_excel_not_found(self, executor):
        """测试导入不存在的Excel"""
        import asyncio
        config = {"file_path": "/nonexistent/file.xlsx"}
        result = asyncio.run(executor._import_excel(config))

        assert result["success"] is False
        assert "不存在" in result["message"]

    def test_import_excel_no_path(self, executor):
        """测试未指定文件路径"""
        import asyncio
        config = {}
        result = asyncio.run(executor._import_excel(config))

        assert result["success"] is False
        assert "未指定" in result["message"]

    def test_merge_excel_success(self, executor, temp_excel_dir):
        """测试成功合并Excel"""
        import asyncio

        df1 = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        df2 = pd.DataFrame({"A": [5, 6], "B": [7, 8]})

        df1.to_excel(os.path.join(temp_excel_dir, "file1.xlsx"), index=False)
        df2.to_excel(os.path.join(temp_excel_dir, "file2.xlsx"), index=False)

        config = {"output_filename": "total_1.xlsx"}
        result = asyncio.run(executor._merge_excel(config))

        assert result["success"] is True
        assert result["files_merged"] == 2
        assert result["rows"] == 4
        assert os.path.exists(result["file_path"])

    def test_merge_excel_excludes_total(self, executor, temp_excel_dir):
        """测试合并时排除total_开头的文件"""
        import asyncio

        df1 = pd.DataFrame({"A": [1, 2]})
        df2 = pd.DataFrame({"A": [3, 4]})

        df1.to_excel(os.path.join(temp_excel_dir, "file1.xlsx"), index=False)
        df2.to_excel(os.path.join(temp_excel_dir, "total_1.xlsx"), index=False)

        config = {"output_filename": "merged.xlsx"}
        result = asyncio.run(executor._merge_excel(config))

        assert result["success"] is True
        assert result["files_merged"] == 1

    def test_merge_excel_excludes_output(self, executor, temp_excel_dir):
        """测试合并时排除output_开头的文件"""
        import asyncio

        df1 = pd.DataFrame({"A": [1, 2]})
        df2 = pd.DataFrame({"A": [3, 4]})

        df1.to_excel(os.path.join(temp_excel_dir, "file1.xlsx"), index=False)
        df2.to_excel(os.path.join(temp_excel_dir, "output_1.xlsx"), index=False)

        config = {"output_filename": "merged.xlsx"}
        result = asyncio.run(executor._merge_excel(config))

        assert result["success"] is True
        assert result["files_merged"] == 1

    def test_merge_excel_no_files(self, executor, temp_excel_dir):
        """测试合并时没有Excel文件"""
        import asyncio
        config = {"output_filename": "total_1.xlsx"}
        result = asyncio.run(executor._merge_excel(config))

        assert result["success"] is False
        assert "没有找到" in result["message"]

    def test_dedup_success(self, executor):
        """测试简单去重"""
        import asyncio

        df = pd.DataFrame({
            "A": [1, 2, 2, 3],
            "B": [4, 5, 5, 6]
        })

        config = {}
        result = asyncio.run(executor._dedup(config, df))

        assert result["success"] is True
        assert result["original_rows"] == 4
        assert result["deduped_rows"] == 3
        assert result["removed_rows"] == 1

    def test_dedup_no_data(self, executor):
        """测试去重无数据"""
        import asyncio

        config = {}
        result = asyncio.run(executor._dedup(config, None))

        assert result["success"] is False
        assert "没有可处理" in result["message"]

    def test_smart_dedup_success(self, executor):
        """测试智能去重（按证券代码和最新公告日）"""
        import asyncio

        df = pd.DataFrame({
            "证券代码": ["002128.SZ", "002128.SZ", "002128.SZ"],
            "证券简称": ["A", "B", "C"],
            "最新公告日": ["2026-04-09", "2026-04-01", "2026-05-01"]
        })

        config = {}
        result = asyncio.run(executor._smart_dedup(config, df))

        assert result["success"] is True
        assert result["original_rows"] == 3
        assert result["deduped_rows"] == 1

        final_df = result["data"]
        assert final_df.iloc[0]["最新公告日"] == "2026-05-01"

    def test_smart_dedup_auto_detect_columns(self, executor):
        """测试智能去重自动检测列名"""
        import asyncio

        df = pd.DataFrame({
            "股票代码": ["001", "001", "002"],
            "名称": ["A", "A", "B"],
            "公告日期": ["2026-04-01", "2026-04-02", "2026-04-03"]
        })

        config = {}
        result = asyncio.run(executor._smart_dedup(config, df))

        assert result["success"] is True
        assert result["original_rows"] == 3
        assert result["deduped_rows"] == 2

    def test_smart_dedup_no_data(self, executor):
        """测试智能去重无数据"""
        import asyncio

        config = {}
        result = asyncio.run(executor._smart_dedup(config, None))

        assert result["success"] is False

    def test_extract_columns_fixed(self, executor, sample_excel_with_date):
        """测试提取固定4列"""
        import asyncio

        df = pd.read_excel(sample_excel_with_date)

        config = {}
        result = asyncio.run(executor._extract_columns(config, df))

        assert result["success"] is True
        extracted_df = result["data"]
        assert len(extracted_df.columns) == 4

    def test_extract_columns_no_data(self, executor):
        """测试提取列无数据"""
        import asyncio

        config = {}
        result = asyncio.run(executor._extract_columns(config, None))

        assert result["success"] is False
        assert "没有可处理" in result["message"]

    def test_export_excel_success(self, executor, temp_excel_dir):
        """测试导出Excel"""
        import asyncio

        df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        config = {"output_filename": "exported.xlsx"}

        result = asyncio.run(executor._export_excel(config, df))

        assert result["success"] is True
        assert "exported.xlsx" in result["file_path"]
        assert os.path.exists(result["file_path"])

    def test_export_excel_no_data(self, executor):
        """测试导出无数据"""
        import asyncio

        config = {"output_filename": "exported.xlsx"}
        result = asyncio.run(executor._export_excel(config, None))

        assert result["success"] is False
        assert "没有可导出" in result["message"]

    def test_unknown_step_type(self, executor):
        """测试未知步骤类型"""
        import asyncio

        config = {}
        result = asyncio.run(executor.execute_step("unknown_type", config))

        assert result["success"] is False
        assert "未知" in result["message"]


class TestSmartDedupLogic:
    """智能去重逻辑测试"""

    def test_keep_latest_date(self):
        """测试保留最新公告日"""
        df = pd.DataFrame({
            "证券代码": ["002128.SZ", "002128.SZ", "002128.SZ"],
            "最新公告日": ["2026-04-09", "2026-04-01", "2026-05-01"]
        })

        df["最新公告日"] = pd.to_datetime(df["最新公告日"])
        df_sorted = df.sort_values("最新公告日", ascending=False)
        df_deduped = df_sorted.drop_duplicates(subset=["证券代码"], keep="first")

        assert len(df_deduped) == 1
        assert df_deduped.iloc[0]["最新公告日"] == pd.Timestamp("2026-05-01")

    def test_multiple_stocks_dedup(self):
        """测试多只股票去重"""
        df = pd.DataFrame({
            "证券代码": ["002128.SZ", "002128.SZ", "600519.SH", "600519.SH"],
            "最新公告日": ["2026-04-01", "2026-04-09", "2026-04-01", "2026-04-05"]
        })

        df["最新公告日"] = pd.to_datetime(df["最新公告日"])
        df_sorted = df.sort_values("最新公告日", ascending=False)
        df_deduped = df_sorted.drop_duplicates(subset=["证券代码"], keep="first")

        assert len(df_deduped) == 2

        stock_codes = set(df_deduped["证券代码"])
        assert "002128.SZ" in stock_codes
        assert "600519.SH" in stock_codes

        sz_row = df_deduped[df_deduped["证券代码"] == "002128.SZ"].iloc[0]
        assert sz_row["最新公告日"] == pd.Timestamp("2026-04-09")

    def test_different_stocks_not_deduped(self):
        """测试不同股票不被去重"""
        df = pd.DataFrame({
            "证券代码": ["002128.SZ", "600519.SH", "000001.SZ"],
            "最新公告日": ["2026-04-01", "2026-04-01", "2026-04-01"]
        })

        df["最新公告日"] = pd.to_datetime(df["最新公告日"])
        df_sorted = df.sort_values("最新公告日", ascending=False)
        df_deduped = df_sorted.drop_duplicates(subset=["证券代码"], keep="first")

        assert len(df_deduped) == 3


class TestMergeExcelStartRow:
    """测试合并Excel时从序号=1开始读取"""

    TEST_DATE = "2026-04-13"

    @pytest.fixture
    def temp_dir(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.fixture
    def executor(self, temp_dir):
        return WorkflowExecutor(base_dir=temp_dir)

    def _get_upload_dir(self, temp_dir):
        """创建并返回日期子目录"""
        upload_dir = os.path.join(temp_dir, self.TEST_DATE)
        os.makedirs(upload_dir, exist_ok=True)
        return upload_dir

    def test_merge_skips_metadata_rows_before_seq1(self, executor, temp_dir):
        """序号=1不在第一行时，跳过前面的元数据行"""
        import asyncio
        upload_dir = self._get_upload_dir(temp_dir)

        # 模拟文件：前2行是元数据，序号=1在第3行
        df = pd.DataFrame({
            "序号": ["合计", "日期:2026-04-02", 1, 2],
            "证券代码": ["", "", "000001.SZ", "000002.SZ"],
            "证券简称": ["", "", "平安银行", "万科A"],
            "最新公告日": ["2026-04-02", "2026-04-02", "2025-12-23", "2026-01-05"]
        })
        df.to_excel(os.path.join(upload_dir, "file1.xlsx"), index=False)

        config = {"output_filename": "total_1.xlsx", "date_str": self.TEST_DATE}
        result = asyncio.run(executor._merge_excel(config, date_str=self.TEST_DATE))

        assert result["success"] is True
        # 应该只有2行数据（序号=1和序号=2），元数据行被跳过
        assert result["rows"] == 2

        output_df = pd.read_excel(result["file_path"])
        codes = output_df["证券代码"].astype(str).tolist()
        assert all(c != "" and c != "nan" for c in codes)

    def test_merge_seq1_at_first_row_works(self, executor, temp_dir):
        """序号=1在第一行时，正常读取所有数据"""
        import asyncio
        upload_dir = self._get_upload_dir(temp_dir)

        df = pd.DataFrame({
            "序号": [1, 2, 3],
            "证券代码": ["000001.SZ", "000002.SZ", "000003.SZ"],
            "证券简称": ["平安银行", "万科A", "国农科技"],
            "最新公告日": ["2026-04-02", "2026-01-05", "2025-12-23"]
        })
        df.to_excel(os.path.join(upload_dir, "file1.xlsx"), index=False)

        config = {"output_filename": "total_1.xlsx"}
        result = asyncio.run(executor._merge_excel(config, date_str=self.TEST_DATE))

        assert result["success"] is True
        assert result["rows"] == 3

    def test_merge_no_seq_column_reads_all(self, executor, temp_dir):
        """无序号列时，读取全部数据"""
        import asyncio
        upload_dir = self._get_upload_dir(temp_dir)

        df = pd.DataFrame({
            "证券代码": ["000001.SZ", "000002.SZ"],
            "证券简称": ["平安银行", "万科A"],
            "最新公告日": ["2026-04-02", "2026-01-05"]
        })
        df.to_excel(os.path.join(upload_dir, "file1.xlsx"), index=False)

        config = {"output_filename": "total_1.xlsx"}
        result = asyncio.run(executor._merge_excel(config, date_str=self.TEST_DATE))

        assert result["success"] is True
        assert result["rows"] == 2

    def test_merge_metadata_date_not_pollute_output(self, executor, temp_dir):
        """元数据行的日期不应污染最终输出的最新公告日"""
        import asyncio
        upload_dir = self._get_upload_dir(temp_dir)

        # 元数据行有一个"假"日期 2099-01-01
        df = pd.DataFrame({
            "序号": ["统计", 1, 2],
            "证券代码": ["汇总", "000001.SZ", "000001.SZ"],
            "证券简称": ["", "平安银行", "平安银行"],
            "最新公告日": ["2099-01-01", "2026-04-02", "2025-12-23"]
        })
        df.to_excel(os.path.join(upload_dir, "file1.xlsx"), index=False)

        config = {"output_filename": "total_1.xlsx"}
        result = asyncio.run(executor._merge_excel(config, date_str=self.TEST_DATE))

        assert result["success"] is True
        output_df = pd.read_excel(result["file_path"])
        # 确认元数据行的假日期 2099-01-01 不在输出中
        dates = output_df["最新公告日"].tolist()
        assert "2099-01-01" not in [str(d)[:10] for d in dates]

    def test_merge_latest_date_preserved_after_full_workflow(self, executor, temp_dir):
        """完整流程：合并+去重后，同一证券代码保留最新公告日"""
        import asyncio
        upload_dir = self._get_upload_dir(temp_dir)

        # 同一证券代码出现3次，不同日期
        df = pd.DataFrame({
            "序号": [1, 2, 3],
            "证券代码": ["000001.SZ", "000001.SZ", "000001.SZ"],
            "证券简称": ["平安银行", "平安银行", "平安银行"],
            "最新公告日": ["2026-01-05", "2026-04-02", "2025-12-23"]
        })
        df.to_excel(os.path.join(upload_dir, "file1.xlsx"), index=False)

        # Step 1: merge
        config = {"output_filename": "total_1.xlsx"}
        merge_result = asyncio.run(executor._merge_excel(config, date_str=self.TEST_DATE))
        assert merge_result["success"] is True

        # Step 2: smart_dedup
        merged_df = pd.read_excel(merge_result["file_path"])
        dedup_result = asyncio.run(executor._smart_dedup({}, merged_df, date_str=self.TEST_DATE))
        assert dedup_result["success"] is True
        assert dedup_result["deduped_rows"] == 1

        # 最新公告日应为最大日期 2026-04-02
        final_df = pd.read_excel(dedup_result["file_path"])
        assert final_df.iloc[0]["最新公告日"] == "2026-04-02"

    def test_merge_multirow_header(self, executor, temp_dir):
        """双行表头格式：第1行分组头、第2行实际列名，序号=1在第3行"""
        import asyncio
        from openpyxl import Workbook

        upload_dir = self._get_upload_dir(temp_dir)

        # 手动构造双行表头 Excel（模拟截图中的格式）
        wb = Workbook()
        ws = wb.active
        # Row 1: 分组表头
        ws["A1"] = "序号"
        ws["B1"] = "披露方"
        ws.merge_cells("B1:C1")
        ws["D1"] = ""
        ws["E1"] = ""
        # Row 2: 实际列名
        ws["A2"] = ""
        ws["B2"] = "证券代码"
        ws["C2"] = "证券简称"
        ws["D2"] = "最新公告日"
        ws["E2"] = "首次公告日"
        # Row 3: 数据
        ws["A3"] = 1
        ws["B3"] = "400272.NQ"
        ws["C3"] = "鹏博士3"
        ws["D3"] = "2026-04-13"
        ws["E3"] = "2025-12-31"
        # Row 4: 数据
        ws["A4"] = 2
        ws["B4"] = "600981.SH"
        ws["C4"] = "苏豪汇鸿"
        ws["D4"] = "2026-04-13"
        ws["E4"] = "2026-04-13"

        filepath = os.path.join(upload_dir, "multirow_header.xlsx")
        wb.save(filepath)

        config = {"output_filename": "total_1.xlsx"}
        result = asyncio.run(executor._merge_excel(config, date_str=self.TEST_DATE))

        assert result["success"] is True
        # 应该读到2行数据（序号1和序号2），不丢失
        assert result["rows"] == 2

        output_df = pd.read_excel(result["file_path"])
        # 列名应已正确重映射为实际列名
        assert "证券代码" in output_df.columns
        assert "证券简称" in output_df.columns
        assert "最新公告日" in output_df.columns
        # 数据内容正确
        codes = output_df["证券代码"].astype(str).tolist()
        assert "600981.SH" in codes

    def test_merge_multirow_header_with_duplicate_subcolumns(self, executor, temp_dir):
        """双行表头+重复子列名（如受让方/转让方各有'名称'列）不报错"""
        import asyncio
        from openpyxl import Workbook

        upload_dir = self._get_upload_dir(temp_dir)
        wb = Workbook()
        ws = wb.active
        # Row 1: 分组表头
        ws["A1"] = "序号"
        ws["B1"] = "披露方"
        ws.merge_cells("B1:C1")
        ws["D1"] = ""
        ws["E1"] = "受让方"
        ws["F1"] = "转让方"
        # Row 2: 子列名（名称重复）
        ws["A2"] = ""
        ws["B2"] = "证券代码"
        ws["C2"] = "证券简称"
        ws["D2"] = "最新公告日"
        ws["E2"] = "名称"
        ws["F2"] = "名称"
        # Row 3-4: 数据
        for row, data in enumerate(
            [(1, "000001.SZ", "平安银行", "2026-04-13", "受让A", "转让B"),
             (2, "600519.SH", "贵州茅台", "2026-04-13", "受让C", "转让D")],
            start=3
        ):
            for col, val in enumerate(data, start=1):
                ws.cell(row=row, column=col, value=val)

        wb.save(os.path.join(upload_dir, "dup_cols.xlsx"))

        config = {"output_filename": "total_1.xlsx"}
        result = asyncio.run(executor._merge_excel(config, date_str=self.TEST_DATE))

        assert result["success"] is True
        assert result["rows"] == 2
        output_df = pd.read_excel(result["file_path"])
        assert "证券代码" in output_df.columns
        assert "最新公告日" in output_df.columns
