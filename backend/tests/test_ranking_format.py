import os
import tempfile
import shutil
import pytest
from openpyxl import Workbook, load_workbook

from services.ranking_format import apply_ranking_format, detect_ranking_layout, LIGHT_RED, DEEP_RED, RANK_6_10_FILL


@pytest.fixture
def tmp_xlsx():
    d = tempfile.mkdtemp()
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _build_extended_sheet():
    """构造扩展布局的 sheet。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "0422"
    headers = ["板块名称", "年初涨跌幅", "B列的数值升序排序结果",
               "月初涨跌幅", "D列的数值升序排序结果",
               "今日涨跌幅", "迄今为止排进前5(次数)", "4月22日"]
    ws.append(headers)
    # 12 行数据：B排名和D排名分别填 1~12 用于测试 Top5 + 6-10 + 11+
    for i in range(1, 13):
        ws.append([f"板块{i}", f"{i*2}.0", i, f"{i}.5", i, 5.0-i*0.5, 1 if i<=5 else 0, i])
    return wb, ws


def test_detect_extended_layout():
    wb, ws = _build_extended_sheet()
    headers = [str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column+1)]
    layout = detect_ranking_layout(headers)
    assert layout["is_extended"] is True
    assert layout["ytd_rank_col"] == 3
    assert layout["mtd_rank_col"] == 5
    assert layout["top5_count_col"] == 7
    assert layout["gold_col"] == 7
    assert layout["date_start_col"] == 8


def test_detect_legacy_layout():
    wb = Workbook()
    ws = wb.active
    ws.append(["板块名称", "今日涨跌幅", "迄今为止排进前5(次数)", "4月22日"])
    headers = [str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column+1)]
    layout = detect_ranking_layout(headers)
    assert layout["is_extended"] is False
    assert layout["gold_col"] == 3
    assert layout["date_start_col"] == 4


def test_6_to_10_light_red_on_rank_cols(tmp_xlsx):
    """年初/月初两个排名列：1-5 深红，6-10 浅红，11+ 无色。"""
    wb, ws = _build_extended_sheet()
    apply_ranking_format(ws, prev_rank_by_sector=None)

    # 检查 C 列（第3列）= 年初排名
    for row in range(2, 14):
        rank = row - 1  # 行对应 rank 1..12
        cell = ws.cell(row=row, column=3)
        if rank <= 5:
            assert cell.fill.start_color.rgb.endswith("C00000"), f"rank {rank} 应深红，实际 {cell.fill.start_color.rgb}"
        elif 6 <= rank <= 10:
            assert cell.fill.start_color.rgb.endswith("E69138"), f"rank {rank} 应橙色，实际 {cell.fill.start_color.rgb}"
        else:
            # 11,12: 无染色或默认（openpyxl 默认是 '00000000'）
            assert not cell.fill.start_color.rgb.endswith("C00000")
            assert not cell.fill.start_color.rgb.endswith("E69138")

    # 检查 E 列（第5列）= 月初排名
    for row in range(2, 14):
        rank = row - 1
        cell = ws.cell(row=row, column=5)
        if rank <= 5:
            assert cell.fill.start_color.rgb.endswith("C00000"), f"E列 rank {rank} 应深红"
        elif 6 <= rank <= 10:
            assert cell.fill.start_color.rgb.endswith("E69138"), f"E列 rank {rank} 应橙色"


def test_date_col_top5_only_no_6_to_10_light(tmp_xlsx):
    """日期列只有 Top5 深红，不触发 6-10 橙色（6-10 橙色只在 YTD/MTD 排名列）"""
    wb, ws = _build_extended_sheet()
    apply_ranking_format(ws, prev_rank_by_sector=None)

    # 日期列 = col 8，值 1..12
    for row in range(2, 14):
        rank = row - 1
        cell = ws.cell(row=row, column=8)
        if rank <= 5:
            assert cell.fill.start_color.rgb.endswith("C00000")
        else:
            # 6+ 不应是橙色（只有 YTD/MTD 排名列才有 6-10 橙色）
            assert not cell.fill.start_color.rgb.endswith("E69138")


def test_legacy_layout_no_6_to_10(tmp_xlsx):
    """老布局：没有 YTD/MTD 排名列 → 不应有 6-10 橙色染色行为"""
    wb = Workbook()
    ws = wb.active
    ws.append(["板块名称", "今日涨跌幅", "迄今为止排进前5(次数)", "4月22日", "4月21日"])
    # 5 行数据
    for i in range(1, 9):
        ws.append([f"板块{i}", 1.0, 1 if i<=5 else 0, i, i])

    apply_ranking_format(ws, prev_rank_by_sector=None)

    # 日期列（4/5）中 rank 6-8：不应染橙色（除非"上日列"有 prev_rank 提升）
    for row in range(7, 10):  # rank 6,7,8
        cell = ws.cell(row=row, column=4)
        assert not cell.fill.start_color.rgb.endswith("E69138"), f"rank {row-1} 老布局日期列不应橙色"
