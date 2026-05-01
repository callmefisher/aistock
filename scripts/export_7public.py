import json
import sys
import pymysql
import pandas as pd
from collections import defaultdict
from datetime import datetime

DB_HOST = "127.0.0.1"
DB_PORT = 3306
DB_USER = "stock_user"
DB_PASS = "stock_password"
DB_NAME = "stock_pool"

START_DATE = "2026-03-18"
END_DATE = "2026-04-29"
OUTPUT_FILE = "/Users/xiayanji/qbox/aistock/data/excel/7public.xlsx"

HIGH_DATE_COL = f"{START_DATE}至{END_DATE}期间百日新高的日期"
HIGH_COUNT_COL = f"{START_DATE}至{END_DATE}期间百日新高次数"


def main():
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER,
        password=DB_PASS, database=DB_NAME,
        charset="utf8mb4"
    )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT date_str, data FROM stock_pools "
                "WHERE is_active=1 AND date_str >= %s AND date_str <= %s "
                "ORDER BY date_str",
                (START_DATE, END_DATE)
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    print(f"查询到 {len(rows)} 条选股池记录")

    all_records = []
    for date_str, raw_data in rows:
        if not raw_data:
            continue
        records = raw_data if isinstance(raw_data, list) else json.loads(raw_data)
        for rec in records:
            rec["_data_date"] = date_str
        all_records.extend(records)

    print(f"共解析 {len(all_records)} 条股票记录")

    code_high_dates = defaultdict(list)
    code_records = defaultdict(list)

    for rec in all_records:
        code = str(rec.get("证券代码", "")).strip()
        if not code:
            continue
        code_records[code].append(rec)

        brixin_val = str(rec.get("百日新高", "")).strip()
        if brixin_val:
            code_high_dates[code].append(rec["_data_date"])

    merged = []
    for code, recs in code_records.items():
        recs_sorted = sorted(
            recs,
            key=lambda r: (
                r.get("最新公告日", "") or "",
                r.get("_data_date", "") or ""
            ),
            reverse=True
        )
        base = recs_sorted[0].copy()
        base.pop("_data_date", None)

        behaviors = []
        seen_behaviors = set()
        for r in recs:
            beh = str(r.get("资本运作行为", "")).strip()
            if not beh:
                continue
            for part in beh.split("、"):
                p = part.strip()
                if p and p not in seen_behaviors:
                    seen_behaviors.add(p)
                    behaviors.append(p)
        base["资本运作行为"] = "、".join(behaviors) if behaviors else str(base.get("资本运作行为", ""))

        high_dates = sorted(set(code_high_dates.get(code, [])))
        base[HIGH_DATE_COL] = ",".join(high_dates)
        base[HIGH_COUNT_COL] = len(high_dates)

        merged.append(base)

    if not merged:
        print("无数据可导出")
        sys.exit(1)

    print(f"合并后 {len(merged)} 条记录")

    before_st = len(merged)
    merged = [r for r in merged if "ST" not in str(r.get("证券简称", "")).upper()]
    print(f"排除含ST的行: {before_st - len(merged)} 条, 剩余 {len(merged)} 条")

    col_order = [
        "证券代码", "证券简称", "最新公告日",
        "百日新高", "站上20日线", "所属板块", "国央企",
        "资本运作行为",
        HIGH_COUNT_COL, HIGH_DATE_COL
    ]

    existing_cols = set(merged[0].keys())
    final_cols = [c for c in col_order if c in existing_cols]
    for c in existing_cols:
        if c not in final_cols:
            final_cols.append(c)

    df = pd.DataFrame(merged, columns=final_cols)
    df = df.sort_values(by="证券代码").reset_index(drop=True)

    df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    green_fill = PatternFill(fill_type="solid", start_color="CAEED3", end_color="CAEED3")
    headers = [c.value for c in ws[1]]

    highlight_cols = {"证券代码", "证券简称", "最新公告日"}
    highlight_col_indices = [i + 1 for i, h in enumerate(headers) if h in highlight_cols]
    high_count_col_idx = None
    for i, h in enumerate(headers):
        if h == HIGH_COUNT_COL:
            high_count_col_idx = i + 1
            break

    for idx, h in enumerate(headers, start=1):
        if not isinstance(h, str):
            ws.column_dimensions[get_column_letter(idx)].width = 18
            continue
        if h == "资本运作行为":
            ws.column_dimensions[get_column_letter(idx)].width = 80
        elif HIGH_DATE_COL in h:
            ws.column_dimensions[get_column_letter(idx)].width = 65
        elif HIGH_COUNT_COL in h:
            ws.column_dimensions[get_column_letter(idx)].width = 22
        elif h in ("证券代码", "证券简称"):
            ws.column_dimensions[get_column_letter(idx)].width = 16
        elif h == "最新公告日":
            ws.column_dimensions[get_column_letter(idx)].width = 16
        elif h == "所属板块":
            ws.column_dimensions[get_column_letter(idx)].width = 16
        elif h == "国央企":
            ws.column_dimensions[get_column_letter(idx)].width = 16
        elif h in ("百日新高", "站上20日线"):
            ws.column_dimensions[get_column_letter(idx)].width = 16
        else:
            ws.column_dimensions[get_column_letter(idx)].width = 20

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align

    if high_count_col_idx is not None:
        for r_idx in range(2, ws.max_row + 1):
            count_val = ws.cell(row=r_idx, column=high_count_col_idx).value
            if count_val is not None and int(count_val) == 1:
                for c_idx in highlight_col_indices:
                    ws.cell(row=r_idx, column=c_idx).fill = green_fill

    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_FILE)

    print(f"导出完成: {OUTPUT_FILE}")
    print(f"总行数: {len(df)}")
    high_count_1 = df[df[HIGH_COUNT_COL] == 1].shape[0]
    print(f"百日新高次数=1 的证券数: {high_count_1}")


if __name__ == "__main__":
    main()
